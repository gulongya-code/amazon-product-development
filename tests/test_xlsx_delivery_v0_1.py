from __future__ import annotations

import ast
from dataclasses import FrozenInstanceError
from io import BytesIO
import json
import os
from pathlib import Path
import subprocess
import sys
import tempfile
import unittest
from zipfile import ZipFile

from openpyxl import load_workbook

from amazon_product_intelligence.contracts import deterministic_id
import amazon_product_intelligence.xlsx_delivery as xlsx_delivery
from amazon_product_intelligence.xlsx_delivery import (
    XLSX_DELIVERY_RULESET_VERSION,
    XlsxDeliveryBuilderV0_1,
    XlsxDeliveryRequest,
    XlsxDeliverySerializationError,
    XlsxDeliverySnapshotV0_1,
    XlsxDeliveryValidationError,
)
from tests.test_operator_export_v0_1 import build_export


EXPECTED_API = {
    "XLSX_DELIVERY_RULESET_VERSION",
    "XlsxDeliveryRequest",
    "XlsxDeliverySnapshotV0_1",
    "XlsxDeliveryBuilderV0_1",
    "XlsxDeliveryError",
    "XlsxDeliveryValidationError",
    "XlsxDeliverySerializationError",
    "WorkbookStyleDefinition",
    "WorksheetRenderDefinition",
    "CellRenderRecord",
    "WorkbookDeliveryRecord",
    "DeliveryCoverageSummary",
    "DeliveryLineageReference",
    "DeliveryDiagnostic",
}
EXPECTED_SHEETS = (
    "01_产品数据",
    "02_关键词需求",
    "03_竞争证据",
    "04_机会分析",
    "05_建议与复核",
)
EXPECTED_COLUMNS = {
    "01_产品数据": (
        "ASIN", "Marketplace", "Title", "Product Facts", "Metrics",
        "Variation", "Reviews", "Quality Indicators", "Source Reference",
    ),
    "02_关键词需求": (
        "Keyword", "Metrics", "Query Status", "Related Products", "Channels",
        "Providers", "Limitations",
    ),
    "03_竞争证据": (
        "Product Endpoint", "Relationship Evidence", "Relationship Type",
        "Channel", "Provider", "Evidence Count", "Variation Evidence",
        "Limitations",
    ),
    "04_机会分析": (
        "Product", "Signals", "Missing Evidence", "Risk Evidence",
        "Score References", "Explanation References",
    ),
    "05_建议与复核": (
        "Recommendation Type", "Rule Reference", "Explanation",
        "Evidence References", "Limitations",
    ),
}


def build_delivery():
    bundles, output, export_request, export_snapshot = build_export()
    request = XlsxDeliveryRequest(
        operator_export_snapshot=export_snapshot.to_dict()
    )
    snapshot = XlsxDeliveryBuilderV0_1().build(request)
    return bundles, output, export_request, export_snapshot, request, snapshot


def reverse_mapping_keys(value):
    if isinstance(value, dict):
        return {
            key: reverse_mapping_keys(value[key])
            for key in reversed(tuple(value))
        }
    if isinstance(value, list):
        return [reverse_mapping_keys(item) for item in value]
    return value


def reidentify_export_row(payload, table_key, column, value):
    table = next(
        item for item in payload["table_definitions"]
        if item["table_key"] == table_key
    )
    row = next(item for item in payload["rows"] if item["table_id"] == table["table_id"])
    old_row_id = row["export_row_id"]
    row["values"][column] = value
    row_content = dict(row)
    row_content.pop("export_row_id")
    row_content.pop("lineage_reference_ids")
    new_row_id = deterministic_id("operator-export-row", row_content)
    row["export_row_id"] = new_row_id
    changed_lineages = {}
    for lineage in payload["lineage_index"]:
        if lineage["export_row_id"] != old_row_id:
            continue
        old_lineage_id = lineage["export_lineage_id"]
        lineage["export_row_id"] = new_row_id
        lineage_content = dict(lineage)
        lineage_content.pop("export_lineage_id")
        new_lineage_id = deterministic_id(
            "operator-export-lineage", lineage_content
        )
        lineage["export_lineage_id"] = new_lineage_id
        changed_lineages[old_lineage_id] = new_lineage_id
    row["lineage_reference_ids"] = sorted(
        changed_lineages.get(item, item) for item in row["lineage_reference_ids"]
    )
    for sheet in payload["sheet_definitions"]:
        sheet["lineage_reference_ids"] = sorted(
            changed_lineages.get(item, item)
            for item in sheet["lineage_reference_ids"]
        )
    payload["rows"] = sorted(payload["rows"], key=lambda item: item["export_row_id"])
    payload["lineage_index"] = sorted(
        payload["lineage_index"], key=lambda item: item["export_lineage_id"]
    )
    snapshot_content = dict(payload)
    snapshot_content.pop("snapshot_id")
    payload["snapshot_id"] = deterministic_id(
        "operator-export-snapshot", snapshot_content
    )
    return new_row_id


def replace_delivery_lineage(payload, index, **changes):
    lineage = payload["lineage_index"][index]
    old_id = lineage["delivery_lineage_id"]
    lineage.update(changes)
    content = dict(lineage)
    content.pop("delivery_lineage_id")
    new_id = deterministic_id("xlsx-lineage", content)
    lineage["delivery_lineage_id"] = new_id
    for cell in payload["cells"]:
        if old_id in cell["lineage_reference_ids"]:
            cell["lineage_reference_ids"] = sorted(
                new_id if item == old_id else item
                for item in cell["lineage_reference_ids"]
            )
    for worksheet in payload["worksheet_definitions"]:
        if old_id in worksheet["lineage_reference_ids"]:
            worksheet["lineage_reference_ids"] = sorted(
                new_id if item == old_id else item
                for item in worksheet["lineage_reference_ids"]
            )


def unescape_formula(value):
    if (
        type(value) is str
        and value.startswith("'")
        and value[1:].lstrip().startswith(("=", "+", "-", "@"))
    ):
        return value[1:]
    return value


class XlsxDeliveryV01Tests(unittest.TestCase):
    @classmethod
    def setUpClass(cls):
        (
            cls.bundles,
            cls.output_snapshot,
            cls.export_request,
            cls.export_snapshot,
            cls.request,
            cls.snapshot,
        ) = build_delivery()

    def workbook(self):
        return load_workbook(BytesIO(self.snapshot.to_xlsx_bytes()), data_only=False)

    def test_public_api_is_exact(self):
        self.assertEqual(set(xlsx_delivery.__all__), EXPECTED_API)
        self.assertEqual(len(xlsx_delivery.__all__), 14)
        for name in EXPECTED_API:
            self.assertIsNotNone(getattr(xlsx_delivery, name))

    def test_ruleset_identity_filename_and_media_type(self):
        self.assertEqual(XLSX_DELIVERY_RULESET_VERSION, "xlsx-delivery-v0.1")
        self.assertEqual(self.snapshot.ruleset_version, XLSX_DELIVERY_RULESET_VERSION)
        self.assertEqual(
            self.snapshot.workbook.filename, "amazon_product_analysis.xlsx"
        )
        self.assertEqual(
            self.snapshot.workbook.media_type,
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        identity = self.snapshot.to_dict()
        snapshot_id = identity.pop("snapshot_id")
        self.assertEqual(
            snapshot_id, deterministic_id("xlsx-delivery-snapshot", identity)
        )
        self.assertEqual(
            self.snapshot.snapshot_id,
            "xlsx-delivery-snapshot:2a316fc50778d888a450753a4ab1e1fa26567219246e8f8e5f1a467746e0eab4",
        )
        self.assertEqual(
            self.snapshot.workbook.content_sha256,
            "5003d07e7c6172291338c01a0a71a80ce9fef36949cfc5040f4793eeb9659657",
        )
        self.assertEqual(self.snapshot.workbook.size_bytes, 52267)

    def test_real_xlsx_is_an_open_packaging_convention_zip(self):
        content = self.snapshot.to_xlsx_bytes()
        self.assertTrue(content.startswith(b"PK"))
        with ZipFile(BytesIO(content)) as package:
            names = set(package.namelist())
        self.assertIn("[Content_Types].xml", names)
        self.assertIn("xl/workbook.xml", names)
        self.assertIn("xl/styles.xml", names)
        self.assertTrue(any(name.startswith("xl/worksheets/sheet") for name in names))

    def test_workbook_opens_with_exact_sheet_order(self):
        workbook = self.workbook()
        self.assertEqual(tuple(workbook.sheetnames), EXPECTED_SHEETS)

    def test_title_and_header_rows_are_exact(self):
        workbook = self.workbook()
        expected_titles = ("产品数据", "关键词需求", "竞争证据", "机会分析", "建议与复核")
        for sheet_name, title in zip(EXPECTED_SHEETS, expected_titles, strict=True):
            worksheet = workbook[sheet_name]
            self.assertEqual(worksheet["A1"].value, title)
            self.assertEqual(
                tuple(
                    worksheet.cell(row=2, column=index).value
                    for index in range(1, len(EXPECTED_COLUMNS[sheet_name]) + 1)
                ),
                EXPECTED_COLUMNS[sheet_name],
            )

    def test_freeze_panes_filters_and_column_widths(self):
        workbook = self.workbook()
        definitions = {
            item.sheet_name: item for item in self.snapshot.worksheet_definitions
        }
        for sheet_name in EXPECTED_SHEETS:
            worksheet = workbook[sheet_name]
            definition = definitions[sheet_name]
            self.assertEqual(worksheet.freeze_panes, "A2")
            self.assertEqual(worksheet.auto_filter.ref, definition.auto_filter_range)
            for column_index, expected_width in enumerate(
                definition.column_widths, start=1
            ):
                actual = worksheet.column_dimensions[
                    worksheet.cell(row=2, column=column_index).column_letter
                ].width
                self.assertEqual(actual, expected_width)

    def test_title_header_and_body_formatting_is_applied(self):
        workbook = self.workbook()
        worksheet = workbook["01_产品数据"]
        self.assertTrue(worksheet["A1"].font.bold)
        self.assertEqual(worksheet["A1"].fill.fill_type, "solid")
        self.assertTrue(worksheet["A2"].font.bold)
        self.assertEqual(worksheet["A2"].alignment.wrap_text, True)
        self.assertEqual(worksheet["A3"].alignment.vertical, "top")
        self.assertEqual(worksheet["A3"].border.left.style, "thin")

    def test_coverage_accounts_for_continuation_rows_without_data_loss(self):
        self.assertEqual(self.snapshot.coverage.worksheet_count, 5)
        self.assertEqual(self.snapshot.coverage.source_export_row_count, 17)
        self.assertEqual(self.snapshot.coverage.rendered_row_count, 22)
        self.assertEqual(self.snapshot.coverage.cell_count, 152)
        self.assertEqual(self.snapshot.coverage.lineage_reference_count, 332)
        self.assertEqual(self.snapshot.coverage.diagnostic_count, 1)

    def test_product_and_keyword_cells_reproduce_export_values(self):
        export_rows = {
            item.export_row_id: item for item in self.export_snapshot.rows
        }
        for sheet_name in ("01_产品数据", "02_关键词需求"):
            definition = next(
                item for item in self.snapshot.worksheet_definitions
                if item.sheet_name == sheet_name
            )
            source_row = export_rows[definition.source_export_row_ids[0]]
            cells = sorted(
                (
                    item for item in self.snapshot.cells
                    if item.source_export_row_id == source_row.export_row_id
                    and item.chunk_index == 0
                ),
                key=lambda item: item.excel_column,
            )
            self.assertEqual(
                tuple(item.value for item in cells),
                tuple(source_row.values[column] for column in definition.columns),
            )

    def test_competition_relationship_is_presentationally_split(self):
        definition = next(
            item for item in self.snapshot.worksheet_definitions
            if item.sheet_name == "03_竞争证据"
        )
        source_row_id = definition.source_export_row_ids[0]
        source_row = next(
            item for item in self.export_snapshot.rows
            if item.export_row_id == source_row_id
        )
        relationship = json.loads(source_row.values["Keyword Relationship"])
        cells = {
            item.column_name: item.value
            for item in self.snapshot.cells
            if item.source_export_row_id == source_row_id
        }
        self.assertEqual(
            json.loads(cells["Relationship Evidence"]),
            relationship["relationship"],
        )
        self.assertEqual(cells["Relationship Type"], relationship["relationship_type"])
        self.assertNotIn("Competitor Ranking", definition.columns)
        self.assertNotIn("Competitor Score", definition.columns)

    def test_opportunity_long_values_are_reconstructable_from_chunks(self):
        definition = next(
            item for item in self.snapshot.worksheet_definitions
            if item.sheet_name == "04_机会分析"
        )
        self.assertGreater(len(definition.delivery_row_ids), 1)
        source_row_id = definition.source_export_row_ids[0]
        source_row = next(
            item for item in self.export_snapshot.rows
            if item.export_row_id == source_row_id
        )
        for column in definition.columns:
            values = [
                item.value
                for item in sorted(
                    (
                        item for item in self.snapshot.cells
                        if item.source_export_row_id == source_row_id
                        and item.column_name == column
                    ),
                    key=lambda item: item.chunk_index,
                )
                if item.value is not None
            ]
            source_value = source_row.values[column]
            if type(source_value) is str:
                reconstructed = "".join(unescape_formula(item) for item in values)
                self.assertEqual(reconstructed, source_value)
            else:
                self.assertEqual(values, [source_value])
        self.assertLessEqual(
            max(
                len(item.value) for item in self.snapshot.cells
                if type(item.value) is str
            ),
            30001,
        )

    def test_recommendations_are_existing_export_values(self):
        definition = next(
            item for item in self.snapshot.worksheet_definitions
            if item.sheet_name == "05_建议与复核"
        )
        export_rows = {
            item.export_row_id: item for item in self.export_snapshot.rows
        }
        for row_id in definition.source_export_row_ids:
            source = export_rows[row_id]
            cells = {
                item.column_name: item.value
                for item in self.snapshot.cells
                if item.source_export_row_id == row_id
            }
            self.assertEqual(
                cells["Recommendation Type"],
                source.values["Recommendation Type"],
            )

    def test_formula_injection_is_escaped_and_not_written_as_formula(self):
        payload = self.export_snapshot.to_dict()
        row_id = reidentify_export_row(
            payload,
            "recommendation",
            "Recommendation Type",
            "=HYPERLINK(\"https://invalid.example\",\"x\")",
        )
        delivered = XlsxDeliveryBuilderV0_1().build(
            XlsxDeliveryRequest(operator_export_snapshot=payload)
        )
        cell = next(
            item for item in delivered.cells
            if item.source_export_row_id == row_id
            and item.column_name == "Recommendation Type"
        )
        self.assertTrue(cell.value.startswith("'="))
        workbook = load_workbook(BytesIO(delivered.to_xlsx_bytes()), data_only=False)
        worksheet = workbook["05_建议与复核"]
        written = worksheet[cell.coordinate]
        self.assertEqual(written.value, cell.value)
        self.assertNotEqual(written.data_type, "f")

    def test_workbook_contains_no_formula_cells(self):
        workbook = self.workbook()
        formula_cells = [
            cell.coordinate
            for worksheet in workbook.worksheets
            for row in worksheet.iter_rows()
            for cell in row
            if cell.data_type == "f"
        ]
        self.assertEqual(formula_cells, [])

    def test_hidden_credentials_in_json_text_are_rejected(self):
        payload = self.export_snapshot.to_dict()
        reidentify_export_row(
            payload,
            "recommendation",
            "Explanation",
            '{"access_token":"must-not-deliver"}',
        )
        with self.assertRaises(XlsxDeliveryValidationError):
            XlsxDeliveryRequest(operator_export_snapshot=payload)

    def test_real_file_can_be_created_once_with_fixed_name(self):
        with tempfile.TemporaryDirectory() as directory:
            target = Path(directory) / "amazon_product_analysis.xlsx"
            actual = self.snapshot.write_xlsx(target)
            self.assertEqual(actual, target)
            self.assertEqual(target.read_bytes(), self.snapshot.to_xlsx_bytes())
            workbook = load_workbook(target, read_only=True, data_only=False)
            self.assertEqual(tuple(workbook.sheetnames), EXPECTED_SHEETS)
            workbook.close()
            with self.assertRaises(XlsxDeliveryValidationError):
                self.snapshot.write_xlsx(target)
        with self.assertRaises(XlsxDeliveryValidationError):
            self.snapshot.write_xlsx("wrong-name.xlsx")

    def test_zip_entries_and_core_properties_have_fixed_timestamps(self):
        with ZipFile(BytesIO(self.snapshot.to_xlsx_bytes())) as package:
            self.assertEqual(
                {item.date_time for item in package.infolist()},
                {(1980, 1, 1, 0, 0, 0)},
            )
            core = package.read("docProps/core.xml").decode("utf-8")
        self.assertIn("2000-01-01T00:00:00Z", core)
        self.assertNotIn("2026-", core)

    def test_same_process_and_mapping_reorder_are_deterministic(self):
        rebuilt = XlsxDeliveryBuilderV0_1().build(self.request)
        reordered = XlsxDeliveryBuilderV0_1().build(
            XlsxDeliveryRequest(
                operator_export_snapshot=reverse_mapping_keys(
                    self.export_snapshot.to_dict()
                )
            )
        )
        for candidate in (rebuilt, reordered):
            self.assertEqual(candidate.snapshot_id, self.snapshot.snapshot_id)
            self.assertEqual(
                candidate.workbook.content_sha256,
                self.snapshot.workbook.content_sha256,
            )
            self.assertEqual(candidate.to_xlsx_bytes(), self.snapshot.to_xlsx_bytes())

    def test_fresh_process_is_deterministic(self):
        script = (
            "from tests.test_operator_export_v0_1 import build_export;"
            "from amazon_product_intelligence.xlsx_delivery import "
            "XlsxDeliveryRequest,XlsxDeliveryBuilderV0_1;"
            "e=build_export()[3];"
            "s=XlsxDeliveryBuilderV0_1().build("
            "XlsxDeliveryRequest(operator_export_snapshot=e.to_dict()));"
            "print(s.snapshot_id+'|'+s.workbook.content_sha256)"
        )
        environment = dict(os.environ)
        environment["PYTHONPATH"] = "src"
        actual = subprocess.check_output(
            [sys.executable, "-c", script],
            cwd=Path(__file__).resolve().parents[1],
            env=environment,
            text=True,
            encoding="utf-8",
        ).strip()
        self.assertEqual(
            actual,
            f"{self.snapshot.snapshot_id}|{self.snapshot.workbook.content_sha256}",
        )

    def test_every_export_lineage_is_delivered_once_with_full_chain(self):
        source = {
            item.export_lineage_id: item
            for item in self.export_snapshot.lineage_index
        }
        delivered = {
            item.source_export_lineage_id: item
            for item in self.snapshot.lineage_index
        }
        self.assertEqual(set(delivered), set(source))
        self.assertEqual(len(delivered), 332)
        for lineage_id, item in delivered.items():
            expected = source[lineage_id]
            self.assertEqual(item.source_export_row_id, expected.export_row_id)
            self.assertEqual(item.source_lineage_id, expected.source_lineage_id)
            self.assertEqual(
                item.canonical_reference_id, expected.canonical_reference_id
            )
            self.assertEqual(item.raw_evidence_id, expected.raw_evidence_id)

    def test_lineage_validates_against_operator_export_snapshot(self):
        self.assertIs(
            self.snapshot.validate_against_export_snapshot(
                self.export_snapshot.to_dict()
            ),
            self.snapshot,
        )
        payload = self.export_snapshot.to_dict()
        payload["snapshot_id"] = "operator-export-snapshot:missing"
        with self.assertRaises(XlsxDeliveryValidationError):
            self.snapshot.validate_against_export_snapshot(payload)

    def test_orphan_lineage_is_rejected(self):
        payload = self.snapshot.to_dict()
        target = payload["lineage_index"][0]["delivery_lineage_id"]
        for cell in payload["cells"]:
            if target in cell["lineage_reference_ids"] and len(cell["lineage_reference_ids"]) > 1:
                cell["lineage_reference_ids"].remove(target)
        with self.assertRaises(XlsxDeliveryValidationError):
            XlsxDeliverySnapshotV0_1.from_dict(payload)

    def test_wrong_worksheet_lineage_is_rejected(self):
        payload = self.snapshot.to_dict()
        current = payload["lineage_index"][0]["worksheet_id"]
        wrong = next(
            item["worksheet_id"] for item in payload["worksheet_definitions"]
            if item["worksheet_id"] != current
        )
        replace_delivery_lineage(payload, 0, worksheet_id=wrong)
        with self.assertRaises(XlsxDeliveryValidationError):
            XlsxDeliverySnapshotV0_1.from_dict(payload)

    def test_missing_export_row_is_rejected(self):
        payload = self.snapshot.to_dict()
        replace_delivery_lineage(
            payload, 0, source_export_row_id="operator-export-row:missing"
        )
        with self.assertRaises(XlsxDeliveryValidationError):
            XlsxDeliverySnapshotV0_1.from_dict(payload)

    def test_fingerprint_mismatch_is_rejected(self):
        payload = self.snapshot.to_dict()
        replace_delivery_lineage(
            payload, 0, source_bundle_fingerprints=["0" * 64]
        )
        with self.assertRaises(XlsxDeliveryValidationError):
            XlsxDeliverySnapshotV0_1.from_dict(payload)

    def test_cell_identity_mismatch_is_rejected(self):
        payload = self.snapshot.to_dict()
        payload["cells"][0]["value"] = "tampered"
        with self.assertRaises(XlsxDeliveryValidationError):
            XlsxDeliverySnapshotV0_1.from_dict(payload)

    def test_strict_serialization_round_trip_and_unknown_fields(self):
        restored = XlsxDeliverySnapshotV0_1.from_dict(self.snapshot.to_dict())
        self.assertEqual(restored.to_dict(), self.snapshot.to_dict())
        self.assertEqual(restored.to_xlsx_bytes(), self.snapshot.to_xlsx_bytes())
        payload = self.snapshot.to_dict()
        payload["unknown"] = True
        with self.assertRaises(XlsxDeliverySerializationError):
            XlsxDeliverySnapshotV0_1.from_dict(payload)

    def test_request_and_snapshot_are_deeply_immutable(self):
        with self.assertRaises(TypeError):
            self.request.operator_export_snapshot["snapshot_id"] = "changed"
        nested = self.request.operator_export_snapshot["rows"][0]["values"]
        with self.assertRaises(TypeError):
            nested[next(iter(nested))] = "changed"
        with self.assertRaises(TypeError):
            self.snapshot.workbook.metadata["format"] = "changed"
        with self.assertRaises(FrozenInstanceError):
            self.snapshot.snapshot_id = "changed"

    def test_builder_does_not_mutate_export_snapshot(self):
        payload = self.export_snapshot.to_dict()
        before = json.dumps(payload, sort_keys=True, ensure_ascii=False)
        request = XlsxDeliveryRequest(operator_export_snapshot=payload)
        XlsxDeliveryBuilderV0_1().build(request)
        self.assertEqual(
            json.dumps(payload, sort_keys=True, ensure_ascii=False), before
        )

    def test_builder_rejects_wrong_request_type(self):
        with self.assertRaises(XlsxDeliveryValidationError):
            XlsxDeliveryBuilderV0_1().build(self.export_snapshot)

    def test_production_dependency_and_safety_boundary(self):
        root = Path(__file__).resolve().parents[1]
        production = root / "src" / "amazon_product_intelligence" / "xlsx_delivery"
        allowed_stdlib = set(sys.stdlib_module_names) | {"__future__"}
        third_party = set()
        forbidden_parts = {
            "adapters", "competition_intelligence", "conflict_resolution",
            "decision_framework", "demand_intelligence", "evidence_evaluation",
            "evidence_policy", "operator_export", "operator_output",
            "opportunity_intelligence", "opportunity_scoring",
            "product_intelligence", "recommendation_framework",
        }
        forbidden_calls = {"eval", "exec", "hash", "repr"}
        for path in sorted(production.glob("*.py")):
            tree = ast.parse(path.read_text(encoding="utf-8"), filename=str(path))
            for node in ast.walk(tree):
                candidates = []
                if isinstance(node, ast.Import):
                    candidates = [alias.name for alias in node.names]
                elif isinstance(node, ast.ImportFrom) and node.level == 0:
                    candidates = [node.module]
                for name in candidates:
                    self.assertTrue(forbidden_parts.isdisjoint(name.split(".")))
                    root_name = name.split(".")[0]
                    if root_name not in allowed_stdlib:
                        third_party.add(root_name)
                if isinstance(node, ast.Call) and isinstance(node.func, ast.Name):
                    self.assertNotIn(node.func.id, forbidden_calls)
        self.assertEqual(third_party, {"openpyxl"})


if __name__ == "__main__":
    unittest.main()
