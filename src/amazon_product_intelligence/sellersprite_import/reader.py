"""Bounded, deterministic CSV/XLSX reader for local SellerSprite exports."""

from __future__ import annotations

import csv
from dataclasses import dataclass
from hashlib import sha256
from io import BytesIO, StringIO
from pathlib import Path
from typing import Any, Iterable

from openpyxl import load_workbook

from .errors import SellerSpriteImportError
from .schema_v1 import (
    CONTRACT_HEADERS,
    MAX_HEADER_SCAN_ROWS,
    MAX_LISTING_ROWS,
    OUT_OF_SCOPE_HEADERS,
    PREFERRED_RAW_SHEET,
    is_header_candidate,
    normalize_header,
)


@dataclass(frozen=True, slots=True, kw_only=True)
class SourceRow:
    row_number: int
    values: dict[str, Any]
    malformed: bool = False


@dataclass(frozen=True, slots=True, kw_only=True)
class RawImportTable:
    source_type: str
    source_basename: str
    source_file_sha256: str
    source_sheet: str | None
    header_row: int
    mapped_headers: tuple[str, ...]
    unmapped_headers: tuple[str, ...]
    out_of_scope_headers: tuple[str, ...]
    rows: tuple[SourceRow, ...]


def _header_mapping(values: tuple[object, ...]) -> tuple[dict[int, str], tuple[str, ...], tuple[str, ...]]:
    mapping: dict[int, str] = {}
    seen: set[str] = set()
    unknown: set[str] = set()
    out_of_scope: set[str] = set()
    for index, raw in enumerate(values):
        header = normalize_header(raw)
        if header is None:
            continue
        if header in CONTRACT_HEADERS:
            if header in seen:
                raise SellerSpriteImportError(
                    "DUPLICATE_MAPPED_HEADER",
                    f"header row maps more than one column to {header!r}",
                )
            seen.add(header)
            if header in OUT_OF_SCOPE_HEADERS:
                out_of_scope.add(header)
            else:
                mapping[index] = header
        else:
            unknown.add(header)
    return mapping, tuple(sorted(unknown)), tuple(sorted(out_of_scope))


def _materialize_rows(
    rows: Iterable[tuple[int, tuple[object, ...]]],
    mapping: dict[int, str],
    header_width: int,
) -> tuple[SourceRow, ...]:
    result: list[SourceRow] = []
    for row_number, row in rows:
        if not any(value is not None and (not isinstance(value, str) or value.strip()) for value in row):
            continue
        if len(result) >= MAX_LISTING_ROWS:
            raise SellerSpriteImportError(
                "ROW_LIMIT_EXCEEDED",
                f"source contains more than {MAX_LISTING_ROWS} nonblank listing rows",
            )
        malformed = any(
            value is not None and (not isinstance(value, str) or value.strip())
            for value in row[header_width:]
        )
        result.append(
            SourceRow(
                row_number=row_number,
                values={header: row[index] if index < len(row) else None for index, header in mapping.items()},
                malformed=malformed,
            )
        )
    return tuple(result)


def _read_csv(payload: bytes, basename: str, digest: str) -> RawImportTable:
    try:
        text = payload.decode("utf-8-sig", errors="strict")
    except UnicodeDecodeError as exc:
        raise SellerSpriteImportError("UNSUPPORTED_CSV_ENCODING", "CSV must be UTF-8 or UTF-8-SIG") from exc
    parsed = [tuple(row) for row in csv.reader(StringIO(text, newline=""), dialect="excel")]
    candidates = [
        index for index, row in enumerate(parsed[:MAX_HEADER_SCAN_ROWS], 1) if is_header_candidate(row)
    ]
    if not candidates:
        raise SellerSpriteImportError("HEADER_SCHEMA_MISMATCH", "no bounded exact-contract header candidate found")
    if len(candidates) != 1:
        raise SellerSpriteImportError("AMBIGUOUS_HEADER", "multiple exact-contract header candidates found")
    header_row = candidates[0]
    raw_header = parsed[header_row - 1]
    mapping, unknown, out_of_scope = _header_mapping(raw_header)
    rows = _materialize_rows(
        ((number, row) for number, row in enumerate(parsed[header_row:], header_row + 1)),
        mapping,
        len(raw_header),
    )
    return RawImportTable(
        source_type="CSV",
        source_basename=basename,
        source_file_sha256=digest,
        source_sheet=None,
        header_row=header_row,
        mapped_headers=tuple(mapping.values()),
        unmapped_headers=unknown,
        out_of_scope_headers=out_of_scope,
        rows=rows,
    )


def _sheet_candidates(worksheet: Any) -> list[tuple[int, tuple[object, ...]]]:
    return [
        (index, tuple(row))
        for index, row in enumerate(
            worksheet.iter_rows(min_row=1, max_row=MAX_HEADER_SCAN_ROWS, values_only=True), 1
        )
        if is_header_candidate(tuple(row))
    ]


def _read_xlsx(
    payload: bytes,
    basename: str,
    digest: str,
    explicit_sheet: str | None,
) -> RawImportTable:
    try:
        workbook = load_workbook(
            BytesIO(payload), read_only=True, data_only=False, keep_links=False
        )
    except Exception as exc:
        raise SellerSpriteImportError("INVALID_XLSX", "file is not a readable XLSX workbook") from exc
    try:
        if explicit_sheet is not None:
            if explicit_sheet not in workbook.sheetnames:
                raise SellerSpriteImportError("SHEET_NOT_FOUND", "explicit sheet does not exist")
            worksheets = [workbook[explicit_sheet]]
        elif PREFERRED_RAW_SHEET in workbook.sheetnames:
            worksheets = [workbook[PREFERRED_RAW_SHEET]]
        else:
            worksheets = list(workbook.worksheets)

        candidates: list[tuple[Any, int, tuple[object, ...]]] = []
        for worksheet in worksheets:
            candidates.extend((worksheet, row_number, row) for row_number, row in _sheet_candidates(worksheet))
        if not candidates:
            raise SellerSpriteImportError("HEADER_SCHEMA_MISMATCH", "no bounded exact-contract header candidate found")
        if len(candidates) != 1:
            raise SellerSpriteImportError("AMBIGUOUS_HEADER", "multiple exact-contract header candidates found")

        worksheet, header_row, raw_header = candidates[0]
        mapping, unknown, out_of_scope = _header_mapping(raw_header)
        raw_rows = (
            (number, tuple(row))
            for number, row in enumerate(
                worksheet.iter_rows(min_row=header_row + 1, values_only=True), header_row + 1
            )
        )
        rows = _materialize_rows(raw_rows, mapping, len(raw_header))
        return RawImportTable(
            source_type="XLSX",
            source_basename=basename,
            source_file_sha256=digest,
            source_sheet=worksheet.title,
            header_row=header_row,
            mapped_headers=tuple(mapping.values()),
            unmapped_headers=unknown,
            out_of_scope_headers=out_of_scope,
            rows=rows,
        )
    finally:
        workbook.close()


def read_local_export(path: str | Path, *, explicit_sheet: str | None = None) -> RawImportTable:
    source = Path(path)
    if not source.is_file():
        raise SellerSpriteImportError("SOURCE_NOT_FOUND", "local source file does not exist")
    payload = source.read_bytes()
    digest = sha256(payload).hexdigest()
    suffix = source.suffix.casefold()
    if suffix == ".csv":
        if explicit_sheet is not None:
            raise SellerSpriteImportError("SHEET_NOT_APPLICABLE", "CSV input does not accept a sheet name")
        return _read_csv(payload, source.name, digest)
    if suffix == ".xlsx":
        return _read_xlsx(payload, source.name, digest, explicit_sheet)
    raise SellerSpriteImportError("UNSUPPORTED_SOURCE_TYPE", "only .xlsx and .csv are accepted")


__all__ = ("RawImportTable", "SourceRow", "read_local_export")
