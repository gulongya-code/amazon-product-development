from __future__ import annotations

from io import BytesIO
import socket
from unittest.mock import patch

from openpyxl import Workbook, load_workbook
from openpyxl.workbook.defined_name import DefinedName
import pytest

from amazon_product_intelligence.contracts import ContractValidationError
from amazon_product_intelligence.operator_template_contract import (
    HIDDEN_SHEET_NAMES,
    RAW_HEADER_CONTRACTS,
    TEMPLATE_CONTRACT_V1,
    TEMPLATE_SCHEMA_FINGERPRINT,
    VISIBLE_SHEET_NAMES,
    FormulaDisposition,
    OperatorTemplateContractV1,
    OperatorTemplateContractValidationError,
    RawHeaderRequirement,
    audit_and_validate_workbook,
    audit_workbook,
    template_schema_fingerprint,
    validate_workbook_audit,
    workbook_audit_fingerprint,
)


def _synthetic_workbook_bytes(
    *,
    raw_headers: tuple[str, ...] | None = None,
    market_formula: str = "=SUM(A1:A9)*0.30",
) -> bytes:
    workbook = Workbook()
    workbook.remove(workbook.active)
    for name in VISIBLE_SHEET_NAMES:
        workbook.create_sheet(name)
    for name in HIDDEN_SHEET_NAMES:
        worksheet = workbook.create_sheet(name)
        worksheet.sheet_state = "hidden"

    headers = raw_headers or tuple(
        reversed(tuple(item.name for item in RAW_HEADER_CONTRACTS))
    )
    for index, value in enumerate(headers, 1):
        workbook["原始数据源"].cell(1, index, value)

    workbook["市场调研"]["A2"] = market_formula
    workbook["不同维度分析"]["B2"] = "=IF(A2>=100,1,0)"
    workbook["自动化配置"]["A2"] = "=0.3"
    workbook["关键词1—数据源"]["C2"] = "=A2"
    workbook["分析模型对比"]["D2"] = "=SUM(A1:A3)"
    workbook["自动化辅助"]["E2"] = "=IF(A2>5,TRUE,FALSE)"

    workbook.defined_names.add(
        DefinedName(
            "PivotSourceKeyword",
            attr_text="'关键词1—数据源'!$A$1:$C$2",
        )
    )
    workbook.defined_names.add(
        DefinedName(
            "PivotSourceCompetitor",
            attr_text="'自动化辅助'!$A$1:$E$2",
        )
    )
    workbook.defined_names.add(
        DefinedName("可选蓝色参数", attr_text="'自动化配置'!$A$1:$A$2")
    )
    workbook["竞品数据"].auto_filter.ref = "A1:B2"
    workbook["关键词1—数据源"].auto_filter.ref = "A1:C2"
    workbook["原始数据源"].auto_filter.ref = "A1:BN2"

    target = BytesIO()
    workbook.save(target)
    workbook.close()
    return target.getvalue()


def test_contract_freezes_exact_11_visible_and_4_hidden_sheets():
    assert VISIBLE_SHEET_NAMES == (
        "综合说明",
        "类目",
        "市场调研",
        "竞品数据",
        "不同维度分析",
        "分析模型对比",
        "top100—日单量分析",
        "产品初步筛选范围",
        "价格核算",
        "样品类型",
        "竞品收集",
    )
    assert HIDDEN_SHEET_NAMES == (
        "自动化配置",
        "原始数据源",
        "关键词1—数据源",
        "自动化辅助",
    )
    assert len(TEMPLATE_CONTRACT_V1.sheets) == 15


def test_contract_freezes_66_unique_headers_by_name():
    names = tuple(item.name for item in RAW_HEADER_CONTRACTS)
    assert len(names) == len(set(names)) == 66
    assert names[0] == "图片"
    assert names[-1] == "标签"
    out_of_scope = tuple(
        item.name for item in RAW_HEADER_CONTRACTS
        if item.requirement is RawHeaderRequirement.OUT_OF_SCOPE
    )
    assert out_of_scope == ("LQS", "SP广告")
    provider_margin = next(item for item in RAW_HEADER_CONTRACTS if item.name == "毛利率")
    assert provider_margin.requirement is RawHeaderRequirement.OPTIONAL
    assert provider_margin.semantic_note == "REFERENCE_ONLY_NOT_PROCUREMENT_TRUTH"


def test_contract_missingness_and_network_are_fail_closed():
    assert (
        TEMPLATE_CONTRACT_V1.numeric_missing_policy
        == "MISSING_BLANK_NA_PARSE_FAILURE_NEVER_ZERO"
    )
    assert TEMPLATE_CONTRACT_V1.external_network_calls_allowed is False


def test_contract_round_trip_and_unknown_field_rejection():
    payload = TEMPLATE_CONTRACT_V1.to_dict()
    rebuilt = OperatorTemplateContractV1.from_dict(payload)
    assert rebuilt == TEMPLATE_CONTRACT_V1
    payload["unknown"] = True
    with pytest.raises(ContractValidationError):
        OperatorTemplateContractV1.from_dict(payload)


def test_schema_fingerprint_is_deterministic():
    first = template_schema_fingerprint(TEMPLATE_CONTRACT_V1)
    rebuilt = OperatorTemplateContractV1.from_dict(TEMPLATE_CONTRACT_V1.to_dict())
    second = template_schema_fingerprint(rebuilt)
    assert first == second == TEMPLATE_SCHEMA_FINGERPRINT
    assert len(first) == 64


def test_reference_formula_census_is_frozen_as_approximately_26738():
    census = {
        item.sheet_name: item.approximate_count
        for item in TEMPLATE_CONTRACT_V1.formula_census_reference
    }
    assert census == {
        "市场调研": 3,
        "不同维度分析": 1150,
        "自动化配置": 2,
        "关键词1—数据源": 961,
        "分析模型对比": 108,
        "自动化辅助": 24514,
    }
    assert sum(census.values()) == 26738


def test_synthetic_workbook_passes_exact_contract_with_shuffled_headers():
    snapshot = audit_and_validate_workbook(_synthetic_workbook_bytes())
    assert tuple(item.name for item in snapshot.sheet_states if item.state == "visible") == VISIBLE_SHEET_NAMES
    assert {item.name for item in snapshot.sheet_states if item.state == "hidden"} == set(HIDDEN_SHEET_NAMES)
    assert len(snapshot.raw_headers) == 66
    assert snapshot.raw_headers[0] == "标签"


def test_sheet_or_visibility_drift_fails_closed():
    content = _synthetic_workbook_bytes()
    workbook = load_workbook(BytesIO(content))
    workbook["竞品收集"].title = "竞品收集-漂移"
    target = BytesIO()
    workbook.save(target)
    workbook.close()
    snapshot = audit_workbook(target.getvalue())
    with pytest.raises(
        OperatorTemplateContractValidationError,
        match="contracted 15 sheets|visible sheet",
    ):
        validate_workbook_audit(snapshot)


def test_very_hidden_support_sheet_fails_closed():
    content = _synthetic_workbook_bytes()
    workbook = load_workbook(BytesIO(content))
    workbook["自动化配置"].sheet_state = "veryHidden"
    target = BytesIO()
    workbook.save(target)
    workbook.close()
    with pytest.raises(
        OperatorTemplateContractValidationError,
        match="hidden sheet|VeryHidden",
    ):
        audit_and_validate_workbook(target.getvalue())


def test_header_missing_extra_or_duplicate_fails_closed():
    names = tuple(item.name for item in RAW_HEADER_CONTRACTS)
    invalid = names[:-1] + (names[0],)
    with pytest.raises(
        OperatorTemplateContractValidationError,
        match="66 unique|missing raw headers",
    ):
        audit_and_validate_workbook(
            _synthetic_workbook_bytes(raw_headers=invalid)
        )


def test_required_named_ranges_and_filters_fail_closed():
    content = _synthetic_workbook_bytes()
    workbook = load_workbook(BytesIO(content))
    del workbook.defined_names["PivotSourceKeyword"]
    workbook["竞品数据"].auto_filter.ref = None
    target = BytesIO()
    workbook.save(target)
    workbook.close()
    with pytest.raises(
        OperatorTemplateContractValidationError,
        match="required named range missing.*required AutoFilter missing",
    ):
        audit_and_validate_workbook(target.getvalue())


def test_formula_census_fingerprint_and_audit_fingerprint_are_deterministic():
    content = _synthetic_workbook_bytes()
    first = audit_workbook(content)
    second = audit_workbook(content)
    assert first == second
    assert first.formula_sheets == second.formula_sheets
    assert workbook_audit_fingerprint(first) == workbook_audit_fingerprint(second)
    census = {item.sheet_name: item.formula_count for item in first.formula_sheets}
    assert census["市场调研"] == 1
    assert census["价格核算"] == 0


def test_formula_change_changes_only_relevant_sheet_fingerprint():
    first = audit_workbook(_synthetic_workbook_bytes(market_formula="=SUM(A1:A9)*0.30"))
    second = audit_workbook(_synthetic_workbook_bytes(market_formula="=SUM(A1:A9)*0.31"))
    first_fp = {item.sheet_name: item.formula_fingerprint for item in first.formula_sheets}
    second_fp = {item.sheet_name: item.formula_fingerprint for item in second.formula_sheets}
    changed = {name for name in first_fp if first_fp[name] != second_fp[name]}
    assert changed == {"市场调研"}


def test_formula_tokenizer_inventory_does_not_treat_cell_rows_as_thresholds():
    snapshot = audit_workbook(_synthetic_workbook_bytes())
    market_literals = tuple(
        item.value for item in snapshot.threshold_literals
        if item.sheet_name == "市场调研"
    )
    assert market_literals == ("0.30",)
    assert "1" not in market_literals
    market_formula = next(
        item for item in snapshot.formula_cells if item.sheet_name == "市场调研"
    )
    assert market_formula.disposition is FormulaDisposition.MOVE_TO_CONFIG
    keyword_formula = next(
        item for item in snapshot.formula_cells if item.sheet_name == "关键词1—数据源"
    )
    assert keyword_formula.disposition is FormulaDisposition.REUSE_AS_FORMULA


def test_product_selection_semantics_forbid_direct_competitor_before_lock():
    semantics = {
        item.sheet_name: item.before_direction_locked
        for item in TEMPLATE_CONTRACT_V1.product_selection_semantics
    }
    assert semantics == {
        "产品初步筛选范围": "CANDIDATE_PRODUCT_ARCHETYPES",
        "样品类型": "HYPOTHESES_AND_SAMPLING_DIRECTIONS",
        "竞品收集": "REPRESENTATIVE_ASINS_ONLY",
    }
    assert all(
        item.direct_competitor_label_allowed_before_lock is False
        for item in TEMPLATE_CONTRACT_V1.product_selection_semantics
    )


def test_audit_constructs_no_network_socket():
    content = _synthetic_workbook_bytes()
    with patch.object(
        socket,
        "socket",
        side_effect=AssertionError("network socket construction is forbidden"),
    ):
        snapshot = audit_and_validate_workbook(content)
    assert snapshot.schema_fingerprint == TEMPLATE_SCHEMA_FINGERPRINT
