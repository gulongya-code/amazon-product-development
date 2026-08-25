"""Operator-first XLSX renderer for Market Report V0.2."""

from __future__ import annotations

import json
from datetime import datetime
from pathlib import Path
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill

from ..models import MarketReportSnapshotV0_2
from .operator_view import compose_operator_view


SHEET_NAMES = (
    "Executive Summary", "Market Overview", "Market Size", "Competition",
    "Distributions", "Competitor Details", "Buyer Needs", "Product Directions",
    "Competitor Shortlist", "Opportunity", "Evidence Gaps", "Audit - Provenance",
)


def _cell(value: Any) -> Any:
    if value is None:
        return "UNKNOWN"
    if isinstance(value, (dict, list, tuple)):
        return json.dumps(value, ensure_ascii=False, sort_keys=True)
    if isinstance(value, str) and value.lstrip().startswith(("=", "+", "-", "@")):
        return "'" + value
    return value


class ExcelReportRendererV0_2:
    def render(self, report: MarketReportSnapshotV0_2, destination: Path) -> None:
        view = compose_operator_view(report)
        p = view.payload
        content = (
            ("Executive Summary", [*view.parity.items(), *p["executive_summary"].items()]),
            ("Market Overview", [("category", p["category"]), ("sample", p["sample"]), ("data_window", p["data_window"]), ("scope_context", p["scope_context"])]),
            ("Market Size", list(p["market_size"].items())),
            ("Competition", [("true_competitor_set", p["true_competitor_set"]), ("competitor_structure", p["competitor_structure"])]),
            ("Distributions", [("items", p["distributions"])]),
            ("Competitor Details", [("items", p["competitor_details"])]),
            ("Buyer Needs", list(p["buyer_needs"].items())),
            ("Product Directions", [("semantics", "HYPOTHESIS"), *p["product_directions"].items()]),
            ("Competitor Shortlist", [("semantics", "REVIEW_ORDER_NOT_RANK"), *p["competitor_shortlist"].items()]),
            ("Opportunity", list(p["opportunity_score"].items())),
            ("Evidence Gaps", [("report_limitations", p["limitations"]), ("sanitized_appendix", p["sanitized_appendix"]), ("external_integrations", p["external_integrations"])]),
            ("Audit - Provenance", [("metadata", p["metadata"]), ("provenance", p["provenance"]), ("evidence_registry", p["evidence_registry"])]),
        )
        workbook = Workbook()
        workbook.properties.creator = "amazon-product-intelligence"
        workbook.properties.lastModifiedBy = "amazon-product-intelligence"
        workbook.properties.created = datetime(2000, 1, 1)
        workbook.properties.modified = datetime(2000, 1, 1)
        workbook.remove(workbook.active)
        for title, rows in content:
            sheet = workbook.create_sheet(title)
            sheet.append(("Field", "Value"))
            for key, value in rows:
                sheet.append((_cell(key), _cell(value)))
            sheet.freeze_panes = "A2"
            sheet.auto_filter.ref = f"A1:B{sheet.max_row}"
            sheet.column_dimensions["A"].width = 34
            sheet.column_dimensions["B"].width = 120
            for cell in sheet[1]:
                cell.font = Font(bold=True, color="FFFFFF")
                cell.fill = PatternFill("solid", fgColor="1F4E78")
            for row in sheet.iter_rows():
                for cell in row:
                    cell.alignment = Alignment(vertical="top", wrap_text=True)
        if tuple(workbook.sheetnames) != SHEET_NAMES:
            raise RuntimeError("V0.2 operator workbook sheet contract drifted")
        destination.parent.mkdir(parents=True, exist_ok=True)
        temporary = destination.with_name(f".{destination.name}.tmp.xlsx")
        workbook.save(temporary)
        temporary.replace(destination)


__all__ = ("SHEET_NAMES", "ExcelReportRendererV0_2")
