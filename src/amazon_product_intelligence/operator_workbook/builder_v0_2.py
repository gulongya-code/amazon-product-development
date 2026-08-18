"""Deterministic nine-sheet Operator Workbook V0.2 presentation builder."""

from __future__ import annotations

import base64
from collections.abc import Mapping, Sequence
from datetime import datetime
from hashlib import sha256
from io import BytesIO
import re
from typing import Any
from zipfile import ZIP_DEFLATED, ZipFile, ZipInfo

import openpyxl
from openpyxl import Workbook
from openpyxl.formatting.rule import FormulaRule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

from amazon_product_intelligence.contracts import canonical_json, deterministic_id

from .errors import OperatorWorkbookValidationError
from .models import (
    OPERATOR_WORKBOOK_RULESET_VERSION,
    WORKBOOK_FILENAME,
    WORKBOOK_MEDIA_TYPE,
    OperatorWorkbookRequest,
    OperatorWorkbookSnapshotV0_2,
    WorkbookCoverageSummary,
    WorkbookDiagnostic,
    WorkbookFieldDefinition,
    WorkbookFileRecord,
    WorkbookLineageReference,
    WorkbookRowRecord,
    WorkbookSheetDefinition,
    _bundle_fingerprint,
)
from .schema_v0_2 import SHEET_SPECS


_FIXED_DOCUMENT_TIME = datetime(2000, 1, 1, 0, 0, 0)
_ZIP_TIME = (1980, 1, 1, 0, 0, 0)
_NOT_AVAILABLE = "NOT_AVAILABLE"
_DANGEROUS_PREFIXES = ("=", "+", "-", "@")


def _escape_formula(value: str) -> str:
    return f"'{value}" if value.lstrip().startswith(_DANGEROUS_PREFIXES) else value


def _scalar(value: Any) -> Any:
    if value is None:
        return _NOT_AVAILABLE
    if type(value) in {str, bool, int, float}:
        return value
    if isinstance(value, Mapping):
        pairs = []
        for key in sorted(value):
            item = value[key]
            if item is None or type(item) in {str, bool, int, float}:
                pairs.append(f"{key}={_NOT_AVAILABLE if item is None else item}")
        return "; ".join(pairs) if pairs else _NOT_AVAILABLE
    if isinstance(value, Sequence) and not isinstance(value, (str, bytes)):
        return " | ".join(str(_scalar(item)) for item in value) or _NOT_AVAILABLE
    return str(value)


def _join(values: Sequence[Any]) -> str:
    rendered = tuple(str(_scalar(item)) for item in values if item not in {None, ""})
    return " | ".join(rendered) if rendered else _NOT_AVAILABLE


def _records(snapshot: Mapping[str, Any], key: str) -> tuple[Mapping[str, Any], ...]:
    value = snapshot.get(key, ())
    if not isinstance(value, Sequence) or isinstance(value, (str, bytes)):
        raise OperatorWorkbookValidationError(f"{key} must be an array")
    return tuple(item for item in value if isinstance(item, Mapping))


def _candidate_view(
    records: Sequence[Mapping[str, Any]],
    *,
    dimension_key: str,
    dimension: str,
) -> tuple[Any, str, tuple[Mapping[str, Any], ...]]:
    selected = tuple(item for item in records if item.get(dimension_key) == dimension)
    candidates = tuple(
        candidate
        for item in selected
        for candidate in item.get("candidates", ())
        if isinstance(candidate, Mapping)
    )
    present = tuple(
        candidate for candidate in candidates
        if candidate.get("presence_status") == "PRESENT"
        or (
            isinstance(candidate.get("value"), Mapping)
            and candidate["value"].get("presence_status") == "PRESENT"
        )
    )
    values: list[Any] = []
    for candidate in present:
        envelope = candidate.get("value")
        if isinstance(envelope, Mapping):
            value = envelope.get("normalized_value", envelope.get("raw_value"))
        else:
            value = candidate.get("normalized_value", candidate.get("raw_value"))
        if canonical_json(value) not in {canonical_json(item) for item in values}:
            values.append(value)
    if not selected:
        return _NOT_AVAILABLE, "NOT_AVAILABLE", ()
    states = tuple(str(item.get("candidate_state", "UNKNOWN")) for item in selected)
    if not values:
        return _NOT_AVAILABLE, _join(states), candidates
    if len(values) == 1:
        state = _join(states)
        if state == "UNKNOWN" or set(states) == {"UNKNOWN"}:
            state = "ONE_DISTINCT_PRESENT_VALUE"
        return values[0], state, candidates
    described = tuple(
        f"{candidate.get('provider', 'UNKNOWN')}={_scalar(candidate.get('normalized_value', candidate.get('raw_value')))}"
        for candidate in present
    )
    return f"MULTIPLE_CANDIDATES: {_join(described)}", "MULTIPLE_CANDIDATES", candidates


def _candidate_unit(candidates: Sequence[Mapping[str, Any]]) -> str:
    units = []
    for candidate in candidates:
        envelope = candidate.get("value")
        unit = envelope.get("unit") if isinstance(envelope, Mapping) else candidate.get("unit")
        if isinstance(unit, Mapping):
            code = unit.get("unit_code")
            if type(code) is str and code not in units:
                units.append(code)
    return _join(units)


def _candidate_providers(candidates: Sequence[Mapping[str, Any]]) -> tuple[str, ...]:
    return tuple(sorted({
        str(item["provider"]) for item in candidates
        if type(item.get("provider")) is str and item["provider"].strip()
    }))


def _source_ids(request: OperatorWorkbookRequest) -> dict[str, str]:
    return {
        "competition_intelligence": request.competition_intelligence_snapshot["snapshot_id"],
        "demand_intelligence": request.demand_intelligence_snapshot["snapshot_id"],
        "evidence_evaluation": request.evidence_evaluation_snapshot["snapshot_id"],
        "operator_export": request.operator_export_snapshot["snapshot_id"],
        "operator_output": request.operator_output_snapshot["snapshot_id"],
        "opportunity_intelligence": request.opportunity_intelligence_snapshot["snapshot_id"],
        "opportunity_scoring": request.opportunity_scoring_snapshot["snapshot_id"],
        "product_intelligence": request.product_intelligence_snapshot["snapshot_id"],
        "recommendation_framework": request.recommendation_framework_snapshot["snapshot_id"],
    }


def _definitions() -> tuple[
    tuple[WorkbookFieldDefinition, ...], tuple[WorkbookSheetDefinition, ...]
]:
    fields: list[WorkbookFieldDefinition] = []
    sheets: list[WorkbookSheetDefinition] = []
    for sheet_ordinal, spec in enumerate(SHEET_SPECS, start=1):
        sheet_fields: list[WorkbookFieldDefinition] = []
        for ordinal, field_spec in enumerate(spec.fields, start=1):
            content = {
                "sheet_key": spec.key,
                "ordinal": ordinal,
                "chinese_name": field_spec.chinese_name,
                "english_name": field_spec.english_name,
                "data_type": field_spec.data_type,
                "source": field_spec.source,
                "visible": not field_spec.default_hidden,
                "default_hidden": field_spec.default_hidden,
                "operator_use": field_spec.operator_use,
                "column_width": float(field_spec.width),
            }
            item = WorkbookFieldDefinition(
                field_id=deterministic_id("operator-workbook-field", content),
                **content,
            )
            fields.append(item)
            sheet_fields.append(item)
        sheet_content = {
            "ordinal": sheet_ordinal,
            "sheet_key": spec.key,
            "sheet_name": spec.name,
            "purpose": spec.purpose,
            "warning": spec.warning,
            "row_grain": spec.row_grain,
            "hidden": spec.hidden,
            "field_ids": tuple(item.field_id for item in sheet_fields),
        }
        sheets.append(WorkbookSheetDefinition(
            sheet_id=deterministic_id("operator-workbook-sheet", sheet_content),
            row_ids=(),
            **sheet_content,
        ))
    return tuple(fields), tuple(sheets)


def _value_map(
    sheet: WorkbookSheetDefinition,
    fields_by_sheet: Mapping[str, tuple[WorkbookFieldDefinition, ...]],
    values: Mapping[str, Any],
) -> dict[str, Any]:
    result: dict[str, Any] = {}
    for field in fields_by_sheet[sheet.sheet_key]:
        result[field.field_id] = values.get(field.english_name, _NOT_AVAILABLE)
    return result


def _display_plans(
    request: OperatorWorkbookRequest,
) -> Mapping[str, tuple[tuple[str, Mapping[str, Any], tuple[str, ...], tuple[str, ...]], ...]]:
    product = request.product_intelligence_snapshot
    demand = request.demand_intelligence_snapshot
    opportunity = request.opportunity_intelligence_snapshot
    scoring = request.opportunity_scoring_snapshot
    output = request.operator_output_snapshot
    facts = _records(product, "product_fact_evidence_sets")
    metrics = _records(product, "product_metric_series")
    demand_metrics = _records(demand, "keyword_metric_evidence_sets")
    product_output = _records(output, "product_rows")[0]
    keyword_output = _records(output, "keyword_rows")[0]
    opportunity_output = _records(output, "opportunity_rows")[0]
    target = product["target_product_identity"]
    keyword = demand["target_keyword_identity"]

    title, title_state, title_candidates = _candidate_view(
        facts, dimension_key="dimension", dimension="title"
    )
    brand, _, brand_candidates = _candidate_view(
        facts, dimension_key="dimension", dimension="brand"
    )
    category, _, category_candidates = _candidate_view(
        facts, dimension_key="dimension", dimension="category"
    )
    product_type, _, type_candidates = _candidate_view(
        facts, dimension_key="dimension", dimension="product_type"
    )
    price, price_state, price_candidates = _candidate_view(
        metrics, dimension_key="metric", dimension="price"
    )
    rating, rating_state, rating_candidates = _candidate_view(
        metrics, dimension_key="metric", dimension="rating"
    )
    review_count, review_state, review_candidates = _candidate_view(
        metrics, dimension_key="metric", dimension="review_count"
    )
    bsr, bsr_state, bsr_candidates = _candidate_view(
        metrics, dimension_key="metric", dimension="bsr"
    )
    sales, sales_state, sales_candidates = _candidate_view(
        metrics, dimension_key="metric", dimension="estimated_monthly_sales"
    )
    source_candidates = (
        title_candidates + brand_candidates + category_candidates + type_candidates
        + price_candidates + rating_candidates + review_candidates + bsr_candidates
        + sales_candidates
    )
    providers = _candidate_providers(source_candidates)
    attributes = []
    attribute_dimensions = tuple(sorted({
        str(fact.get("dimension")) for fact in facts
        if fact.get("dimension") not in {"title", "brand", "category", "description"}
    }))
    for dimension in attribute_dimensions:
        value, state, _ = _candidate_view(
            facts, dimension_key="dimension", dimension=dimension
        )
        attributes.append(f"{dimension}={_scalar(value)} [{state}]")
    topology = product.get("variation_topology", {})
    parent_asin = target.get("parent_asin") or _NOT_AVAILABLE
    child_count = len(topology.get("child_product_identities", ())) if isinstance(topology, Mapping) else 0
    variation_role = "CHILD" if target.get("parent_asin") else "UNKNOWN"
    data_state = product.get("evidence_coverage_summary", {}).get("coverage_status", "EVIDENCE_PRESENT")
    conflict_state = "MULTIPLE_CANDIDATES" if any(
        state == "MULTIPLE_CANDIDATES" for state in (title_state, price_state, rating_state, review_state)
    ) else "NO_MATERIAL_DISPLAY_CONFLICT"
    product_values = {
        "ASIN": target.get("asin"), "Marketplace": target.get("marketplace"),
        "Display Title": title, "Title State": title_state, "Brand": brand,
        "Category": category, "Product Type": product_type, "Price": price,
        "Price Currency": _candidate_unit(price_candidates), "Price State": price_state,
        "Rating": rating, "Rating State": rating_state,
        "Review Evidence Count": review_count, "BSR": bsr,
        "BSR Context": bsr_state, "Sales Evidence Value": sales,
        "Sales Evidence Unit": _candidate_unit(sales_candidates),
        "Sales Evidence Type": sales_candidates[0].get("evidence_type", sales_state) if sales_candidates else sales_state,
        "Variation Role": variation_role, "Parent ASIN": parent_asin,
        "Child Count": child_count, "Attribute Summary": _join(attributes),
        "Seller": _NOT_AVAILABLE, "FBA Status": _NOT_AVAILABLE,
        "Data Sources": _join(providers), "Data State": data_state,
        "Conflict State": conflict_state, "Time / Period Status": "SOURCE_TIME_STATUS_PRESERVED",
        "Product Snapshot ID": product["snapshot_id"], "Output Row ID": product_output["output_row_id"],
    }

    search_volume, search_state, search_candidates = _candidate_view(
        demand_metrics, dimension_key="metric", dimension="search_volume"
    )
    cpc, cpc_state, cpc_candidates = _candidate_view(
        demand_metrics, dimension_key="metric", dimension="cpc"
    )
    aba, aba_state, aba_candidates = _candidate_view(
        demand_metrics, dimension_key="metric", dimension="aba_search_frequency_rank"
    )
    difficulty, difficulty_state, difficulty_candidates = _candidate_view(
        demand_metrics, dimension_key="metric", dimension="competition_difficulty"
    )
    query_records = _records(demand, "query_execution_evidence")
    relationship_groups = _records(demand, "relationship_evidence_groups")
    related_inventory = _records(demand, "related_product_evidence_inventory")
    related_ids = tuple(sorted({
        str(identity.get("asin") or identity.get("product_id"))
        for item in related_inventory
        for identity in item.get("product_identities", ())
        if isinstance(identity, Mapping)
    }))
    channels = tuple(sorted({
        str(item.get("channel")) for item in relationship_groups if item.get("channel")
    }))
    directions = tuple(sorted({
        str(item.get("direction")) for item in query_records if item.get("direction")
    }))
    query_statuses = tuple(sorted({
        str(item.get("result_status") or item.get("outcome") or "UNKNOWN") for item in query_records
    }))
    keyword_candidates = search_candidates + cpc_candidates + aba_candidates + difficulty_candidates
    keyword_values = {
        "Keyword": keyword.get("normalized_text") or keyword.get("raw_text"),
        "Marketplace": keyword.get("marketplace"), "Locale": keyword.get("locale"),
        "Search Volume": search_volume, "Search Volume State": search_state,
        "Search Volume Unit": _candidate_unit(search_candidates), "CPC": cpc,
        "CPC Currency": _candidate_unit(cpc_candidates), "CPC State": cpc_state,
        "ABA Rank": aba, "ABA Rank State": aba_state, "Difficulty": difficulty,
        "Difficulty State": difficulty_state, "Related Product Count": len(related_ids),
        "Related Product ASINs": _join(related_ids), "Channel": _join(channels),
        "Query Direction": _join(directions), "Query Status": _join(query_statuses),
        "Provider": _join(_candidate_providers(keyword_candidates)),
        "Estimate Method Status": _join(tuple(
            str(item.get("estimate_method_status", "UNKNOWN")) for item in keyword_candidates
        )),
        "Period Status": _join(tuple(
            str(item.get("time", {}).get("period_type", "UNKNOWN")) for item in keyword_candidates
        )),
        "Limitations": _join(keyword_output.get("limitations", ())),
        "Demand Snapshot ID": demand["snapshot_id"],
    }

    risks = _records(opportunity, "risk_evidence")
    missing_inventory = opportunity.get("missing_evidence", {})
    missing_items = missing_inventory.get("items", ()) if isinstance(missing_inventory, Mapping) else ()
    quality_profiles = _records(request.evidence_evaluation_snapshot, "quality_profiles")
    market_values = {
        "Marketplace": target.get("marketplace"), "Category Candidate": category,
        "Market Size Evidence Metric": "search_volume (provider evidence)",
        "Metric Value": search_volume, "Unit": _candidate_unit(search_candidates),
        "Observed Product Count": len(product.get("included_product_identities", ())),
        "Data Sources": _join(tuple(sorted(set(providers) | set(_candidate_providers(keyword_candidates))))),
        "Evidence-backed Trend": _NOT_AVAILABLE,
        "Risk Alerts": _join(tuple(item.get("risk_type", item.get("code", "RISK_EVIDENCE")) for item in risks)),
        "Evidence Quality": _join(tuple(item.get("overall_quality", item.get("quality_status", "QUALITY_RECORDED")) for item in quality_profiles)),
        "Analysis Limitations": "NO_MARKET_GUARANTEE | NO_TREND_INFERENCE | PROVIDER_METHODS_PRESERVED",
        "Snapshot ID": opportunity["snapshot_id"],
    }

    top_values = {
        "Product ASIN": target.get("asin"), "Display Title": title,
        "Marketplace": target.get("marketplace"), "Source Rank Value": _NOT_AVAILABLE,
        "Rank Metric": _NOT_AVAILABLE, "Rank Context": _NOT_AVAILABLE,
        "Channel": _NOT_AVAILABLE, "Rank Provider": _NOT_AVAILABLE,
        "Rank Status": "NOT_AVAILABLE", "Rank Period": _NOT_AVAILABLE,
        "Price": price, "Review Evidence Count": review_count,
        "Rating Evidence": rating, "Product Features": _join(attributes),
        "Data Limitations": "NO_EXPLICIT_RANK_EVIDENCE | NOT_BEST_PRODUCT",
        "Rank Observation ID": _NOT_AVAILABLE,
    }

    price_values = []
    for candidate in price_candidates:
        value = candidate.get("normalized_value", candidate.get("raw_value"))
        if type(value) in {int, float}:
            price_values.append(float(value))
    structure_values = {
        "Marketplace": target.get("marketplace"), "Product Type": product_type,
        "Product Count": len(product.get("included_product_identities", ())),
        "Observed Share": 1.0 if product.get("included_product_identities") else _NOT_AVAILABLE,
        "Sales Evidence Summary": f"{_scalar(sales)} {_candidate_unit(sales_candidates)} [{sales_state}]",
        "Minimum Comparable Price": min(price_values) if price_values else _NOT_AVAILABLE,
        "Maximum Comparable Price": max(price_values) if price_values else _NOT_AVAILABLE,
        "Currency": _candidate_unit(price_candidates), "Observed Feature Inventory": _join(attributes),
        "Data State": data_state, "Provider Count": len(providers),
        "Limitations": "EXACT_OBSERVED_GROUP_ONLY | NO_CLUSTERING | NO_MARKET_SHARE_INFERENCE",
        "Member Product IDs": _join(tuple(
            item.get("product_id", item.get("asin", _NOT_AVAILABLE))
            for item in product.get("included_product_identities", ()) if isinstance(item, Mapping)
        )),
    }

    competition_plans = []
    for row in _records(output, "competition_rows"):
        endpoint = row.get("product_endpoint", {})
        relationship = row.get("keyword_relationship", {})
        keyword_identity = relationship.get("keyword_identity", {}) if isinstance(relationship, Mapping) else {}
        values = {
            "Product ASIN": endpoint.get("asin", endpoint.get("product_id", _NOT_AVAILABLE)),
            "Keyword": keyword_identity.get("normalized_text", keyword_identity.get("raw_text", _NOT_AVAILABLE)),
            "Relationship Direction": relationship.get("direction", _NOT_AVAILABLE),
            "Observed Relationship": "OBSERVED_PRODUCT_KEYWORD_RELATIONSHIP",
            "Observed Relationship Type": row.get("relationship_type"),
            "Channel": row.get("channel"), "Provider": row.get("provider"),
            "Evidence Count": row.get("evidence_count"),
            "Evidence Classification": "OBSERVED_RELATIONSHIP_EVIDENCE",
            "Variation Evidence Count": len(row.get("variation_evidence", ())),
            "Query Status": _join(relationship.get("query_result_statuses", ())),
            "Limitations": _join(row.get("limitations", ())),
            "Competition Output Row ID": row.get("output_row_id"),
        }
        competition_plans.append((
            str(row["output_row_id"]), values, (str(row["output_row_id"]),),
            (str(row["source_snapshot_id"]),),
        ))

    observed_signals = _records(opportunity, "observed_signals")
    derived_signals = _records(opportunity, "derived_signals")
    signal_types = tuple(sorted({
        str(item.get("signal_type", item.get("dimension", "SIGNAL")))
        for item in observed_signals + derived_signals
    }))
    demand_signal_types = tuple(
        item for item in signal_types if "KEYWORD_METRIC" in item or "QUERY" in item
    )
    competition_signal_types = tuple(
        item for item in signal_types if "RELATIONSHIP" in item or "VARIATION" in item
    )
    product_signal_types = tuple(
        item for item in signal_types
        if item not in set(demand_signal_types) | set(competition_signal_types)
        and ("PRODUCT" in item or "REVIEW" in item)
    )
    calculations = _records(scoring, "calculations")
    explanations = {item.get("calculation_id"): item for item in _records(scoring, "explanations")}
    product_ids = _join(tuple(
        item.get("asin", item.get("product_id", _NOT_AVAILABLE))
        for item in opportunity_output.get("product", ()) if isinstance(item, Mapping)
    ))
    opportunity_plans = []
    for calculation in calculations:
        explanation = explanations.get(calculation.get("calculation_id"), {})
        values = {
            "Product": product_ids,
            "Demand Signal": _join(demand_signal_types),
            "Competition Signal": _join(competition_signal_types),
            "Product Signal": _join(product_signal_types),
            "Signal Classification": _join(tuple(sorted({
                str(item.get("classification", "SOURCE_SIGNAL"))
                for item in observed_signals + derived_signals
            }))),
            "Missing Evidence": _join(tuple(
                item.get("missing_type", item.get("requirement", item.get("missing_evidence_id", "MISSING_EVIDENCE")))
                for item in missing_items if isinstance(item, Mapping)
            )),
            "Risk Evidence": _join(tuple(
                item.get("risk_type", item.get("risk_id", "RISK_EVIDENCE")) for item in risks
            )),
            "Score Factor": calculation.get("factor_id"),
            "Rule Process Score": calculation.get("result_value", _NOT_AVAILABLE),
            "Score Status": calculation.get("result_status"),
            "Score Reference": calculation.get("calculation_id"),
            "Score Interpretation": explanation.get("result_interpretation", "RULE_PROCESS_SCORE_ONLY"),
            "Explanation Reference": explanation.get("explanation_id", _NOT_AVAILABLE),
            "Limitations": "RULE_PROCESS_SCORE_ONLY | NO_SUCCESS_PROBABILITY | NO_OPPORTUNITY_GUARANTEE",
            "Opportunity Output Row ID": opportunity_output.get("output_row_id"),
        }
        opportunity_plans.append((
            str(calculation.get("calculation_id")), values,
            (str(opportunity_output["output_row_id"]),),
            tuple(str(item) for item in opportunity_output.get("source_snapshot_ids", ())),
        ))

    rec_plans = []
    for row in _records(output, "recommendation_rows"):
        explanation = row.get("explanation", {})
        rule = row.get("rule_reference", {})
        rec_type = str(row.get("recommendation_type"))
        label = {
            "FURTHER_RESEARCH": "进一步调研",
            "SUPPLEMENT_DATA": "补充数据",
            "MANUAL_REVIEW": "人工审核",
        }.get(rec_type, "人工复核")
        values = {
            "Product": product_ids, "Recommendation Type": rec_type,
            "Recommendation Display Label": label,
            "Reason": explanation.get("rule_explanation", rule.get("description", _NOT_AVAILABLE)),
            "Rule Reference": rule.get("rule_id", row.get("source_record_id")),
            "Policy Status": (
                "POLICY_REFERENCES_PRESENT"
                if explanation.get("policy_evaluation_ids") else "NO_POLICY_REFERENCE"
            ),
            "Conflict Status": (
                "CONFLICT_REFERENCES_PRESENT"
                if explanation.get("conflict_ids") else "NO_CONFLICT_REFERENCE"
            ),
            "Missing Requirements": _join(explanation.get("missing_requirements", ())),
            "Evidence References": _join(row.get("evidence_references", ())),
            "Evidence Count": len(row.get("evidence_references", ())),
            "Limitations": _join(row.get("limitations", ())),
            "Manual Review Status": "未复核", "Recommendation Record ID": row.get("source_record_id"),
            "Source Snapshot ID": row.get("source_snapshot_id"),
            "Operator Output Row ID": row.get("output_row_id"),
        }
        rec_plans.append((
            str(row["source_record_id"]), values, (str(row["output_row_id"]),),
            (str(row["source_snapshot_id"]),),
        ))

    return {
        "market_overview": ((
            f"{target.get('marketplace')}:{_scalar(category)}", market_values,
            (str(product_output["output_row_id"]), str(opportunity_output["output_row_id"])),
            (str(product["snapshot_id"]), str(opportunity["snapshot_id"])),
        ),),
        "product_database": ((str(target.get("product_id")), product_values, (str(product_output["output_row_id"]),), (str(product["snapshot_id"]),)),),
        "top_products": ((f"rank:{target.get('product_id')}", top_values, (str(product_output["output_row_id"]),), (str(product["snapshot_id"]),)),),
        "keyword_demand": ((str(keyword.get("keyword_id")), keyword_values, (str(keyword_output["output_row_id"]),), (str(demand["snapshot_id"]),)),),
        "competition_evidence": tuple(competition_plans),
        "product_structure": ((f"structure:{target.get('marketplace')}:{_scalar(product_type)}", structure_values, (str(product_output["output_row_id"]),), (str(product["snapshot_id"]),)),),
        "opportunity_analysis": tuple(opportunity_plans),
        "action_recommendations": tuple(rec_plans),
    }


def _copy_lineage(
    *, row_id: str, sheet_id: str, export_snapshot_id: str,
    output_snapshot_id: str, output_lineage: Mapping[str, Any],
    export_lineage: Mapping[str, Any],
) -> WorkbookLineageReference:
    if (
        export_lineage["source_output_lineage_id"] != output_lineage["output_lineage_id"]
        or export_lineage["source_output_row_id"] != output_lineage["output_row_id"]
    ):
        raise OperatorWorkbookValidationError(
            "operator export lineage does not match operator output lineage"
        )
    content = {
        "row_id": row_id, "sheet_id": sheet_id,
        "source_export_snapshot_id": export_snapshot_id,
        "source_export_row_id": export_lineage["export_row_id"],
        "source_export_lineage_id": export_lineage["export_lineage_id"],
        "source_output_snapshot_id": output_snapshot_id,
        "source_output_row_id": output_lineage["output_row_id"],
        "source_output_lineage_id": output_lineage["output_lineage_id"],
        "source_snapshot_id": output_lineage["source_snapshot_id"],
        "source_record_id": output_lineage["source_record_id"],
        "source_lineage_id": output_lineage["source_lineage_id"],
        "canonical_reference_id": output_lineage["canonical_reference_id"],
        "canonical_reference_type": output_lineage["canonical_reference_type"],
        "semantic_observation_id": output_lineage.get("semantic_observation_id"),
        "transformation_run_id": output_lineage["transformation_run_id"],
        "mapping_version": output_lineage["mapping_version"],
        "raw_evidence_id": output_lineage["raw_evidence_id"],
        "collection_run_id": output_lineage["collection_run_id"],
        "provider": output_lineage["provider"],
        "source_tool": output_lineage["source_tool"],
        "source_field": output_lineage["source_field"],
        "source_bundle_fingerprints": tuple(sorted(output_lineage["source_bundle_fingerprints"])),
    }
    return WorkbookLineageReference(
        workbook_lineage_id=deterministic_id("operator-workbook-lineage", content),
        **content,
    )


def _normalize_zip(content: bytes) -> bytes:
    source_buffer = BytesIO(content)
    target_buffer = BytesIO()
    with ZipFile(source_buffer, "r") as source, ZipFile(
        target_buffer, "w", compression=ZIP_DEFLATED, compresslevel=9, allowZip64=True
    ) as target:
        for name in sorted(source.namelist()):
            payload = source.read(name)
            if name == "docProps/core.xml":
                payload = re.sub(
                    rb"(<dcterms:modified[^>]*>)[^<]*(</dcterms:modified>)",
                    rb"\g<1>2000-01-01T00:00:00Z\g<2>", payload, count=1,
                )
            info = ZipInfo(filename=name, date_time=_ZIP_TIME)
            info.compress_type = ZIP_DEFLATED
            info.create_system = 3
            info.external_attr = 0o600 << 16
            target.writestr(info, payload, compress_type=ZIP_DEFLATED, compresslevel=9)
    return target_buffer.getvalue()


def _render_value(value: Any) -> Any:
    if value is None:
        return _NOT_AVAILABLE
    if type(value) is str:
        return _escape_formula(value[:30000])
    if type(value) in {bool, int, float}:
        return value
    return _escape_formula(str(_scalar(value))[:30000])


def _render_workbook(
    *, fields: tuple[WorkbookFieldDefinition, ...],
    sheets: tuple[WorkbookSheetDefinition, ...], rows: tuple[WorkbookRowRecord, ...],
    source_snapshot_ids: Mapping[str, str],
) -> bytes:
    workbook = Workbook()
    workbook.remove(workbook.active)
    workbook.properties.creator = "Amazon Product Intelligence"
    workbook.properties.lastModifiedBy = "Amazon Product Intelligence"
    workbook.properties.title = "Amazon Product Intelligence Workbook V0.2"
    workbook.properties.subject = "运营展示层工作簿（可审计）"
    workbook.properties.description = "Deterministic nine-sheet operator presentation; no analysis recomputation."
    workbook.properties.keywords = "amazon,operator,evidence,lineage,v0.2"
    workbook.properties.category = "Operator Workbook"
    workbook.properties.created = _FIXED_DOCUMENT_TIME
    workbook.properties.modified = _FIXED_DOCUMENT_TIME
    workbook.calculation.calcMode = "manual"
    workbook.calculation.fullCalcOnLoad = False
    workbook.calculation.forceFullCalc = False
    workbook.custom_doc_props.append(openpyxl.packaging.custom.StringProperty(
        name="RulesetVersion", value=OPERATOR_WORKBOOK_RULESET_VERSION
    ))
    workbook.custom_doc_props.append(openpyxl.packaging.custom.StringProperty(
        name="SourceOutputSnapshotId", value=source_snapshot_ids["operator_output"]
    ))

    field_by_id = {item.field_id: item for item in fields}
    row_by_id = {item.row_id: item for item in rows}
    navy = PatternFill("solid", fgColor="17365D")
    blue = PatternFill("solid", fgColor="4472C4")
    pale = PatternFill("solid", fgColor="FFF2CC")
    alternate = PatternFill("solid", fgColor="F4F7FB")
    risk_fill = PatternFill("solid", fgColor="FCE4D6")
    thin = Side(style="thin", color="D9E2F3")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    for sheet in sheets:
        worksheet = workbook.create_sheet(sheet.sheet_name)
        sheet_fields = tuple(field_by_id[item] for item in sheet.field_ids)
        last_col = get_column_letter(len(sheet_fields))
        worksheet.merge_cells(f"A1:{last_col}1")
        worksheet["A1"] = f"{sheet.sheet_name}｜{sheet.purpose}"
        worksheet["A1"].fill = navy
        worksheet["A1"].font = Font(name="Microsoft YaHei", size=15, bold=True, color="FFFFFF")
        worksheet["A1"].alignment = Alignment(vertical="center")
        worksheet.merge_cells(f"A2:{last_col}2")
        worksheet["A2"] = f"提示：{sheet.warning}"
        worksheet["A2"].fill = pale
        worksheet["A2"].font = Font(name="Microsoft YaHei", size=10, bold=True, color="7F6000")
        worksheet["A2"].alignment = Alignment(vertical="center", wrap_text=True)
        worksheet.row_dimensions[1].height = 28
        worksheet.row_dimensions[2].height = 34
        for col, field in enumerate(sheet_fields, start=1):
            cell = worksheet.cell(row=3, column=col, value=field.chinese_name)
            cell.fill = blue
            cell.font = Font(name="Microsoft YaHei", size=10, bold=True, color="FFFFFF")
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.border = border
            dimension = worksheet.column_dimensions[get_column_letter(col)]
            dimension.width = field.column_width
            if field.default_hidden and not sheet.hidden:
                dimension.hidden = True
        worksheet.row_dimensions[3].height = 34
        row_heights = {
            "market_overview": 62,
            "product_database": 112,
            "top_products": 112,
            "keyword_demand": 72,
            "competition_evidence": 62,
            "product_structure": 92,
            "opportunity_analysis": 108,
            "action_recommendations": 108,
            "data_audit": 72,
        }
        for row_number, row_id in enumerate(sheet.row_ids, start=4):
            row = row_by_id[row_id]
            for col, field in enumerate(sheet_fields, start=1):
                value = _render_value(row.values[field.field_id])
                cell = worksheet.cell(row=row_number, column=col, value=value)
                cell.font = Font(name="Microsoft YaHei", size=9, color="1F1F1F")
                cell.alignment = Alignment(vertical="top", wrap_text=True)
                cell.border = border
                if row_number % 2 == 0:
                    cell.fill = alternate
                if type(value) is str:
                    cell.number_format = "@"
            worksheet.row_dimensions[row_number].height = row_heights[sheet.sheet_key]
        worksheet.freeze_panes = "A4"
        worksheet.auto_filter.ref = f"A3:{last_col}{max(3, 3 + len(sheet.row_ids))}"
        worksheet.sheet_view.showGridLines = False
        worksheet.sheet_properties.pageSetUpPr.fitToPage = True
        worksheet.page_setup.fitToWidth = 1
        worksheet.page_setup.fitToHeight = 0
        if sheet.sheet_key == "action_recommendations":
            manual_index = next(
                index for index, item in enumerate(sheet_fields, start=1)
                if item.english_name == "Manual Review Status"
            )
            letter = get_column_letter(manual_index)
            validation = DataValidation(
                type="list",
                formula1='"未复核,进一步调研,补充数据,人工审核,已复核"',
                allow_blank=False,
            )
            validation.error = "请选择预设人工复核状态"
            validation.errorTitle = "无效状态"
            validation.prompt = "仅记录人工流程状态，不改变系统建议。"
            validation.promptTitle = "人工复核"
            worksheet.add_data_validation(validation)
            validation.add(f"{letter}4:{letter}{max(4, 3 + len(sheet.row_ids))}")
        if sheet.sheet_key in {"market_overview", "opportunity_analysis", "action_recommendations"}:
            worksheet.conditional_formatting.add(
                f"A4:{last_col}{max(4, 3 + len(sheet.row_ids))}",
                FormulaRule(formula=['ISNUMBER(SEARCH("RISK",A4))'], fill=risk_fill),
            )
        if sheet.hidden:
            worksheet.sheet_state = "hidden"
    raw = BytesIO()
    workbook.save(raw)
    return _normalize_zip(raw.getvalue())


class OperatorWorkbookBuilderV0_2:
    """Project validated upstream snapshots into the V0.2 operator workbook."""

    def build(self, request: OperatorWorkbookRequest) -> OperatorWorkbookSnapshotV0_2:
        if not isinstance(request, OperatorWorkbookRequest):
            raise OperatorWorkbookValidationError("request must be OperatorWorkbookRequest")
        fields, base_sheets = _definitions()
        fields_by_sheet = {
            sheet.sheet_key: tuple(item for item in fields if item.sheet_key == sheet.sheet_key)
            for sheet in base_sheets
        }
        plans = _display_plans(request)
        output_lineages = _records(request.operator_output_snapshot, "lineage_index")
        export_lineages = _records(request.operator_export_snapshot, "lineage_index")
        export_lineage_by_output = {
            str(item["source_output_lineage_id"]): item for item in export_lineages
        }
        if len(export_lineage_by_output) != len(export_lineages):
            raise OperatorWorkbookValidationError(
                "operator export contains duplicate output-lineage references"
            )
        lineages_by_row: dict[str, tuple[Mapping[str, Any], ...]] = {}
        for output_row_id in sorted({str(item["output_row_id"]) for item in output_lineages}):
            lineages_by_row[output_row_id] = tuple(sorted(
                (item for item in output_lineages if item["output_row_id"] == output_row_id),
                key=lambda item: str(item["output_lineage_id"]),
            ))

        rows: list[WorkbookRowRecord] = []
        lineages: list[WorkbookLineageReference] = []
        sheet_row_ids: dict[str, list[str]] = {item.sheet_key: [] for item in base_sheets}
        output_snapshot_id = str(request.operator_output_snapshot["snapshot_id"])
        export_snapshot_id = str(request.operator_export_snapshot["snapshot_id"])
        for sheet in base_sheets[:-1]:
            for row_key, values, output_row_ids, source_snapshot_ids in plans[sheet.sheet_key]:
                value_map = _value_map(sheet, fields_by_sheet, values)
                identity = {
                    "sheet_id": sheet.sheet_id, "row_key": row_key, "values": value_map,
                    "source_output_row_ids": tuple(sorted(output_row_ids)),
                    "source_snapshot_ids": tuple(sorted(source_snapshot_ids)),
                }
                row_id = deterministic_id("operator-workbook-row", identity)
                row_lineages = tuple(
                    _copy_lineage(
                        row_id=row_id, sheet_id=sheet.sheet_id,
                        export_snapshot_id=export_snapshot_id,
                        output_snapshot_id=output_snapshot_id, output_lineage=item,
                        export_lineage=export_lineage_by_output[str(item["output_lineage_id"])],
                    )
                    for output_row_id in sorted(output_row_ids)
                    for item in lineages_by_row.get(output_row_id, ())
                )
                if not row_lineages:
                    raise OperatorWorkbookValidationError(
                        f"display row {row_key} has no Operator Output lineage"
                    )
                row = WorkbookRowRecord(
                    row_id=row_id, lineage_reference_ids=tuple(
                        item.workbook_lineage_id for item in row_lineages
                    ), **identity,
                )
                rows.append(row)
                lineages.extend(row_lineages)
                sheet_row_ids[sheet.sheet_key].append(row_id)

        audit_sheet = base_sheets[-1]
        display_rows = tuple(rows)
        display_lineages = tuple(lineages)
        display_lineage_by_id = {item.workbook_lineage_id: item for item in display_lineages}
        source_sheet_by_id = {item.sheet_id: item for item in base_sheets}
        for display_row in display_rows:
            source_sheet = source_sheet_by_id[display_row.sheet_id]
            excel_row = 4 + sheet_row_ids[source_sheet.sheet_key].index(display_row.row_id)
            for display_lineage_id in display_row.lineage_reference_ids:
                source_lineage = display_lineage_by_id[display_lineage_id]
                audit_key = f"audit:{display_lineage_id}"
                audit_values = {
                    "Audit Record ID": audit_key, "Source Sheet": source_sheet.sheet_name,
                    "Display Row Key": display_row.row_key, "Excel Row": excel_row,
                    "Display Field": "ROW_LINEAGE", "Excel Cell": f"ROW:{excel_row}",
                    "Export Row ID": source_lineage.source_export_row_id,
                    "Output Row ID": source_lineage.source_output_row_id,
                    "Evidence ID": source_lineage.canonical_reference_id,
                    "Provider": source_lineage.provider, "Source Tool": source_lineage.source_tool,
                    "Source Field": source_lineage.source_field,
                    "Raw Evidence Reference": source_lineage.raw_evidence_id,
                    "Collection Run ID": source_lineage.collection_run_id,
                    "Transformation Run ID": source_lineage.transformation_run_id,
                    "Mapping Version": source_lineage.mapping_version,
                    "Canonical Reference ID": source_lineage.canonical_reference_id,
                    "Lineage ID": source_lineage.source_lineage_id,
                    "Source Snapshot ID": source_lineage.source_snapshot_id,
                    "Source Bundle Fingerprint": _join(source_lineage.source_bundle_fingerprints),
                }
                value_map = _value_map(audit_sheet, fields_by_sheet, audit_values)
                identity = {
                    "sheet_id": audit_sheet.sheet_id, "row_key": audit_key,
                    "values": value_map,
                    "source_output_row_ids": (source_lineage.source_output_row_id,),
                    "source_snapshot_ids": (source_lineage.source_snapshot_id,),
                }
                audit_row_id = deterministic_id("operator-workbook-row", identity)
                audit_lineage = _copy_lineage(
                    row_id=audit_row_id, sheet_id=audit_sheet.sheet_id,
                    export_snapshot_id=export_snapshot_id,
                    output_snapshot_id=output_snapshot_id,
                    output_lineage=next(
                        item for item in output_lineages
                        if item["output_lineage_id"] == source_lineage.source_output_lineage_id
                    ),
                    export_lineage=export_lineage_by_output[
                        source_lineage.source_output_lineage_id
                    ],
                )
                audit_row = WorkbookRowRecord(
                    row_id=audit_row_id,
                    lineage_reference_ids=(audit_lineage.workbook_lineage_id,),
                    **identity,
                )
                rows.append(audit_row)
                lineages.append(audit_lineage)
                sheet_row_ids[audit_sheet.sheet_key].append(audit_row_id)

        sheets = tuple(WorkbookSheetDefinition(
            sheet_id=sheet.sheet_id, ordinal=sheet.ordinal, sheet_key=sheet.sheet_key,
            sheet_name=sheet.sheet_name, purpose=sheet.purpose, warning=sheet.warning,
            row_grain=sheet.row_grain, hidden=sheet.hidden, field_ids=sheet.field_ids,
            row_ids=tuple(sheet_row_ids[sheet.sheet_key]),
        ) for sheet in base_sheets)
        rows_tuple = tuple(rows)
        lineage_tuple = tuple(lineages)
        source_ids = _source_ids(request)
        workbook_bytes = _render_workbook(
            fields=fields, sheets=sheets, rows=rows_tuple, source_snapshot_ids=source_ids
        )
        workbook_content = {
            "filename": WORKBOOK_FILENAME, "media_type": WORKBOOK_MEDIA_TYPE,
            "content_base64": base64.b64encode(workbook_bytes).decode("ascii"),
            "content_sha256": sha256(workbook_bytes).hexdigest(),
            "size_bytes": len(workbook_bytes),
            "sheet_ids": tuple(item.sheet_id for item in sheets),
            "metadata": {
                "format": "XLSX", "ruleset_version": OPERATOR_WORKBOOK_RULESET_VERSION,
                "renderer": "openpyxl", "renderer_version": openpyxl.__version__,
                "formula_escape": "LEADING_APOSTROPHE", "worksheet_count": 9,
                "source_output_snapshot_id": output_snapshot_id,
                "source_export_snapshot_id": export_snapshot_id,
                "deterministic_zip_timestamp": "1980-01-01T00:00:00",
            },
        }
        workbook = WorkbookFileRecord(
            workbook_id=deterministic_id("operator-workbook-file", workbook_content),
            **workbook_content,
        )
        diagnostic_content = {
            "code": "WORKBOOK_PRESENTATION_ONLY", "severity": "INFO",
            "message": (
                "Nine-sheet workbook reorganizes validated upstream evidence, score references, "
                "and recommendation records without recomputation or candidate resolution."
            ),
            "source_snapshot_ids": tuple(sorted(source_ids.values())),
        }
        diagnostics = (WorkbookDiagnostic(
            diagnostic_id=deterministic_id("operator-workbook-diagnostic", diagnostic_content),
            **diagnostic_content,
        ),)
        coverage = WorkbookCoverageSummary(
            sheet_count=9, field_count=len(fields),
            display_row_count=len(display_rows),
            audit_row_count=len(sheet_row_ids[audit_sheet.sheet_key]),
            lineage_reference_count=len(lineage_tuple),
            diagnostic_count=len(diagnostics),
            row_counts_by_sheet={sheet.sheet_name: len(sheet.row_ids) for sheet in sheets},
        )
        content = {
            "ruleset_version": OPERATOR_WORKBOOK_RULESET_VERSION,
            "source_bundle_fingerprints": tuple(sorted(
                _bundle_fingerprint(item) for item in request.canonical_bundles
            )),
            "source_snapshot_ids": source_ids, "fields": fields, "sheets": sheets,
            "rows": rows_tuple, "workbook": workbook, "coverage": coverage,
            "diagnostics": diagnostics, "lineage_index": lineage_tuple,
        }
        snapshot = OperatorWorkbookSnapshotV0_2(
            snapshot_id=deterministic_id("operator-workbook-snapshot", content),
            **content,
        )
        return snapshot.validate_against_bundles(request.canonical_bundles)


__all__ = ("OperatorWorkbookBuilderV0_2",)
