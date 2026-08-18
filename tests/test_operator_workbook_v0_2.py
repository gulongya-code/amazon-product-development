from __future__ import annotations

import ast
from dataclasses import FrozenInstanceError
from io import BytesIO
import json
import os
from pathlib import Path
import subprocess
import sys
import tempfile
import unittest
from zipfile import ZipFile

from openpyxl import load_workbook

import amazon_product_intelligence.operator_workbook as operator_workbook
from amazon_product_intelligence.operator_workbook import (
    OPERATOR_WORKBOOK_RULESET_VERSION,
    WORKBOOK_FILENAME,
    OperatorWorkbookBuilderV0_2,
    OperatorWorkbookRequest,
    OperatorWorkbookSerializationError,
    OperatorWorkbookSnapshotV0_2,
    OperatorWorkbookValidationError,
)
from amazon_product_intelligence.operator_workbook.schema_v0_2 import (
    EXPECTED_FIELD_COUNT,
    EXPECTED_SHEET_NAMES,
)
from amazon_product_intelligence.operator_export import (
    OperatorExportBuilderV0_1,
    OperatorExportRequest,
)
from tests.test_conflict_resolution_v0_1 import build_evaluation, synthetic_bundles
from tests.test_operator_output_v0_1 import build_fixture


EXPECTED_API = {
    "OPERATOR_WORKBOOK_RULESET_VERSION",
    "WORKBOOK_FILENAME",
    "OperatorWorkbookRequest",
    "OperatorWorkbookSnapshotV0_2",
    "OperatorWorkbookBuilderV0_2",
    "OperatorWorkbookError",
    "OperatorWorkbookValidationError",
    "OperatorWorkbookSerializationError",
    "WorkbookFieldDefinition",
    "WorkbookSheetDefinition",
    "WorkbookRowRecord",
    "WorkbookLineageReference",
    "WorkbookFileRecord",
    "WorkbookCoverageSummary",
    "WorkbookDiagnostic",
}


def build_workbook():
    bundles, sources, _, output = build_fixture()
    evaluation = build_evaluation(*synthetic_bundles())
    product, demand, competition, opportunity, scoring, recommendation = sources
    export = OperatorExportBuilderV0_1().build(OperatorExportRequest(
        canonical_bundles=bundles,
        operator_output_snapshot=output.to_dict(),
    ))
    request = OperatorWorkbookRequest(
        canonical_bundles=bundles,
        product_intelligence_snapshot=product.to_dict(),
        demand_intelligence_snapshot=demand.to_dict(),
        competition_intelligence_snapshot=competition.to_dict(),
        opportunity_intelligence_snapshot=opportunity.to_dict(),
        evidence_evaluation_snapshot=evaluation.to_dict(),
        opportunity_scoring_snapshot=scoring.to_dict(),
        recommendation_framework_snapshot=recommendation.to_dict(),
        operator_export_snapshot=export.to_dict(),
        operator_output_snapshot=output.to_dict(),
    )
    return bundles, request, OperatorWorkbookBuilderV0_2().build(request)


def reverse_mapping_keys(value):
    if isinstance(value, dict):
        return {key: reverse_mapping_keys(value[key]) for key in reversed(tuple(value))}
    if isinstance(value, list):
        return [reverse_mapping_keys(item) for item in value]
    return value


class OperatorWorkbookV02Tests(unittest.TestCase):
    @classmethod
    def setUpClass(cls):
        cls.bundles, cls.request, cls.snapshot = build_workbook()

    def workbook(self):
        return load_workbook(BytesIO(self.snapshot.to_xlsx_bytes()), data_only=False)

    def english_values(self, sheet_name):
        sheet = next(item for item in self.snapshot.sheets if item.sheet_name == sheet_name)
        fields = {item.field_id: item for item in self.snapshot.fields}
        row_by_id = {item.row_id: item for item in self.snapshot.rows}
        return tuple({
            fields[field_id].english_name: row_by_id[row_id].values[field_id]
            for field_id in sheet.field_ids
        } for row_id in sheet.row_ids)

    def test_public_api_is_exact(self):
        self.assertEqual(set(operator_workbook.__all__), EXPECTED_API)
        self.assertEqual(len(operator_workbook.__all__), 15)

    def test_ruleset_filename_and_real_xlsx(self):
        self.assertEqual(OPERATOR_WORKBOOK_RULESET_VERSION, "operator-workbook-v0.2")
        self.assertEqual(WORKBOOK_FILENAME, "amazon_product_analysis_v0.2.xlsx")
        self.assertEqual(self.snapshot.workbook.filename, WORKBOOK_FILENAME)
        content = self.snapshot.to_xlsx_bytes()
        self.assertTrue(content.startswith(b"PK"))
        with ZipFile(BytesIO(content)) as package:
            self.assertIn("xl/workbook.xml", package.namelist())

    def test_exact_nine_sheet_order_and_hidden_audit(self):
        workbook = self.workbook()
        self.assertEqual(tuple(workbook.sheetnames), EXPECTED_SHEET_NAMES)
        self.assertTrue(workbook["09_数据审计"].sheet_state == "hidden")
        self.assertTrue(all(workbook[name].sheet_state == "visible" for name in EXPECTED_SHEET_NAMES[:-1]))

    def test_exact_field_dictionary_and_chinese_headers(self):
        self.assertEqual(len(self.snapshot.fields), EXPECTED_FIELD_COUNT)
        workbook = self.workbook()
        fields = {item.field_id: item for item in self.snapshot.fields}
        for sheet in self.snapshot.sheets:
            expected = tuple(fields[item].chinese_name for item in sheet.field_ids)
            actual = tuple(
                workbook[sheet.sheet_name].cell(row=3, column=index).value
                for index in range(1, len(expected) + 1)
            )
            self.assertEqual(actual, expected)
            self.assertTrue(all(type(item) is str and item for item in actual))

    def test_title_warning_freeze_filters_widths_and_wrap(self):
        workbook = self.workbook()
        fields = {item.field_id: item for item in self.snapshot.fields}
        for sheet in self.snapshot.sheets:
            worksheet = workbook[sheet.sheet_name]
            self.assertIn(sheet.purpose, worksheet["A1"].value)
            self.assertIn(sheet.warning, worksheet["A2"].value)
            self.assertEqual(worksheet.freeze_panes, "A4")
            self.assertIsNotNone(worksheet.auto_filter.ref)
            self.assertTrue(worksheet["A2"].alignment.wrap_text)
            for index, field_id in enumerate(sheet.field_ids, start=1):
                letter = worksheet.cell(row=3, column=index).column_letter
                self.assertEqual(worksheet.column_dimensions[letter].width, fields[field_id].column_width)

    def test_default_hidden_audit_columns_on_visible_sheets(self):
        workbook = self.workbook()
        fields = {item.field_id: item for item in self.snapshot.fields}
        for sheet in self.snapshot.sheets[:-1]:
            worksheet = workbook[sheet.sheet_name]
            for index, field_id in enumerate(sheet.field_ids, start=1):
                letter = worksheet.cell(row=3, column=index).column_letter
                self.assertEqual(
                    bool(worksheet.column_dimensions[letter].hidden),
                    fields[field_id].default_hidden,
                )

    def test_business_fields_are_populated_without_candidate_resolution(self):
        product = self.english_values("02_产品数据库")[0]
        self.assertEqual(product["ASIN"], "B0G2VV4RBW")
        self.assertEqual(product["Marketplace"], "US")
        self.assertEqual(product["Brand"], "SKLSSVF")
        self.assertEqual(product["Category"], "VALVE")
        self.assertEqual(product["Price"], 18.99)
        self.assertIn(product["Rating State"], {"MULTIPLE_CANDIDATES", "ONE_DISTINCT_PRESENT_VALUE"})
        if product["Rating State"] == "MULTIPLE_CANDIDATES":
            self.assertIn("MULTIPLE_CANDIDATES", product["Rating"])

    def test_keyword_metrics_are_source_values_with_states(self):
        row = self.english_values("04_关键词需求分析")[0]
        self.assertEqual(row["Keyword"], "plastic spoons")
        self.assertEqual(row["Search Volume"], 41910)
        self.assertEqual(row["CPC"], 2.74)
        self.assertEqual(row["ABA Rank"], 2922)
        self.assertEqual(row["Difficulty"], 63)
        self.assertEqual(
            set(row["Estimate Method Status"].split(" | ")),
            {"DOCUMENTED", "PARTIALLY_DOCUMENTED", "UNKNOWN"},
        )

    def test_top_product_does_not_invent_rank(self):
        row = self.english_values("03_TOP产品分析")[0]
        self.assertEqual(row["Source Rank Value"], "NOT_AVAILABLE")
        self.assertEqual(row["Rank Status"], "NOT_AVAILABLE")
        self.assertIn("NOT_BEST_PRODUCT", row["Data Limitations"])

    def test_competition_is_relationship_evidence_not_ranking(self):
        rows = self.english_values("05_市场竞争证据")
        self.assertEqual(len(rows), 10)
        self.assertTrue(all(item["Evidence Count"] > 0 for item in rows))
        self.assertTrue(all(item["Observed Relationship"] == "OBSERVED_PRODUCT_KEYWORD_RELATIONSHIP" for item in rows))
        all_text = json.dumps(rows, ensure_ascii=False)
        self.assertNotIn("competitor ranking", all_text.lower())

    def test_opportunity_only_reproduces_score_references(self):
        rows = self.english_values("07_机会分析")
        source_ids = {item["calculation_id"] for item in self.request.opportunity_scoring_snapshot["calculations"]}
        self.assertEqual({item["Score Reference"] for item in rows}, source_ids)
        self.assertTrue(all(item["Demand Signal"] != "NOT_AVAILABLE" for item in rows))
        self.assertTrue(all(item["Competition Signal"] != "NOT_AVAILABLE" for item in rows))
        self.assertTrue(all("OBSERVED_SIGNAL" in item["Signal Classification"] for item in rows))
        self.assertTrue(all("NO_SUCCESS_PROBABILITY" in item["Limitations"] for item in rows))

    def test_recommendations_are_existing_records_and_manual_status_is_editable(self):
        rows = self.english_values("08_行动建议")
        source_types = {item["recommendation_type"] for item in self.request.operator_output_snapshot["recommendation_rows"]}
        self.assertEqual({item["Recommendation Type"] for item in rows}, source_types)
        self.assertTrue(all(item["Reason"] != "NOT_AVAILABLE" for item in rows))
        self.assertTrue(all(item["Policy Status"] == "POLICY_REFERENCES_PRESENT" for item in rows))
        self.assertTrue(all(item["Manual Review Status"] == "未复核" for item in rows))
        workbook = self.workbook()
        self.assertEqual(len(workbook["08_行动建议"].data_validations.dataValidation), 1)

    def test_no_formula_cells_and_formula_escape_helper(self):
        workbook = self.workbook()
        self.assertEqual([
            cell.coordinate
            for worksheet in workbook.worksheets
            for row in worksheet.iter_rows()
            for cell in row if cell.data_type == "f"
        ], [])
        from amazon_product_intelligence.operator_workbook.builder_v0_2 import _escape_formula
        for value in ("=1+1", "+cmd", " -2", " @x"):
            self.assertTrue(_escape_formula(value).startswith("'"))

    def test_workbook_security_scan_has_no_sensitive_or_raw_payload_fields(self):
        forbidden = (
            "access_token", "api_key", "credential", "password", "private_key",
            "raw api payload", "raw_provider_payload", "secret", "authorization",
        )
        workbook = self.workbook()
        text = "\n".join(
            str(cell.value)
            for worksheet in workbook.worksheets
            for row in worksheet.iter_rows()
            for cell in row if cell.value is not None
        ).lower()
        for marker in forbidden:
            self.assertNotIn(marker, text)

    def test_coverage_and_row_counts_are_exact(self):
        self.assertEqual(self.snapshot.coverage.sheet_count, 9)
        self.assertEqual(self.snapshot.coverage.field_count, 157)
        self.assertEqual(self.snapshot.coverage.display_row_count, 23)
        self.assertEqual(self.snapshot.coverage.audit_row_count, 1037)
        self.assertEqual(sum(self.snapshot.coverage.row_counts_by_sheet.values()), len(self.snapshot.rows))

    def test_every_display_row_has_output_and_canonical_lineage(self):
        audit_sheet_id = self.snapshot.sheets[-1].sheet_id
        lineage = {item.workbook_lineage_id: item for item in self.snapshot.lineage_index}
        for row in self.snapshot.rows:
            if row.sheet_id == audit_sheet_id:
                continue
            self.assertTrue(row.source_output_row_ids)
            self.assertTrue(row.lineage_reference_ids)
            for lineage_id in row.lineage_reference_ids:
                item = lineage[lineage_id]
                self.assertIn(item.source_output_row_id, row.source_output_row_ids)
                self.assertTrue(item.canonical_reference_id)
                self.assertTrue(item.raw_evidence_id)

    def test_audit_rows_reproduce_display_lineage(self):
        audit_rows = self.english_values("09_数据审计")
        self.assertEqual(len(audit_rows), self.snapshot.coverage.audit_row_count)
        export_row_ids = {
            item["export_row_id"]
            for item in self.request.operator_export_snapshot["rows"]
        }
        self.assertTrue(all(item["Export Row ID"] in export_row_ids for item in audit_rows))
        self.assertTrue(all(item["Output Row ID"] and item["Canonical Reference ID"] for item in audit_rows))
        self.assertTrue(all(
            item.source_export_snapshot_id
            == self.request.operator_export_snapshot["snapshot_id"]
            and item.source_export_row_id in export_row_ids
            and item.source_export_lineage_id
            for item in self.snapshot.lineage_index
        ))

    def test_lineage_validates_against_canonical_bundles(self):
        self.assertIs(self.snapshot.validate_against_bundles(self.bundles), self.snapshot)
        with self.assertRaises(OperatorWorkbookValidationError):
            self.snapshot.validate_against_bundles(self.bundles[:-1])

    def test_same_process_and_mapping_reorder_are_deterministic(self):
        rebuilt = OperatorWorkbookBuilderV0_2().build(self.request)
        payload = self.request.to_dict()
        reordered = OperatorWorkbookBuilderV0_2().build(
            OperatorWorkbookRequest.from_dict(reverse_mapping_keys(payload))
        )
        for candidate in (rebuilt, reordered):
            self.assertEqual(candidate.snapshot_id, self.snapshot.snapshot_id)
            self.assertEqual(candidate.workbook.content_sha256, self.snapshot.workbook.content_sha256)
            self.assertEqual(candidate.to_xlsx_bytes(), self.snapshot.to_xlsx_bytes())

    def test_fresh_process_is_deterministic(self):
        script = (
            "from tests.test_operator_workbook_v0_2 import build_workbook;"
            "s=build_workbook()[2];"
            "print(s.snapshot_id+'|'+s.workbook.content_sha256)"
        )
        environment = dict(os.environ)
        environment["PYTHONPATH"] = "src"
        actual = subprocess.check_output(
            [sys.executable, "-c", script],
            cwd=Path(__file__).resolve().parents[1], env=environment,
            text=True, encoding="utf-8",
        ).strip()
        self.assertEqual(actual, f"{self.snapshot.snapshot_id}|{self.snapshot.workbook.content_sha256}")

    def test_zip_and_document_timestamps_are_fixed(self):
        with ZipFile(BytesIO(self.snapshot.to_xlsx_bytes())) as package:
            self.assertEqual({item.date_time for item in package.infolist()}, {(1980, 1, 1, 0, 0, 0)})
            core = package.read("docProps/core.xml").decode("utf-8")
        self.assertIn("2000-01-01T00:00:00Z", core)
        self.assertNotIn("2026-", core)

    def test_strict_round_trip_unknown_fields_and_immutability(self):
        restored = OperatorWorkbookSnapshotV0_2.from_dict(self.snapshot.to_dict())
        self.assertEqual(restored.to_dict(), self.snapshot.to_dict())
        payload = self.snapshot.to_dict()
        payload["unknown"] = True
        with self.assertRaises(OperatorWorkbookSerializationError):
            OperatorWorkbookSnapshotV0_2.from_dict(payload)
        with self.assertRaises(TypeError):
            self.snapshot.workbook.metadata["format"] = "changed"
        with self.assertRaises(TypeError):
            self.snapshot.rows[0].values[next(iter(self.snapshot.rows[0].values))] = "changed"
        with self.assertRaises(FrozenInstanceError):
            self.snapshot.snapshot_id = "changed"

    def test_tampered_row_value_is_rejected(self):
        payload = self.snapshot.to_dict()
        first_key = next(iter(payload["rows"][0]["values"]))
        payload["rows"][0]["values"][first_key] = "tampered"
        with self.assertRaises(OperatorWorkbookSerializationError):
            OperatorWorkbookSnapshotV0_2.from_dict(payload)

    def test_real_file_can_be_written_once_with_fixed_name(self):
        with tempfile.TemporaryDirectory() as directory:
            target = Path(directory) / WORKBOOK_FILENAME
            self.assertEqual(self.snapshot.write_xlsx(target), target)
            self.assertEqual(target.read_bytes(), self.snapshot.to_xlsx_bytes())
            with self.assertRaises(OperatorWorkbookValidationError):
                self.snapshot.write_xlsx(target)
        with self.assertRaises(OperatorWorkbookValidationError):
            self.snapshot.write_xlsx("wrong.xlsx")

    def test_builder_rejects_wrong_request_type(self):
        with self.assertRaises(OperatorWorkbookValidationError):
            OperatorWorkbookBuilderV0_2().build(self.snapshot)

    def test_request_rejects_tampered_operator_export_lineage(self):
        payload = self.request.to_dict()
        payload["operator_export_snapshot"]["lineage_index"][0][
            "source_output_lineage_id"
        ] = "operator-output-lineage:missing"
        with self.assertRaises(OperatorWorkbookSerializationError):
            OperatorWorkbookRequest.from_dict(payload)

    def test_production_import_boundary_and_no_nondeterministic_calls(self):
        root = Path(__file__).resolve().parents[1]
        production = root / "src" / "amazon_product_intelligence" / "operator_workbook"
        forbidden_modules = {"random", "secrets", "time", "uuid"}
        forbidden_calls = {"eval", "exec", "hash", "repr"}
        third_party = set()
        allowed_stdlib = set(sys.stdlib_module_names) | {"__future__"}
        for path in sorted(production.glob("*.py")):
            tree = ast.parse(path.read_text(encoding="utf-8"), filename=str(path))
            for node in ast.walk(tree):
                candidates = []
                if isinstance(node, ast.Import):
                    candidates = [alias.name for alias in node.names]
                elif isinstance(node, ast.ImportFrom) and node.level == 0:
                    candidates = [node.module]
                for name in candidates:
                    root_name = name.split(".")[0]
                    self.assertNotIn(root_name, forbidden_modules)
                    if root_name not in allowed_stdlib and root_name != "amazon_product_intelligence":
                        third_party.add(root_name)
                if isinstance(node, ast.Call) and isinstance(node.func, ast.Name):
                    self.assertNotIn(node.func.id, forbidden_calls)
        self.assertEqual(third_party, {"openpyxl"})


if __name__ == "__main__":
    unittest.main()
