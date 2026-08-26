"""Deterministic, read-only XLSX audit for Operator Template Contract V1."""

from __future__ import annotations

from collections.abc import Iterable
from io import BytesIO
from pathlib import Path
from typing import BinaryIO

from openpyxl import load_workbook
from openpyxl.formula import Tokenizer

from .errors import OperatorTemplateContractValidationError
from .models import (
    DefinedNameAudit,
    FormulaCellAudit,
    FormulaDisposition,
    FormulaSheetAudit,
    OperatorTemplateContractV1,
    SheetRangeAudit,
    SheetStateAudit,
    SheetVisibility,
    ThresholdLiteralAudit,
    WorkbookTemplateAuditSnapshot,
    canonical_sha256,
    template_schema_fingerprint,
)
from .schema_v1 import TEMPLATE_CONTRACT_V1


WorkbookSource = str | Path | bytes | bytearray | BinaryIO


def _formula_text(value: object) -> str:
    if isinstance(value, str):
        return value
    text = getattr(value, "text", None)
    if isinstance(text, str):
        return text
    return str(value)


def _tokenize(formula: str) -> tuple[tuple[str, ...], tuple[tuple[int, str], ...]]:
    tokens = Tokenizer(formula).items
    signature = tuple(
        f"{token.type}:{token.subtype}:{token.value}" for token in tokens
    )
    numeric = tuple(
        (index, token.value)
        for index, token in enumerate(tokens)
        if token.type == "OPERAND" and token.subtype == "NUMBER"
    )
    return signature, numeric


def _policy_for(
    contract: OperatorTemplateContractV1,
    sheet_name: str,
    has_numeric_literal: bool,
) -> FormulaDisposition:
    policy = next(
        (item for item in contract.formula_policies if item.sheet_name == sheet_name),
        None,
    )
    if policy is None:
        return FormulaDisposition.REUSE_AS_FORMULA
    if has_numeric_literal and policy.numeric_literals_to_config:
        return FormulaDisposition.MOVE_TO_CONFIG
    return policy.disposition


def _defined_names(workbook: object) -> tuple[DefinedNameAudit, ...]:
    values: Iterable[object] = workbook.defined_names.values()
    result = tuple(
        DefinedNameAudit(
            name=str(item.name),
            refers_to=str(item.attr_text or ""),
            local_sheet_id=item.localSheetId,
        )
        for item in values
    )
    return tuple(
        sorted(
            result,
            key=lambda item: (
                item.name,
                -1 if item.local_sheet_id is None else item.local_sheet_id,
                item.refers_to,
            ),
        )
    )


def _sheet_ranges(workbook: object) -> tuple[SheetRangeAudit, ...]:
    result: list[SheetRangeAudit] = []
    for worksheet in workbook.worksheets:
        if worksheet.auto_filter.ref:
            result.append(
                SheetRangeAudit(
                    sheet_name=worksheet.title,
                    kind="AUTO_FILTER",
                    name=worksheet.title,
                    range_ref=str(worksheet.auto_filter.ref),
                )
            )
        for table in worksheet.tables.values():
            result.append(
                SheetRangeAudit(
                    sheet_name=worksheet.title,
                    kind="TABLE",
                    name=str(table.name),
                    range_ref=str(table.ref),
                )
            )
        for index, pivot in enumerate(getattr(worksheet, "_pivots", ()), 1):
            name = getattr(pivot, "name", None)
            if not name:
                name = f"pivot-{index}"
            location = getattr(pivot, "location", None)
            ref = getattr(location, "ref", "") if location is not None else ""
            result.append(
                SheetRangeAudit(
                    sheet_name=worksheet.title,
                    kind="PIVOT",
                    name=str(name),
                    range_ref=str(ref or "UNAVAILABLE"),
                )
            )
    return tuple(
        sorted(
            result,
            key=lambda item: (item.sheet_name, item.kind, item.name, item.range_ref),
        )
    )


def audit_workbook(
    source: WorkbookSource,
    *,
    contract: OperatorTemplateContractV1 = TEMPLATE_CONTRACT_V1,
) -> WorkbookTemplateAuditSnapshot:
    """Read an XLSX without evaluating or rewriting formulas."""

    if not isinstance(contract, OperatorTemplateContractV1):
        raise OperatorTemplateContractValidationError(
            "contract must be OperatorTemplateContractV1"
        )
    workbook_source: object = BytesIO(bytes(source)) if isinstance(
        source, (bytes, bytearray)
    ) else source
    workbook = load_workbook(
        filename=workbook_source,
        data_only=False,
        read_only=False,
        keep_links=True,
    )
    try:
        sheet_states = tuple(
            SheetStateAudit(name=sheet.title, state=str(sheet.sheet_state))
            for sheet in workbook.worksheets
        )
        raw_headers: tuple[str, ...] = ()
        if "原始数据源" in workbook.sheetnames:
            raw_sheet = workbook["原始数据源"]
            values = [cell.value for cell in raw_sheet[1]]
            while values and values[-1] is None:
                values.pop()
            raw_headers = tuple(
                value if isinstance(value, str) else "" if value is None else str(value)
                for value in values
            )

        formula_cells: list[FormulaCellAudit] = []
        threshold_literals: list[ThresholdLiteralAudit] = []
        formula_sheets: list[FormulaSheetAudit] = []
        for worksheet in workbook.worksheets:
            sheet_cells: list[FormulaCellAudit] = []
            for row in worksheet.iter_rows():
                for cell in row:
                    if cell.data_type != "f" and not (
                        isinstance(cell.value, str) and cell.value.startswith("=")
                    ):
                        continue
                    formula = _formula_text(cell.value)
                    signature, numeric = _tokenize(formula)
                    disposition = _policy_for(
                        contract, worksheet.title, bool(numeric)
                    )
                    record = FormulaCellAudit(
                        sheet_name=worksheet.title,
                        coordinate=cell.coordinate,
                        formula=formula,
                        token_signature=signature,
                        disposition=disposition,
                    )
                    sheet_cells.append(record)
                    for token_index, value in numeric:
                        threshold_literals.append(
                            ThresholdLiteralAudit(
                                sheet_name=worksheet.title,
                                coordinate=cell.coordinate,
                                token_index=token_index,
                                value=value,
                                disposition=(
                                    FormulaDisposition.MOVE_TO_CONFIG
                                    if disposition is FormulaDisposition.MOVE_TO_CONFIG
                                    else disposition
                                ),
                            )
                        )
            formula_cells.extend(sheet_cells)
            formula_sheets.append(
                FormulaSheetAudit(
                    sheet_name=worksheet.title,
                    formula_count=len(sheet_cells),
                    formula_fingerprint=canonical_sha256(sheet_cells),
                )
            )

        return WorkbookTemplateAuditSnapshot(
            schema_fingerprint=template_schema_fingerprint(contract),
            sheet_states=sheet_states,
            raw_headers=raw_headers,
            formula_cells=tuple(formula_cells),
            formula_sheets=tuple(formula_sheets),
            defined_names=_defined_names(workbook),
            sheet_ranges=_sheet_ranges(workbook),
            threshold_literals=tuple(threshold_literals),
        )
    finally:
        workbook.close()


def validate_workbook_audit(
    snapshot: WorkbookTemplateAuditSnapshot,
    *,
    contract: OperatorTemplateContractV1 = TEMPLATE_CONTRACT_V1,
) -> WorkbookTemplateAuditSnapshot:
    """Fail closed on sheet, visibility, header, or required dependency drift."""

    if not isinstance(snapshot, WorkbookTemplateAuditSnapshot):
        raise OperatorTemplateContractValidationError(
            "snapshot must be WorkbookTemplateAuditSnapshot"
        )
    violations: list[str] = []
    expected_names = tuple(sheet.name for sheet in contract.sheets)
    observed_names = tuple(sheet.name for sheet in snapshot.sheet_states)
    if len(observed_names) != len(expected_names) or set(observed_names) != set(expected_names):
        violations.append("workbook must contain exactly the contracted 15 sheets")

    expected_visible = tuple(
        sheet.name for sheet in contract.sheets
        if sheet.visibility is SheetVisibility.VISIBLE
    )
    observed_visible = tuple(
        sheet.name for sheet in snapshot.sheet_states if sheet.state == "visible"
    )
    if observed_visible != expected_visible:
        violations.append("visible sheet names/order mismatch")

    expected_hidden = {
        sheet.name for sheet in contract.sheets
        if sheet.visibility is SheetVisibility.HIDDEN
    }
    observed_hidden = {
        sheet.name for sheet in snapshot.sheet_states if sheet.state == "hidden"
    }
    if observed_hidden != expected_hidden:
        violations.append("hidden sheet set/state mismatch")
    invalid_states = tuple(
        sheet.name for sheet in snapshot.sheet_states
        if sheet.state not in {"visible", "hidden"}
    )
    if invalid_states:
        violations.append(
            "VeryHidden or unsupported sheet state is forbidden: "
            + ", ".join(invalid_states)
        )

    expected_headers = {header.name for header in contract.raw_headers}
    observed_headers = set(snapshot.raw_headers)
    if len(snapshot.raw_headers) != 66 or len(observed_headers) != 66:
        violations.append("原始数据源 must contain 66 unique header names")
    missing_headers = sorted(expected_headers - observed_headers)
    extra_headers = sorted(observed_headers - expected_headers)
    if missing_headers:
        violations.append("missing raw headers: " + ", ".join(missing_headers))
    if extra_headers:
        violations.append("unexpected raw headers: " + ", ".join(extra_headers))

    defined_names = {item.name for item in snapshot.defined_names}
    auto_filters = {
        item.sheet_name for item in snapshot.sheet_ranges
        if item.kind == "AUTO_FILTER"
    }
    for dependency in contract.dependencies:
        if not dependency.required:
            continue
        if dependency.kind == "NAMED_RANGE" and dependency.name not in defined_names:
            violations.append(f"required named range missing: {dependency.name}")
        if dependency.kind == "AUTO_FILTER" and dependency.sheet_name not in auto_filters:
            violations.append(
                f"required AutoFilter missing: {dependency.sheet_name}"
            )

    expected_fingerprint = template_schema_fingerprint(contract)
    if snapshot.schema_fingerprint != expected_fingerprint:
        violations.append("template schema fingerprint mismatch")
    if violations:
        raise OperatorTemplateContractValidationError("; ".join(violations))
    return snapshot


def audit_and_validate_workbook(
    source: WorkbookSource,
    *,
    contract: OperatorTemplateContractV1 = TEMPLATE_CONTRACT_V1,
) -> WorkbookTemplateAuditSnapshot:
    return validate_workbook_audit(
        audit_workbook(source, contract=contract), contract=contract
    )


__all__ = (
    "WorkbookSource",
    "audit_and_validate_workbook",
    "audit_workbook",
    "validate_workbook_audit",
)
