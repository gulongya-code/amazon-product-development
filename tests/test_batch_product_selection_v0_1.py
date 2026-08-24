from __future__ import annotations

from copy import deepcopy
from hashlib import sha256
from io import StringIO
import json
import os
from pathlib import Path
from tempfile import TemporaryDirectory
import unittest
from unittest.mock import patch
import zipfile

from openpyxl import load_workbook

import amazon_product_intelligence.batch_product_selection as batch_api
from amazon_product_intelligence.batch_product_selection import (
    BATCH_INPUT_CONTRACT_VERSION,
    BATCH_RANKING_STATUS,
    BATCH_RESULT_CONTRACT_VERSION,
    BatchCandidateDefinition,
    BatchInputValidationError,
    BatchProductSelectionOrchestrator,
    BatchSelectionError,
    BatchSelectionRequest,
    BatchStatus,
    BatchSummaryExcelRenderer,
    BatchSummaryMarkdownRenderer,
    CandidateExecutionSource,
    CandidateRecoveryDisposition,
    parse_batch_request,
)
from amazon_product_intelligence.connectors import (
    BoundedTransientRetryPolicy,
    ProviderConfig,
    ProviderRegistry,
    TransportResponse,
    XiYouProvider,
)
from amazon_product_intelligence.contracts import canonical_json
from amazon_product_intelligence.market_report.delivery import OperatorReportDeliveryResult
from amazon_product_intelligence.operator_workflow import OPERATOR_WORKFLOW_RULESET_VERSION
from amazon_product_intelligence.production_pipeline import (
    ProductionPipelineOrchestrator,
    ProductionRunMode,
    ProviderCreditSemantics,
)
from amazon_product_intelligence.production_pipeline.cli import main
from amazon_product_intelligence.production_pipeline.orchestrator import ProviderRuntime
from amazon_product_intelligence.production_pipeline.providers import (
    FixtureTransport,
    RecordingTransport,
)


ROOT = Path(__file__).resolve().parents[1]
FIXTURE_PATH = (
    ROOT
    / "src"
    / "amazon_product_intelligence"
    / "production_pipeline"
    / "fixtures"
    / "dog_water_bottle_v0_1.json"
)
ASINS = ("B0DWB00001", "B0DWB00002", "B0DWB00003")


def directory_hashes(path: Path) -> dict[str, str]:
    return {
        item.relative_to(path).as_posix(): sha256(item.read_bytes()).hexdigest()
        for item in sorted(path.rglob("*"))
        if item.is_file()
    }


class DeterministicCandidateDelivery:
    def deliver(self, source, output_directory, *, operator_workflow=None):
        output = Path(output_directory)
        output.mkdir(parents=True, exist_ok=True)
        markdown = output / "operator_market_report.md"
        xlsx = output / "operator_market_report.xlsx"
        markdown.write_text(
            f"# Candidate fixture\n\n{operator_workflow.operator_action.value}\n",
            encoding="utf-8",
        )
        with zipfile.ZipFile(xlsx, "w") as package:
            package.writestr("[Content_Types].xml", "<Types/>")
        return OperatorReportDeliveryResult(
            source_report_id=source.report_id,
            delivery_version="operator-market-report-delivery-v0.1",
            xlsx_path=xlsx,
            markdown_path=markdown,
            xlsx_sha256=sha256(xlsx.read_bytes()).hexdigest(),
            markdown_sha256=sha256(markdown.read_bytes()).hexdigest(),
            operator_workflow_id=operator_workflow.snapshot_id,
        )


class DeterministicBatchDelivery:
    def deliver(self, result, output_directory):
        output = Path(output_directory)
        json_path = output / "batch_selection_result.json"
        markdown = output / "batch_selection_summary.md"
        xlsx = output / "batch_selection_summary.xlsx"
        markdown.write_text(
            "# Batch Operator Brief\n\n"
            + "\n".join(
                f"{item.candidate_id}: {item.operator_action or 'UNAVAILABLE'}"
                for item in result.candidates
            )
            + "\n",
            encoding="utf-8",
        )
        with zipfile.ZipFile(xlsx, "w") as package:
            package.writestr("[Content_Types].xml", "<Types/>")
        json_path.write_text(
            json.dumps(
                result.to_dict(),
                ensure_ascii=False,
                allow_nan=False,
                sort_keys=True,
                indent=2,
            )
            + "\n",
            encoding="utf-8",
        )


class FaultInjectingFixtureTransport:
    def __init__(self, fixture, outcomes=None):
        self.fixture = FixtureTransport(fixture)
        self.outcomes = {key: list(value) for key, value in (outcomes or {}).items()}
        self.calls: list[str] = []
        self.network_call_count = 0

    def execute(self, request):
        key = (
            request.operation
            if request.operation == "asin_info"
            else f"{request.operation}:{request.parameters.get('asin')}"
        )
        self.calls.append(key)
        planned = self.outcomes.get(key)
        if planned:
            outcome = planned.pop(0)
            if isinstance(outcome, Exception):
                raise outcome
            return outcome
        return self.fixture.execute(request)


def response(status: int, credits: float | None = None) -> TransportResponse:
    metadata = {} if credits is None else {"cost_credits": credits}
    return TransportResponse(status_code=status, payload={}, metadata=metadata)


def fixture_runtime_factory(transport):
    def factory(_request):
        metadata = json.loads(FIXTURE_PATH.read_text(encoding="utf-8"))
        recording = RecordingTransport(transport)
        provider = XiYouProvider(
            recording,
            environment={"SP038_FIXTURE_SECRET": "sp038-fixture-secret"},
            retry_policy=BoundedTransientRetryPolicy(),
        )
        registry = ProviderRegistry()
        registry.register(
            provider,
            ProviderConfig(
                provider_id="xiyou",
                enabled=True,
                priority=1,
                credential_env="SP038_FIXTURE_SECRET",
                timeout_seconds=1.0,
                max_attempts=2,
            ),
        )
        return ProviderRuntime(
            registry=registry,
            provider=provider,
            recording_transport=recording,
            metadata=metadata,
            credit_semantics=ProviderCreditSemantics.FIXTURE_REFERENCE,
        )

    return factory


def batch_payload() -> dict:
    return {
        "contract_version": BATCH_INPUT_CONTRACT_VERSION,
        "batch_id": "fixture-batch-001",
        "marketplace": "us",
        "category_name": "dog water bottle",
        "mode": "fixture",
        "provider_preference": "xiyou",
        "provider_config_reference": "environment",
        "candidates": [
            {"candidate_id": "candidate-b", "asins": [ASINS[1].lower()]},
            {"candidate_id": "candidate-a", "asins": [ASINS[0]]},
        ],
    }


def three_candidate_request(root: Path, name: str, *, resume_from: Path | None = None):
    return BatchSelectionRequest(
        batch_id="recovery-batch-001",
        marketplace="US",
        category_name="dog water bottle",
        mode=ProductionRunMode.FIXTURE,
        candidates=tuple(
            BatchCandidateDefinition(candidate_id=f"candidate-{letter}", asins=(asin,))
            for letter, asin in zip(("a", "b", "c"), ASINS, strict=True)
        ),
        output_directory=root / name,
        resume_from=resume_from,
    )


class BatchProductSelectionContractTests(unittest.TestCase):
    def setUp(self) -> None:
        self.temp = TemporaryDirectory(dir=ROOT / "outputs")
        self.addCleanup(self.temp.cleanup)
        self.root = Path(self.temp.name)

    def test_public_contract_and_normalized_deterministic_input(self) -> None:
        request = parse_batch_request(
            batch_payload(), output_directory=self.root / "first"
        )
        reversed_payload = batch_payload()
        reversed_payload["candidates"] = list(reversed(reversed_payload["candidates"]))
        second = parse_batch_request(
            reversed_payload, output_directory=self.root / "second"
        )
        self.assertEqual("US", request.marketplace)
        self.assertEqual(("candidate-a", "candidate-b"), tuple(
            item.candidate_id for item in request.candidates
        ))
        self.assertEqual((ASINS[1],), request.candidates[1].asins)
        self.assertEqual(request.input_fingerprint, second.input_fingerprint)
        self.assertEqual("batch-selection-result-v0.1", BATCH_RESULT_CONTRACT_VERSION)
        self.assertEqual("UNAVAILABLE", BATCH_RANKING_STATUS)
        self.assertIn("BatchProductSelectionOrchestrator", batch_api.__all__)
        self.assertIn("CandidateRecoveryDisposition", batch_api.__all__)

    def test_invalid_duplicate_discovery_and_path_unsafe_inputs_are_rejected(self) -> None:
        cases = []
        duplicate_asin = batch_payload()
        duplicate_asin["candidates"][0]["asins"] = [ASINS[1], ASINS[1].lower()]
        cases.append(duplicate_asin)
        duplicate_cohort = batch_payload()
        duplicate_cohort["candidates"][1]["asins"] = [ASINS[1]]
        cases.append(duplicate_cohort)
        unsafe = batch_payload()
        unsafe["candidates"][0]["candidate_id"] = "../candidate-b"
        cases.append(unsafe)
        discovery = batch_payload()
        discovery["seed_keyword"] = "dog bottle"
        cases.append(discovery)
        for payload in cases:
            with self.subTest(payload=payload):
                with self.assertRaises(BatchInputValidationError):
                    parse_batch_request(payload, output_directory=self.root / "invalid")

        invalid_file = self.root / "invalid-batch.json"
        invalid_file.write_text(json.dumps(unsafe), encoding="utf-8")

        class ForbiddenBatchOrchestrator:
            calls = 0

            def run(self, _request):
                self.calls += 1
                raise AssertionError("invalid batch must not execute")

        forbidden = ForbiddenBatchOrchestrator()
        stdout, stderr = StringIO(), StringIO()
        code = main(
            [
                "batch",
                "--batch-file",
                str(invalid_file),
                "--output-dir",
                str(self.root / "invalid-cli-output"),
            ],
            batch_orchestrator=forbidden,
            stdout=stdout,
            stderr=stderr,
        )
        self.assertEqual(2, code)
        self.assertEqual(0, forbidden.calls)
        self.assertEqual("", stdout.getvalue())
        self.assertIn("INVALID_BATCH_INPUT", stderr.getvalue())

    def test_output_ownership_conflict_precedes_pipeline_factory(self) -> None:
        destination = self.root / "owned"
        destination.mkdir()
        stale = destination / "batch_selection_result.json"
        stale.write_text('{"owner":"older-batch"}\n', encoding="utf-8")
        original = sha256(stale.read_bytes()).hexdigest()
        calls: list[str] = []

        def forbidden(candidate_id):
            calls.append(candidate_id)
            raise AssertionError("pipeline factory must not run")

        request = parse_batch_request(batch_payload(), output_directory=destination)
        with self.assertRaises(BatchSelectionError) as caught:
            BatchProductSelectionOrchestrator(pipeline_factory=forbidden).run(request)
        self.assertEqual("BATCH_OUTPUT_CONFLICT", caught.exception.code.value)
        self.assertEqual([], calls)
        self.assertEqual(original, sha256(stale.read_bytes()).hexdigest())


class BatchProductSelectionFixtureAcceptanceTests(unittest.TestCase):
    @classmethod
    def setUpClass(cls) -> None:
        if not os.environ.get("MARKET_REPORT_NODE_EXECUTABLE") or not os.environ.get(
            "MARKET_REPORT_NODE_MODULES"
        ):
            raise unittest.SkipTest("artifact-tool runtime paths are not configured")
        cls.temp = TemporaryDirectory(dir=ROOT / "outputs")
        cls.root = Path(cls.temp.name)
        cls.output = cls.root / "acceptance"
        cls.request = parse_batch_request(
            batch_payload(), output_directory=cls.output
        )
        with patch(
            "amazon_product_intelligence.production_pipeline.orchestrator.HttpJsonTransport.execute",
            side_effect=AssertionError("fixture batch must not access HTTP"),
        ) as http_execute:
            cls.result = BatchProductSelectionOrchestrator().run(cls.request)
            cls.http_calls = http_execute.call_count
        cls.payload = json.loads(
            (cls.output / "batch_selection_result.json").read_text(encoding="utf-8")
        )
        cls.markdown = (cls.output / "batch_selection_summary.md").read_text(
            encoding="utf-8"
        )
        cls.workbook = load_workbook(
            cls.output / "batch_selection_summary.xlsx",
            read_only=True,
            data_only=False,
        )

    @classmethod
    def tearDownClass(cls) -> None:
        if hasattr(cls, "workbook"):
            cls.workbook.close()
        if hasattr(cls, "temp"):
            cls.temp.cleanup()

    def test_candidate_pipeline_reuse_artifacts_and_aggregate_outputs(self) -> None:
        self.assertIs(self.result.status, BatchStatus.SUCCEEDED)
        self.assertEqual(0, self.http_calls)
        self.assertEqual(2, self.result.succeeded_count)
        self.assertEqual(0, self.result.failed_count)
        for candidate in self.result.candidates:
            self.assertIs(candidate.execution_source, CandidateExecutionSource.NEW_EXECUTION)
            self.assertEqual(
                {"market_report_json", "operator_xlsx", "operator_markdown", "run_manifest"},
                set(candidate.artifact_paths),
            )
            self.assertTrue(all(Path(path).is_file() for path in candidate.artifact_paths.values()))
        self.assertTrue(all(Path(path).is_file() for path in self.result.batch_artifact_paths.values()))

    def test_aggregate_action_status_and_null_semantics_match_candidate_workflows(self) -> None:
        for candidate in self.result.candidates:
            manifest = json.loads(
                Path(candidate.artifact_paths["run_manifest"]).read_text(encoding="utf-8")
            )
            report = json.loads(
                Path(candidate.artifact_paths["market_report_json"]).read_text(encoding="utf-8")
            )
            workflow = manifest["operator_workflow"]
            self.assertEqual(workflow["operator_action"], candidate.operator_action)
            self.assertEqual(workflow["recommendation_type"], candidate.recommendation_type)
            self.assertEqual(workflow["evidence_readiness"], candidate.evidence_readiness)
            self.assertEqual(
                OPERATOR_WORKFLOW_RULESET_VERSION,
                candidate.operator_workflow_ruleset_version,
            )
            self.assertEqual("PENDING_DATA", candidate.opportunity_score_status)
            self.assertIsNone(candidate.opportunity_score_value)
            self.assertIsNone(report["opportunity_score"]["score_value"])
            self.assertEqual(BATCH_RANKING_STATUS, candidate.ranking_status)
            self.assertIn(candidate.operator_action, self.markdown)
        self.assertNotIn('"opportunity_score_value": 0', canonical_json(self.payload))
        self.assertIn("value=null", self.markdown)

    def test_xlsx_markdown_json_executive_views_agree(self) -> None:
        self.assertEqual("# Batch Operator Brief", self.markdown.splitlines()[0])
        self.assertEqual(
            ["Batch Summary", "Candidate Actions", "Evidence Gaps", "Run Health", "Audit Lineage"],
            self.workbook.sheetnames,
        )
        summary = self.workbook["Batch Summary"]
        rows = {
            summary.cell(row=row, column=1).value: {
                "status": summary.cell(row=row, column=2).value,
                "action": summary.cell(row=row, column=3).value,
                "opportunity": summary.cell(row=row, column=7).value,
                "ranking": summary.cell(row=row, column=8).value,
                "recovery": summary.cell(row=row, column=15).value,
            }
            for row in range(5, 5 + self.result.candidate_count)
        }
        for candidate in self.result.candidates:
            self.assertEqual(candidate.production_run_status, rows[candidate.candidate_id]["status"])
            self.assertEqual(candidate.operator_action, rows[candidate.candidate_id]["action"])
            self.assertEqual("null (PENDING_DATA)", rows[candidate.candidate_id]["opportunity"])
            self.assertEqual("UNAVAILABLE", rows[candidate.candidate_id]["ranking"])
            self.assertEqual(
                "No recovery action is required.",
                rows[candidate.candidate_id]["recovery"],
            )
        repeated = self.root / "repeat-batch-summary.xlsx"
        BatchSummaryExcelRenderer().render(self.result, repeated)
        self.assertEqual(
            sha256((self.output / "batch_selection_summary.xlsx").read_bytes()).hexdigest(),
            sha256(repeated.read_bytes()).hexdigest(),
        )

    def test_usage_is_fixture_reference_not_billed_and_secret_safe(self) -> None:
        self.assertEqual(4, self.result.usage.total_logical_operations)
        self.assertEqual(4, self.result.usage.new_transport_attempts)
        self.assertEqual(4.0, self.result.usage.current_run_observed_credits)
        self.assertEqual("FIXTURE_REFERENCE", self.result.usage.credit_semantics)
        self.assertIn("not billed", self.result.usage.billing_note)
        for path in self.output.rglob("*"):
            if path.is_file():
                content = path.read_bytes().lower()
                self.assertNotIn(b"sp038-fixture-secret", content)
                self.assertNotIn(b"x-api-key", content)
                self.assertNotIn(b"authorization", content)

    def test_batch_cli_success_contract(self) -> None:
        class ReturningOrchestrator:
            def __init__(self, result):
                self.result = result
                self.requests = []

            def run(self, request):
                self.requests.append(request)
                return self.result

        returning = ReturningOrchestrator(self.result)
        source = self.root / "cli-batch.json"
        source.write_text(json.dumps(batch_payload()), encoding="utf-8")
        stdout, stderr = StringIO(), StringIO()
        code = main(
            ["batch", "--batch-file", str(source), "--output-dir", str(self.root / "unused")],
            batch_orchestrator=returning,
            stdout=stdout,
            stderr=stderr,
        )
        self.assertEqual(0, code)
        self.assertEqual(1, len(returning.requests))
        self.assertIn("FIXTURE_REFERENCE", stdout.getvalue())
        self.assertIn("not billed", stdout.getvalue())
        self.assertEqual("", stderr.getvalue())


class BatchProductSelectionRecoveryTests(unittest.TestCase):
    def setUp(self) -> None:
        self.temp = TemporaryDirectory(dir=ROOT / "outputs")
        self.addCleanup(self.temp.cleanup)
        self.root = Path(self.temp.name)
        self.fixture = json.loads(FIXTURE_PATH.read_text(encoding="utf-8"))

    def pipeline_factory(self, transports, *, fault_candidate=None):
        def factory(candidate_id):
            outcomes = {}
            if candidate_id == fault_candidate:
                outcomes[f"asin_keywords:{ASINS[1]}"] = [
                    response(503, 0.25),
                    response(503, 0.25),
                ]
            transport = FaultInjectingFixtureTransport(self.fixture, outcomes)
            transports[candidate_id] = transport
            return ProductionPipelineOrchestrator(
                provider_runtime_factory=fixture_runtime_factory(transport),
                delivery=DeterministicCandidateDelivery(),
            )

        return factory

    def test_fault_isolated_partial_then_resume_uses_sp036_and_matches_uninterrupted(self) -> None:
        first_transports = {}
        source_request = three_candidate_request(self.root, "partial-source")
        with patch(
            "amazon_product_intelligence.connectors.transport.urlopen",
            side_effect=AssertionError("network access attempted"),
        ):
            partial = BatchProductSelectionOrchestrator(
                pipeline_factory=self.pipeline_factory(
                    first_transports, fault_candidate="candidate-b"
                ),
                delivery=DeterministicBatchDelivery(),
            ).run(source_request)
        self.assertIs(partial.status, BatchStatus.PARTIAL)
        self.assertEqual(("candidate-a", "candidate-b", "candidate-c"), tuple(first_transports))
        failed = next(item for item in partial.candidates if item.candidate_id == "candidate-b")
        self.assertEqual("FAILED", failed.production_run_status)
        self.assertEqual({"run_manifest"}, set(failed.artifact_paths))
        failed_directory = self.root / "partial-source" / "candidates" / "candidate-b"
        self.assertFalse((failed_directory / "market_report.json").exists())
        self.assertFalse((failed_directory / "operator_market_report.xlsx").exists())
        self.assertFalse((failed_directory / "operator_market_report.md").exists())
        partial_markdown = BatchSummaryMarkdownRenderer().render(partial)
        self.assertIn("# Batch Operator Brief", partial_markdown)
        self.assertIn("candidate-b", partial_markdown)
        self.assertIn("UNAVAILABLE — candidate failed", partial_markdown)
        self.assertIn("Failures Requiring Rerun or Resume", partial_markdown)
        self.assertIn("Checkpoint resume available", partial_markdown)
        self.assertIs(
            failed.recovery_disposition,
            CandidateRecoveryDisposition.CHECKPOINT_RESUME_AVAILABLE,
        )
        if os.environ.get("MARKET_REPORT_NODE_EXECUTABLE") and os.environ.get(
            "MARKET_REPORT_NODE_MODULES"
        ):
            partial_xlsx = self.root / "partial-visible.xlsx"
            BatchSummaryExcelRenderer().render(partial, partial_xlsx)
            workbook = load_workbook(partial_xlsx, read_only=False, data_only=False)
            self.addCleanup(workbook.close)
            summary = workbook["Batch Summary"]
            candidate_rows = {
                summary.cell(row=row, column=1).value: row
                for row in range(5, 5 + partial.candidate_count)
            }
            failed_row = candidate_rows["candidate-b"]
            self.assertEqual("FAILED", summary.cell(row=failed_row, column=2).value)
            self.assertEqual(
                "UNAVAILABLE — candidate failed",
                summary.cell(row=failed_row, column=3).value,
            )
            self.assertTrue(
                str(summary.cell(row=failed_row, column=2).fill.fgColor.rgb).endswith(
                    "FCE4D6"
                )
            )
            self.assertIn(
                "Checkpoint resume available",
                summary.cell(row=failed_row, column=15).value,
            )
            health = workbook["Run Health"]
            health_rows = {
                health.cell(row=row, column=1).value: row
                for row in range(5, 5 + partial.candidate_count)
            }
            self.assertEqual(
                "CHECKPOINT_RESUME_AVAILABLE",
                health.cell(row=health_rows["candidate-b"], column=4).value,
            )

        class ReturningPartial:
            def run(self, _request):
                return partial

        batch_file = self.root / "partial-cli-input.json"
        payload = batch_payload()
        batch_file.write_text(json.dumps(payload), encoding="utf-8")
        stdout, stderr = StringIO(), StringIO()
        exit_code = main(
            [
                "batch",
                "--batch-file",
                str(batch_file),
                "--output-dir",
                str(self.root / "partial-cli-unused"),
            ],
            batch_orchestrator=ReturningPartial(),
            stdout=stdout,
            stderr=stderr,
        )
        self.assertEqual(1, exit_code)
        self.assertEqual("", stdout.getvalue())
        self.assertIn("partial", stderr.getvalue())

        source_hashes = directory_hashes(self.root / "partial-source")
        resumed_transports = {}
        resumed = BatchProductSelectionOrchestrator(
            pipeline_factory=self.pipeline_factory(resumed_transports),
            delivery=DeterministicBatchDelivery(),
        ).run(
            three_candidate_request(
                self.root,
                "resumed",
                resume_from=self.root / "partial-source",
            )
        )
        self.assertIs(resumed.status, BatchStatus.SUCCEEDED)
        self.assertEqual(["candidate-b"], list(resumed_transports))
        self.assertEqual(
            [f"asin_keywords:{ASINS[1]}"], resumed_transports["candidate-b"].calls
        )
        self.assertEqual(source_hashes, directory_hashes(self.root / "partial-source"))
        self.assertEqual(1, resumed.usage.new_transport_attempts)
        self.assertEqual(1, resumed.usage.executed_operations)
        self.assertEqual(1, resumed.usage.checkpoint_replayed_operations)
        self.assertEqual(4, resumed.usage.reused_source_operations)
        self.assertEqual(1.0, resumed.usage.current_run_observed_credits)
        resumed_by_id = {item.candidate_id: item for item in resumed.candidates}
        self.assertIs(
            resumed_by_id["candidate-a"].execution_source,
            CandidateExecutionSource.REUSED_SUCCESS,
        )
        self.assertIs(
            resumed_by_id["candidate-b"].execution_source,
            CandidateExecutionSource.CHECKPOINT_RESUME,
        )
        self.assertIs(
            resumed_by_id["candidate-b"].recovery_disposition,
            CandidateRecoveryDisposition.NOT_APPLICABLE,
        )
        self.assertTrue(
            resumed_by_id["candidate-a"].artifact_paths["run_manifest"].startswith(
                str((self.root / "partial-source").resolve())
            )
        )

        uninterrupted_transports = {}
        uninterrupted = BatchProductSelectionOrchestrator(
            pipeline_factory=self.pipeline_factory(uninterrupted_transports),
            delivery=DeterministicBatchDelivery(),
        ).run(three_candidate_request(self.root, "uninterrupted"))
        self.assertEqual(
            uninterrupted.semantic_fingerprint,
            resumed.semantic_fingerprint,
        )
        for expected, actual in zip(
            uninterrupted.candidates, resumed.candidates, strict=True
        ):
            self.assertEqual(expected.operator_semantic_fingerprint, actual.operator_semantic_fingerprint)
            self.assertEqual(expected.operator_action, actual.operator_action)
            self.assertEqual(expected.next_actions, actual.next_actions)
        resumed_b = resumed_by_id["candidate-b"]
        uninterrupted_b = next(
            item for item in uninterrupted.candidates if item.candidate_id == "candidate-b"
        )
        self.assertNotEqual(uninterrupted_b.run_health, resumed_b.run_health)

    def test_resume_fingerprint_mismatch_fails_before_pipeline_and_source_is_immutable(self) -> None:
        transports = {}
        source = BatchProductSelectionOrchestrator(
            pipeline_factory=self.pipeline_factory(transports),
            delivery=DeterministicBatchDelivery(),
        ).run(three_candidate_request(self.root, "source-success"))
        self.assertIs(source.status, BatchStatus.SUCCEEDED)
        before = directory_hashes(self.root / "source-success")
        mismatched = BatchSelectionRequest(
            batch_id="different-batch",
            marketplace="US",
            category_name="dog water bottle",
            mode=ProductionRunMode.FIXTURE,
            candidates=three_candidate_request(self.root, "unused").candidates,
            output_directory=self.root / "mismatch-destination",
            resume_from=self.root / "source-success",
        )
        calls: list[str] = []

        def forbidden(candidate_id):
            calls.append(candidate_id)
            raise AssertionError("pipeline factory must not run")

        with self.assertRaises(BatchSelectionError) as caught:
            BatchProductSelectionOrchestrator(
                pipeline_factory=forbidden,
                delivery=DeterministicBatchDelivery(),
            ).run(mismatched)
        self.assertEqual("INCOMPATIBLE_BATCH_RESUME", caught.exception.code.value)
        self.assertEqual([], calls)
        self.assertEqual(before, directory_hashes(self.root / "source-success"))
        self.assertFalse((self.root / "mismatch-destination").exists())

        corrupted = (
            self.root
            / "source-success"
            / "candidates"
            / "candidate-a"
            / "operator_market_report.md"
        )
        corrupted.write_text("tampered after source completion\n", encoding="utf-8")
        matching = three_candidate_request(
            self.root,
            "integrity-destination",
            resume_from=self.root / "source-success",
        )
        with self.assertRaises(BatchSelectionError) as integrity:
            BatchProductSelectionOrchestrator(
                pipeline_factory=forbidden,
                delivery=DeterministicBatchDelivery(),
            ).run(matching)
        self.assertEqual("BATCH_ARTIFACT_INTEGRITY", integrity.exception.code.value)
        self.assertEqual([], calls)
        self.assertFalse((self.root / "integrity-destination").exists())

    def test_three_generation_chained_resume_reuses_origins_and_sp036_work(self) -> None:
        first_transports = {}
        with patch(
            "amazon_product_intelligence.connectors.transport.urlopen",
            side_effect=AssertionError("network access attempted"),
        ):
            first = BatchProductSelectionOrchestrator(
                pipeline_factory=self.pipeline_factory(
                    first_transports, fault_candidate="candidate-b"
                ),
                delivery=DeterministicBatchDelivery(),
            ).run(three_candidate_request(self.root, "chain-batch-1"))
            first_hashes = directory_hashes(self.root / "chain-batch-1")

            second_transports = {}
            second = BatchProductSelectionOrchestrator(
                pipeline_factory=self.pipeline_factory(
                    second_transports, fault_candidate="candidate-b"
                ),
                delivery=DeterministicBatchDelivery(),
            ).run(
                three_candidate_request(
                    self.root,
                    "chain-batch-2",
                    resume_from=self.root / "chain-batch-1",
                )
            )
            second_hashes = directory_hashes(self.root / "chain-batch-2")

            third_transports = {}
            third = BatchProductSelectionOrchestrator(
                pipeline_factory=self.pipeline_factory(third_transports),
                delivery=DeterministicBatchDelivery(),
            ).run(
                three_candidate_request(
                    self.root,
                    "chain-batch-3",
                    resume_from=self.root / "chain-batch-2",
                )
            )

        self.assertIs(first.status, BatchStatus.PARTIAL)
        self.assertIs(second.status, BatchStatus.PARTIAL)
        self.assertIs(third.status, BatchStatus.SUCCEEDED)
        self.assertEqual(first_hashes, directory_hashes(self.root / "chain-batch-1"))
        self.assertEqual(second_hashes, directory_hashes(self.root / "chain-batch-2"))
        self.assertEqual(["candidate-b"], list(second_transports))
        self.assertEqual(["candidate-b"], list(third_transports))
        self.assertEqual(
            [f"asin_keywords:{ASINS[1]}", f"asin_keywords:{ASINS[1]}"],
            second_transports["candidate-b"].calls,
        )
        self.assertEqual(
            [f"asin_keywords:{ASINS[1]}"],
            third_transports["candidate-b"].calls,
        )
        second_by_id = {item.candidate_id: item for item in second.candidates}
        third_by_id = {item.candidate_id: item for item in third.candidates}
        for candidate_id in ("candidate-a", "candidate-c"):
            self.assertIs(
                second_by_id[candidate_id].execution_source,
                CandidateExecutionSource.REUSED_SUCCESS,
            )
            self.assertIs(
                third_by_id[candidate_id].execution_source,
                CandidateExecutionSource.REUSED_SUCCESS,
            )
            self.assertTrue(
                third_by_id[candidate_id].artifact_paths["run_manifest"].startswith(
                    str((self.root / "chain-batch-1").resolve())
                )
            )
            self.assertEqual(
                second_by_id[candidate_id].artifact_hashes,
                third_by_id[candidate_id].artifact_hashes,
            )
        self.assertIs(
            second_by_id["candidate-b"].execution_source,
            CandidateExecutionSource.CHECKPOINT_RESUME,
        )
        self.assertIs(
            second_by_id["candidate-b"].recovery_disposition,
            CandidateRecoveryDisposition.CHECKPOINT_RESUME_AVAILABLE,
        )
        self.assertIs(
            third_by_id["candidate-b"].execution_source,
            CandidateExecutionSource.CHECKPOINT_RESUME,
        )
        self.assertEqual(1, third.usage.new_transport_attempts)
        self.assertEqual(1, third.usage.executed_operations)
        self.assertEqual(1, third.usage.checkpoint_replayed_operations)
        self.assertEqual(4, third.usage.reused_source_operations)
        self.assertEqual(1.0, third.usage.current_run_observed_credits)
        self.assertEqual("FIXTURE_REFERENCE", third.usage.credit_semantics)
        self.assertIn("not billed", third.usage.billing_note)

        uninterrupted_transports = {}
        uninterrupted = BatchProductSelectionOrchestrator(
            pipeline_factory=self.pipeline_factory(uninterrupted_transports),
            delivery=DeterministicBatchDelivery(),
        ).run(three_candidate_request(self.root, "chain-uninterrupted"))
        self.assertEqual(uninterrupted.semantic_fingerprint, third.semantic_fingerprint)
        for expected, actual in zip(
            uninterrupted.candidates, third.candidates, strict=True
        ):
            self.assertEqual(
                expected.operator_semantic_fingerprint,
                actual.operator_semantic_fingerprint,
            )
            self.assertEqual(expected.operator_action, actual.operator_action)
            self.assertEqual(expected.next_actions, actual.next_actions)

    def test_batch_layer_failure_without_safe_manifest_fresh_runs_next_batch(self) -> None:
        first_transports = {}
        normal_factory = self.pipeline_factory(first_transports)

        def batch_layer_fault(candidate_id):
            if candidate_id == "candidate-b":
                raise RuntimeError("deterministic batch-layer fault before production state")
            return normal_factory(candidate_id)

        first = BatchProductSelectionOrchestrator(
            pipeline_factory=batch_layer_fault,
            delivery=DeterministicBatchDelivery(),
        ).run(three_candidate_request(self.root, "fresh-batch-1"))
        first_hashes = directory_hashes(self.root / "fresh-batch-1")
        first_b = next(item for item in first.candidates if item.candidate_id == "candidate-b")
        self.assertIs(first.status, BatchStatus.PARTIAL)
        self.assertIs(first_b.execution_source, CandidateExecutionSource.NEW_EXECUTION)
        self.assertIs(
            first_b.recovery_disposition,
            CandidateRecoveryDisposition.FRESH_EXECUTION_REQUIRED,
        )
        self.assertEqual({}, dict(first_b.artifact_paths))
        self.assertIn(
            "No safe checkpoint is available",
            BatchSummaryMarkdownRenderer().render(first),
        )
        if os.environ.get("MARKET_REPORT_NODE_EXECUTABLE") and os.environ.get(
            "MARKET_REPORT_NODE_MODULES"
        ):
            first_xlsx = self.root / "fresh-rerun-guidance.xlsx"
            BatchSummaryExcelRenderer().render(first, first_xlsx)
            workbook = load_workbook(first_xlsx, read_only=True, data_only=False)
            self.addCleanup(workbook.close)
            summary = workbook["Batch Summary"]
            candidate_rows = {
                summary.cell(row=row, column=1).value: row
                for row in range(5, 5 + first.candidate_count)
            }
            self.assertIn(
                "requires a fresh execution",
                summary.cell(
                    row=candidate_rows["candidate-b"], column=15
                ).value,
            )

        resumed_transports = {}
        resumed = BatchProductSelectionOrchestrator(
            pipeline_factory=self.pipeline_factory(resumed_transports),
            delivery=DeterministicBatchDelivery(),
        ).run(
            three_candidate_request(
                self.root,
                "fresh-batch-2",
                resume_from=self.root / "fresh-batch-1",
            )
        )
        self.assertIs(resumed.status, BatchStatus.SUCCEEDED)
        self.assertEqual(["candidate-b"], list(resumed_transports))
        resumed_b = next(
            item for item in resumed.candidates if item.candidate_id == "candidate-b"
        )
        self.assertIs(resumed_b.execution_source, CandidateExecutionSource.NEW_EXECUTION)
        self.assertEqual(
            ["asin_info", f"asin_keywords:{ASINS[1]}"],
            resumed_transports["candidate-b"].calls,
        )
        self.assertEqual(first_hashes, directory_hashes(self.root / "fresh-batch-1"))

    def test_checkpoint_attempt_exception_records_truthful_lineage_then_fresh_runs(self) -> None:
        first_transports = {}
        first = BatchProductSelectionOrchestrator(
            pipeline_factory=self.pipeline_factory(
                first_transports, fault_candidate="candidate-b"
            ),
            delivery=DeterministicBatchDelivery(),
        ).run(three_candidate_request(self.root, "truthful-batch-1"))
        self.assertIs(first.status, BatchStatus.PARTIAL)

        calls: list[str] = []

        def resume_layer_fault(candidate_id):
            calls.append(candidate_id)
            raise RuntimeError("deterministic exception during checkpoint attempt")

        second = BatchProductSelectionOrchestrator(
            pipeline_factory=resume_layer_fault,
            delivery=DeterministicBatchDelivery(),
        ).run(
            three_candidate_request(
                self.root,
                "truthful-batch-2",
                resume_from=self.root / "truthful-batch-1",
            )
        )
        self.assertEqual(["candidate-b"], calls)
        second_b = next(
            item for item in second.candidates if item.candidate_id == "candidate-b"
        )
        self.assertIs(
            second_b.execution_source, CandidateExecutionSource.CHECKPOINT_RESUME
        )
        self.assertIs(
            second_b.recovery_disposition,
            CandidateRecoveryDisposition.FRESH_EXECUTION_REQUIRED,
        )
        self.assertEqual({}, dict(second_b.artifact_paths))

        third_transports = {}
        third = BatchProductSelectionOrchestrator(
            pipeline_factory=self.pipeline_factory(third_transports),
            delivery=DeterministicBatchDelivery(),
        ).run(
            three_candidate_request(
                self.root,
                "truthful-batch-3",
                resume_from=self.root / "truthful-batch-2",
            )
        )
        third_b = next(
            item for item in third.candidates if item.candidate_id == "candidate-b"
        )
        self.assertIs(third.status, BatchStatus.SUCCEEDED)
        self.assertIs(third_b.execution_source, CandidateExecutionSource.NEW_EXECUTION)
        self.assertEqual(
            ["asin_info", f"asin_keywords:{ASINS[1]}"],
            third_transports["candidate-b"].calls,
        )

    def test_chained_artifact_tampering_fails_closed_before_provider_factory(self) -> None:
        def build_chain(suffix):
            first_transports = {}
            BatchProductSelectionOrchestrator(
                pipeline_factory=self.pipeline_factory(
                    first_transports, fault_candidate="candidate-b"
                ),
                delivery=DeterministicBatchDelivery(),
            ).run(three_candidate_request(self.root, f"integrity-{suffix}-1"))
            second_transports = {}
            BatchProductSelectionOrchestrator(
                pipeline_factory=self.pipeline_factory(
                    second_transports, fault_candidate="candidate-b"
                ),
                delivery=DeterministicBatchDelivery(),
            ).run(
                three_candidate_request(
                    self.root,
                    f"integrity-{suffix}-2",
                    resume_from=self.root / f"integrity-{suffix}-1",
                )
            )
            return (
                self.root / f"integrity-{suffix}-1",
                self.root / f"integrity-{suffix}-2",
            )

        first, second = build_chain("artifact")
        artifact = first / "candidates" / "candidate-a" / "operator_market_report.md"
        artifact.write_text("tampered chained artifact\n", encoding="utf-8")
        calls: list[str] = []

        def forbidden(candidate_id):
            calls.append(candidate_id)
            raise AssertionError("integrity must fail before pipeline construction")

        with self.assertRaises(BatchSelectionError) as tampered:
            BatchProductSelectionOrchestrator(
                pipeline_factory=forbidden,
                delivery=DeterministicBatchDelivery(),
            ).run(
                three_candidate_request(
                    self.root,
                    "integrity-artifact-3",
                    resume_from=second,
                )
            )
        self.assertEqual("BATCH_ARTIFACT_INTEGRITY", tampered.exception.code.value)
        self.assertEqual([], calls)

        first, second = build_chain("path")
        batch_result_path = second / "batch_selection_result.json"
        payload = json.loads(batch_result_path.read_text(encoding="utf-8"))
        candidate_a = next(
            item for item in payload["candidates"] if item["candidate_id"] == "candidate-a"
        )
        candidate_c = next(
            item for item in payload["candidates"] if item["candidate_id"] == "candidate-c"
        )
        candidate_a["artifact_paths"]["operator_markdown"] = candidate_c[
            "artifact_paths"
        ]["operator_markdown"]
        candidate_a["artifact_hashes"]["operator_markdown"] = candidate_c[
            "artifact_hashes"
        ]["operator_markdown"]
        batch_result_path.write_text(
            json.dumps(payload, ensure_ascii=False, sort_keys=True, indent=2) + "\n",
            encoding="utf-8",
        )
        with self.assertRaises(BatchSelectionError) as path_tampered:
            BatchProductSelectionOrchestrator(
                pipeline_factory=forbidden,
                delivery=DeterministicBatchDelivery(),
            ).run(
                three_candidate_request(
                    self.root,
                    "integrity-path-3",
                    resume_from=second,
                )
            )
        self.assertEqual(
            "BATCH_ARTIFACT_INTEGRITY", path_tampered.exception.code.value
        )
        self.assertEqual([], calls)

        first, second = build_chain("missing")
        missing = first / "candidates" / "candidate-a" / "operator_market_report.md"
        missing.unlink()
        with self.assertRaises(BatchSelectionError) as missing_artifact:
            BatchProductSelectionOrchestrator(
                pipeline_factory=forbidden,
                delivery=DeterministicBatchDelivery(),
            ).run(
                three_candidate_request(
                    self.root,
                    "integrity-missing-3",
                    resume_from=second,
                )
            )
        self.assertEqual(
            "BATCH_ARTIFACT_INTEGRITY", missing_artifact.exception.code.value
        )
        self.assertEqual([], calls)


if __name__ == "__main__":
    unittest.main()
