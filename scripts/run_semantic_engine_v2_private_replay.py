"""Run the mandatory privacy-safe five-category TASK-SP-041S2 replay.

The manifest and optional operator-review rows are private external inputs.  This
script emits aggregate JSON only and refuses to publish listing-grain content.
"""

from __future__ import annotations

import argparse
from collections import Counter
from dataclasses import dataclass, replace
import json
from pathlib import Path
import re
import time
from typing import Any

from amazon_product_intelligence.semantic_engine_v2 import (
    ConsumptionLifecycle,
    RelationRole,
    SemanticEngineV2Error,
    SemanticScope,
    UniversalSemanticRole,
    build_semantic_engine_v2_result,
    load_category_semantic_profile,
)
from amazon_product_intelligence.sellersprite_import import (
    ImportContext,
    import_sellersprite_file,
)


EXPECTED_COUNTS = {
    "CAL_SHOWER_CADDY": 998,
    "CAL_DOG_WATER_BOTTLE": 400,
    "CAL_VACUUM_FILTER": 300,
    "CAL_FOOD_STORAGE_SET": 150,
    "CAL_AIR_FRYER_MIXED": 300,
}
CORE_FLOORS = {
    "CAL_SHOWER_CADDY": (
        (UniversalSemanticRole.INSTALLATION_ARCHITECTURE, 0.842),
    ),
    "CAL_VACUUM_FILTER": (
        (UniversalSemanticRole.COMPATIBILITY, 0.770),
    ),
    "CAL_FOOD_STORAGE_SET": (
        (UniversalSemanticRole.STRUCTURAL_FORM, 0.033),
    ),
    "CAL_AIR_FRYER_MIXED": (
        (UniversalSemanticRole.STRUCTURAL_FORM, 0.373),
    ),
}
BOUNDARY_TAGS = frozenset((
    "OBVIOUS_OTHER", "NONPRIMARY", "USE_CASE_ONLY", "INVALID_QUANTITY",
    "AMBIGUOUS_UNIT", "HOST_CAPACITY_ACCESSORY",
))
_MANIFEST_KEYS = frozenset((
    "calibration_id", "input", "profile", "marketplace", "category",
    "observed_date", "sheet",
))
_OPERATOR_KEYS = frozenset((
    "calibration_id", "listing_reference", "expected_relation_role",
    "expected_primary_cohort", "boundary_tags",
))

_OPERATOR_SHEET = "01_中文辅助审核"
_OPERATOR_HEADERS = frozenset((
    "校准类目", "ASIN", "我的市场范围建议", "我的商品角色建议",
    "你的决定", "你修改的市场范围", "你修改的商品角色",
    "最终市场范围", "最终商品角色",
))
_CALIBRATION_IDS_BY_LABEL = {
    "淋浴置物架": "CAL_SHOWER_CADDY",
    "狗狗便携饮水瓶": "CAL_DOG_WATER_BOTTLE",
    "吸尘器滤芯/过滤系统": "CAL_VACUUM_FILTER",
    "食品收纳盒套装": "CAL_FOOD_STORAGE_SET",
    "空气炸锅": "CAL_AIR_FRYER_MIXED",
}
_ACCEPT_DECISIONS = frozenset(("接受我的判断", "ACCEPT"))
_MODIFY_DECISIONS = frozenset(("修改", "MODIFY"))
_REVIEW_DECISIONS = frozenset(("需要和我讨论", "REVIEW", "REVIEW_REQUIRED"))
_ROLE_LABELS = {
    "主商品": RelationRole.PRIMARY_PRODUCT,
    "配件": RelationRole.ACCESSORY,
    "替换件": RelationRole.REPLACEMENT,
    "补充/填充品": RelationRole.REFILL,
    "UNKNOWN": RelationRole.UNKNOWN,
    "未知": RelationRole.UNKNOWN,
    "REVIEW_REQUIRED": RelationRole.REVIEW_REQUIRED,
    "待讨论": RelationRole.REVIEW_REQUIRED,
}
_CONSUMABLE_LABEL = "耗材"
_MARKET_SCOPE_LABELS = {
    "核心目标商品": (True, ()),
    "相关目标商品": (False, ()),
    "目标周边配件/替换/耗材": (False, ("NONPRIMARY",)),
    # The bounded S1 OTHER_PRODUCT cohort is also the only authoritative
    # off-target/use-case boundary available in the original raw workbook.
    # USE_CASE_ONLY measures target-identity inclusion below, not cohort entry.
    "其他商品": (False, ("OBVIOUS_OTHER", "USE_CASE_ONLY")),
}


@dataclass(frozen=True, slots=True)
class OperatorReviewInput:
    """Private operator labels plus aggregate-only input-quality counters."""

    rows: tuple[dict[str, Any], ...]
    raw_row_count: int
    valid_decision_row_count: int
    malformed_decision_excluded_count: int
    malformed_relation_label_excluded_count: int
    consumable_lifecycle_label_count: int
    source_format: str


def _load_json(path: Path) -> Any:
    try:
        return json.loads(path.read_text(encoding="utf-8"))
    except (OSError, UnicodeError, json.JSONDecodeError) as exc:
        raise SemanticEngineV2Error("PRIVATE_REPLAY_INPUT_INVALID", "private JSON is unreadable") from exc


def _manifest(path: Path) -> tuple[dict[str, Any], ...]:
    payload = _load_json(path)
    if not isinstance(payload, list) or len(payload) != len(EXPECTED_COUNTS):
        raise SemanticEngineV2Error("PRIVATE_REPLAY_INPUT_INVALID", "manifest must contain five entries")
    entries: list[dict[str, Any]] = []
    for raw in payload:
        if not isinstance(raw, dict) or set(raw) != _MANIFEST_KEYS:
            raise SemanticEngineV2Error("PRIVATE_REPLAY_INPUT_INVALID", "manifest schema mismatch")
        if raw["calibration_id"] not in EXPECTED_COUNTS:
            raise SemanticEngineV2Error("PRIVATE_REPLAY_INPUT_INVALID", "unknown calibration ID")
        if any(not isinstance(raw[name], str) or not raw[name].strip() for name in (
            "input", "profile", "marketplace", "category",
        )):
            raise SemanticEngineV2Error("PRIVATE_REPLAY_INPUT_INVALID", "manifest text is blank")
        if raw["observed_date"] is not None and not isinstance(raw["observed_date"], str):
            raise SemanticEngineV2Error("PRIVATE_REPLAY_INPUT_INVALID", "observed_date is invalid")
        if raw["sheet"] is not None and not isinstance(raw["sheet"], str):
            raise SemanticEngineV2Error("PRIVATE_REPLAY_INPUT_INVALID", "sheet is invalid")
        entries.append(raw)
    if {item["calibration_id"] for item in entries} != set(EXPECTED_COUNTS):
        raise SemanticEngineV2Error("PRIVATE_REPLAY_INPUT_INVALID", "calibration IDs must be unique")
    return tuple(sorted(entries, key=lambda item: item["calibration_id"]))


def _cell_text(value: Any) -> str | None:
    if not isinstance(value, str):
        return None
    normalized = value.strip()
    return normalized or None


def _json_operator_rows(path: Path) -> OperatorReviewInput:
    payload = _load_json(path)
    if not isinstance(payload, list):
        raise SemanticEngineV2Error("PRIVATE_REPLAY_INPUT_INVALID", "operator review must be a list")
    rows: list[dict[str, Any]] = []
    identities: set[tuple[str, str]] = set()
    for raw in payload:
        if not isinstance(raw, dict) or set(raw) != _OPERATOR_KEYS:
            raise SemanticEngineV2Error("PRIVATE_REPLAY_INPUT_INVALID", "operator schema mismatch")
        if raw["calibration_id"] not in EXPECTED_COUNTS:
            raise SemanticEngineV2Error("PRIVATE_REPLAY_INPUT_INVALID", "operator category is invalid")
        if not isinstance(raw["listing_reference"], str) or not raw["listing_reference"].strip():
            raise SemanticEngineV2Error("PRIVATE_REPLAY_INPUT_INVALID", "operator listing reference is blank")
        if raw["expected_relation_role"] is not None:
            RelationRole(raw["expected_relation_role"])
        if raw["expected_primary_cohort"] is not None and type(raw["expected_primary_cohort"]) is not bool:
            raise SemanticEngineV2Error("PRIVATE_REPLAY_INPUT_INVALID", "operator cohort label is invalid")
        tags = raw["boundary_tags"]
        if not isinstance(tags, list) or any(item not in BOUNDARY_TAGS for item in tags):
            raise SemanticEngineV2Error("PRIVATE_REPLAY_INPUT_INVALID", "operator boundary tag is invalid")
        identity = (raw["calibration_id"], raw["listing_reference"])
        if identity in identities:
            raise SemanticEngineV2Error("PRIVATE_REPLAY_INPUT_INVALID", "operator row is duplicated")
        identities.add(identity)
        rows.append(raw)
    unlabeled = sum(item["expected_relation_role"] is None for item in rows)
    return OperatorReviewInput(
        rows=tuple(rows), raw_row_count=len(rows), valid_decision_row_count=len(rows),
        malformed_decision_excluded_count=0,
        malformed_relation_label_excluded_count=unlabeled,
        consumable_lifecycle_label_count=0, source_format="JSON",
    )


def _legacy_relation_role(
    effective_label: str | None, suggested_label: str | None,
) -> tuple[RelationRole | None, ConsumptionLifecycle | None]:
    """Project the frozen one-column S1 labels into the orthogonal V1.1 role.

    A consumable override changes lifecycle, not the already proposed relation.
    The original reviewed rows whose proposal itself was ``耗材`` are the frozen
    paper/liner pattern and therefore project to ACCESSORY + CONSUMABLE.  No
    listing title or other private cell is inspected to manufacture a label.
    """

    if effective_label in _ROLE_LABELS:
        return _ROLE_LABELS[effective_label], None
    if effective_label != _CONSUMABLE_LABEL:
        return None, None
    if suggested_label == _CONSUMABLE_LABEL:
        relation = RelationRole.ACCESSORY
    else:
        relation = _ROLE_LABELS.get(suggested_label)
    return relation, ConsumptionLifecycle.CONSUMABLE


def _xlsx_operator_rows(path: Path) -> OperatorReviewInput:
    try:
        from openpyxl import load_workbook
    except ImportError as exc:  # pragma: no cover - dependency is required by SP-041B
        raise SemanticEngineV2Error(
            "PRIVATE_REPLAY_INPUT_INVALID", "XLSX reader dependency is unavailable",
        ) from exc

    try:
        workbook = load_workbook(path, read_only=True, data_only=True, keep_links=False)
    except (OSError, ValueError) as exc:
        raise SemanticEngineV2Error(
            "PRIVATE_REPLAY_INPUT_INVALID", "operator XLSX is unreadable",
        ) from exc
    try:
        if _OPERATOR_SHEET not in workbook.sheetnames:
            raise SemanticEngineV2Error(
                "PRIVATE_REPLAY_INPUT_INVALID", "operator XLSX sheet is missing",
            )
        sheet = workbook[_OPERATOR_SHEET]
        iterator = sheet.iter_rows(values_only=True)
        try:
            header_row = next(iterator)
        except StopIteration as exc:
            raise SemanticEngineV2Error(
                "PRIVATE_REPLAY_INPUT_INVALID", "operator XLSX is empty",
            ) from exc
        header_names = tuple(_cell_text(value) for value in header_row)
        if any(name not in header_names for name in _OPERATOR_HEADERS):
            raise SemanticEngineV2Error(
                "PRIVATE_REPLAY_INPUT_INVALID", "operator XLSX schema mismatch",
            )
        positions = {name: header_names.index(name) for name in _OPERATOR_HEADERS}
        raw_rows = tuple(
            values for values in iterator
            if any(_cell_text(value) is not None for value in values)
        )
    finally:
        workbook.close()

    rows: list[dict[str, Any]] = []
    identities: set[tuple[str, str]] = set()
    malformed_decisions = 0
    malformed_relations = 0
    valid_decisions = 0
    consumable_labels = 0
    for values in raw_rows:
        def cell(name: str) -> str | None:
            position = positions[name]
            return _cell_text(values[position] if position < len(values) else None)

        calibration_id = _CALIBRATION_IDS_BY_LABEL.get(cell("校准类目"))
        listing_reference = cell("ASIN")
        if calibration_id is None or listing_reference is None:
            raise SemanticEngineV2Error(
                "PRIVATE_REPLAY_INPUT_INVALID", "operator XLSX identity is invalid",
            )
        identity = (calibration_id, listing_reference)
        if identity in identities:
            raise SemanticEngineV2Error(
                "PRIVATE_REPLAY_INPUT_INVALID", "operator row is duplicated",
            )
        identities.add(identity)

        decision = cell("你的决定")
        if decision not in _ACCEPT_DECISIONS | _MODIFY_DECISIONS | _REVIEW_DECISIONS:
            malformed_decisions += 1
            continue
        valid_decisions += 1

        suggested_role = cell("我的商品角色建议")
        final_role = cell("最终商品角色")
        if decision in _REVIEW_DECISIONS:
            expected_relation = RelationRole.REVIEW_REQUIRED
            expected_lifecycle = None
        else:
            expected_relation, expected_lifecycle = _legacy_relation_role(
                final_role, suggested_role,
            )
        if expected_relation is None:
            # A valid MODIFY with an unfilled/invalid role override is not repaired
            # from the proposal.  It remains outside the agreement denominator.
            malformed_relations += 1
        if expected_lifecycle is not None:
            consumable_labels += 1

        final_scope = cell("最终市场范围")
        scope_projection = _MARKET_SCOPE_LABELS.get(final_scope)
        expected_cohort, boundary_tags = (
            scope_projection if scope_projection is not None else (None, ())
        )
        rows.append({
            "calibration_id": calibration_id,
            "listing_reference": listing_reference,
            "expected_relation_role": (
                None if expected_relation is None else expected_relation.value
            ),
            "expected_consumption_lifecycle": (
                None if expected_lifecycle is None else expected_lifecycle.value
            ),
            "expected_primary_cohort": expected_cohort,
            "boundary_tags": list(boundary_tags),
        })
    return OperatorReviewInput(
        rows=tuple(rows), raw_row_count=len(raw_rows),
        valid_decision_row_count=valid_decisions,
        malformed_decision_excluded_count=malformed_decisions,
        malformed_relation_label_excluded_count=malformed_relations,
        consumable_lifecycle_label_count=consumable_labels,
        source_format="XLSX",
    )


def _operator_rows(path: Path | None) -> OperatorReviewInput:
    if path is None:
        return OperatorReviewInput((), 0, 0, 0, 0, 0, "NONE")
    suffix = path.suffix.casefold()
    if suffix == ".json":
        return _json_operator_rows(path)
    if suffix == ".xlsx":
        return _xlsx_operator_rows(path)
    raise SemanticEngineV2Error(
        "PRIVATE_REPLAY_INPUT_INVALID", "operator review must be JSON or XLSX",
    )


def _count_rates(counts: dict[str, int], total: int) -> dict[str, float]:
    return {
        name: count / total if total else 0.0
        for name, count in sorted(counts.items())
    }


def _category_summary(
    result: Any, *, deterministic_match: bool, runtime: dict[str, float],
) -> dict[str, Any]:
    coverage = {
        role: available / total if total else 0.0
        for role, available, total in result.role_coverage_summary
    }
    identity_counts = dict(result.identity_status_counts)
    relation_counts = dict(result.relation_role_counts)
    lifecycle_counts = dict(result.lifecycle_counts)
    relationship_counts = dict(result.relationship_state_counts)
    cohort_counts = dict(result.cohort_state_counts)
    relation_governed = sum(
        count for name, count in relation_counts.items()
        if name not in {RelationRole.UNKNOWN.value, RelationRole.REVIEW_REQUIRED.value}
    )
    lifecycle_governed = sum(
        count for name, count in lifecycle_counts.items()
        if name not in {
            ConsumptionLifecycle.UNKNOWN.value,
            ConsumptionLifecycle.REVIEW_REQUIRED.value,
        }
    )
    return {
        "accepted_listing_count": result.listing_count,
        "listing_count": result.listing_count,
        "profile_id": result.profile_id,
        "profile_version": result.profile_version,
        "profile_fingerprint": result.profile_fingerprint,
        "result_fingerprint": result.semantic_fingerprint,
        "deterministic_replay_match": deterministic_match,
        "identity_status_counts": identity_counts,
        "identity_status_rates": _count_rates(identity_counts, result.listing_count),
        "relationship_state_counts": relationship_counts,
        "relationship_state_rates": _count_rates(
            relationship_counts, sum(relationship_counts.values()),
        ),
        "relation_role_counts": relation_counts,
        "relation_role_rates": _count_rates(relation_counts, result.listing_count),
        "relation_role_governed_count": relation_governed,
        "relation_role_governed_coverage": (
            relation_governed / result.listing_count if result.listing_count else 0.0
        ),
        "lifecycle_counts": lifecycle_counts,
        "lifecycle_rates": _count_rates(lifecycle_counts, result.listing_count),
        "lifecycle_governed_count": lifecycle_governed,
        "lifecycle_governed_coverage": (
            lifecycle_governed / result.listing_count if result.listing_count else 0.0
        ),
        "cohort_state_counts": cohort_counts,
        "cohort_state_rates": _count_rates(cohort_counts, result.listing_count),
        "role_coverage_rates": dict(sorted(coverage.items())),
        "review_listing_count": result.review_listing_count,
        "unknown_identity_count": result.unknown_identity_count,
        "network_calls": dict(result.diagnostics)["network_calls"],
        "llm_authoritative_decisions": dict(result.diagnostics)[
            "llm_authoritative_decisions"
        ],
        "runtime_seconds": dict(sorted(runtime.items())),
    }


def _operator_metrics(
    results: dict[str, Any], review: OperatorReviewInput,
) -> dict[str, Any]:
    listings = {
        (calibration_id, listing.listing_reference): listing
        for calibration_id, result in results.items() for listing in result.listings
    }
    matched = 0
    agreement = 0
    lifecycle_matched = 0
    lifecycle_agreement = 0
    boundary_failures = Counter({tag: 0 for tag in BOUNDARY_TAGS})
    boundary_samples = Counter({tag: 0 for tag in BOUNDARY_TAGS})
    cohort_labeled = 0
    cohort_mismatches = 0
    missing = 0
    for row in review.rows:
        listing = listings.get((row["calibration_id"], row["listing_reference"]))
        if listing is None:
            missing += 1
            continue
        expected = row["expected_relation_role"]
        if expected is not None:
            matched += 1
            agreement += listing.product_role.relation_role.value == expected
        expected_lifecycle = row.get("expected_consumption_lifecycle")
        if expected_lifecycle is not None:
            lifecycle_matched += 1
            lifecycle_agreement += (
                listing.product_role.consumption_lifecycle.value == expected_lifecycle
            )
        eligible = listing.market_cohort_eligibility.eligible_for_primary_cohort
        roles = {fact.role for fact in listing.facts}
        for tag in row["boundary_tags"]:
            boundary_samples[tag] += 1
            if tag in {"OBVIOUS_OTHER", "NONPRIMARY"}:
                boundary_failures[tag] += int(eligible)
            elif tag == "USE_CASE_ONLY":
                # Use-case language must not establish target Product Identity,
                # even when a separate role gate would also keep cohort false.
                boundary_failures[tag] += int(
                    listing.product_identity.is_target_identity is True
                )
            elif tag in {"INVALID_QUANTITY", "AMBIGUOUS_UNIT"}:
                boundary_failures[tag] += int(bool(roles & {
                    UniversalSemanticRole.QUANTITY, UniversalSemanticRole.SIZE_CAPACITY,
                }))
            elif tag == "HOST_CAPACITY_ACCESSORY":
                boundary_failures[tag] += sum(
                    fact.role is UniversalSemanticRole.SIZE_CAPACITY
                    and fact.semantic_scope is not SemanticScope.HOST_DEVICE
                    for fact in listing.facts
                )
        expected_cohort = row["expected_primary_cohort"]
        if expected_cohort is not None:
            cohort_labeled += 1
            cohort_mismatches += int(eligible != expected_cohort)
    return {
        "operator_row_count": review.raw_row_count,
        "operator_valid_decision_row_count": review.valid_decision_row_count,
        "operator_evaluable_row_count": len(review.rows),
        "malformed_decision_excluded_count": review.malformed_decision_excluded_count,
        "malformed_relation_label_excluded_count": review.malformed_relation_label_excluded_count,
        "operator_source_format": review.source_format,
        "operator_listing_missing_count": missing,
        "relation_role_labeled_count": matched,
        "relation_role_agreement_count": agreement,
        "relation_role_agreement_rate": agreement / matched if matched else None,
        "consumable_lifecycle_labeled_count": lifecycle_matched,
        "consumable_lifecycle_agreement_count": lifecycle_agreement,
        "consumable_lifecycle_agreement_rate": (
            lifecycle_agreement / lifecycle_matched if lifecycle_matched else None
        ),
        "primary_cohort_labeled_count": cohort_labeled,
        "primary_cohort_mismatch_count": cohort_mismatches,
        "obvious_other_sample_count": boundary_samples["OBVIOUS_OTHER"],
        "obvious_other_false_include_count": boundary_failures["OBVIOUS_OTHER"],
        "nonprimary_sample_count": boundary_samples["NONPRIMARY"],
        "nonprimary_leakage_count": boundary_failures["NONPRIMARY"],
        "use_case_only_sample_count": boundary_samples["USE_CASE_ONLY"],
        "use_case_identity_include_count": boundary_failures["USE_CASE_ONLY"],
        "invalid_quantity_sample_count": boundary_samples["INVALID_QUANTITY"],
        "invalid_quantity_accepted_count": boundary_failures["INVALID_QUANTITY"],
        "ambiguous_unit_sample_count": boundary_samples["AMBIGUOUS_UNIT"],
        "ambiguous_unit_coerced_count": boundary_failures["AMBIGUOUS_UNIT"],
        "host_capacity_accessory_sample_count": boundary_samples["HOST_CAPACITY_ACCESSORY"],
        "host_capacity_assigned_to_accessory_count": boundary_failures["HOST_CAPACITY_ACCESSORY"],
    }


def _operator_gate_results(operator: dict[str, Any]) -> dict[str, bool]:
    xlsx_input_shape_valid = (
        operator["operator_source_format"] != "XLSX"
        or (
            operator["operator_row_count"] == 60
            and operator["operator_valid_decision_row_count"] == 56
            and operator["malformed_decision_excluded_count"] == 4
            and operator["malformed_relation_label_excluded_count"] == 1
            and operator["relation_role_labeled_count"] == 55
        )
    )
    return {
        "operator_input_denominator_integrity": xlsx_input_shape_valid,
        "relation_role_operator_agreement": (
            operator["operator_row_count"] == 60
            and operator["operator_listing_missing_count"] == 0
            and operator["relation_role_labeled_count"] > 0
            and operator["relation_role_agreement_rate"] is not None
            and operator["relation_role_agreement_rate"] >= 0.90
        ),
        "obvious_other_false_include_zero": (
            operator["obvious_other_sample_count"] > 0
            and operator["obvious_other_false_include_count"] == 0
        ),
        "nonprimary_leakage_zero": (
            operator["nonprimary_sample_count"] > 0
            and operator["nonprimary_leakage_count"] == 0
        ),
        "use_case_identity_include_zero": (
            operator["use_case_only_sample_count"] > 0
            and operator["use_case_identity_include_count"] == 0
        ),
    }


def run_private_replay(manifest_path: Path, operator_path: Path | None) -> dict[str, Any]:
    replay_started = time.perf_counter()
    entries = _manifest(manifest_path)
    operator_review = _operator_rows(operator_path)
    results: dict[str, Any] = {}
    runtime_by_category: dict[str, dict[str, float]] = {}
    deterministic_by_category: dict[str, bool] = {}
    deterministic = True
    count_match = True
    for entry in entries:
        import_started = time.perf_counter()
        dataset = import_sellersprite_file(
            entry["input"],
            context=ImportContext(
                marketplace=entry["marketplace"], category=entry["category"],
                imported_at="2026-08-29T00:00:00+00:00",
                observed_date=entry["observed_date"], sheet_name=entry["sheet"],
            ),
        )
        import_seconds = time.perf_counter() - import_started
        profile = load_category_semantic_profile(entry["profile"])
        semantic_started = time.perf_counter()
        result = build_semantic_engine_v2_result(dataset, profile=profile)
        semantic_seconds = time.perf_counter() - semantic_started
        replay_dataset = replace(
            dataset, imported_at="2030-01-01T00:00:00+00:00",
            records=tuple(reversed(dataset.records)),
        )
        determinism_started = time.perf_counter()
        replay = build_semantic_engine_v2_result(replay_dataset, profile=profile)
        determinism_seconds = time.perf_counter() - determinism_started
        category_deterministic = result.to_json() == replay.to_json()
        deterministic_by_category[entry["calibration_id"]] = category_deterministic
        deterministic = deterministic and category_deterministic
        count_match = count_match and result.listing_count == EXPECTED_COUNTS[entry["calibration_id"]]
        results[entry["calibration_id"]] = result
        runtime_by_category[entry["calibration_id"]] = {
            "governed_import": round(import_seconds, 6),
            "semantic_build": round(semantic_seconds, 6),
            "determinism_rebuild": round(determinism_seconds, 6),
        }

    summaries = {
        calibration_id: _category_summary(
            result,
            deterministic_match=deterministic_by_category[calibration_id],
            runtime=runtime_by_category[calibration_id],
        )
        for calibration_id, result in sorted(results.items())
    }
    operator = _operator_metrics(results, operator_review)
    core_floor_results = {}
    for calibration_id, floors in CORE_FLOORS.items():
        for role, floor in floors:
            actual = summaries[calibration_id]["role_coverage_rates"][role.value]
            core_floor_results[f"{calibration_id}:{role.value}"] = {
                "calibration_id": calibration_id,
                "role": role.value,
                "floor": floor,
                "actual": actual,
                "passed": actual >= floor,
            }
    root = Path(__file__).resolve().parents[1]
    generic_text = "\n".join(
        path.read_text(encoding="utf-8")
        for path in sorted((root / "src/amazon_product_intelligence/semantic_engine_v2").glob("*.py"))
    ).casefold()
    category_literals = (
        "shower caddy", "dog water bottle", "vacuum filter",
        "food storage", "air fryer",
    )
    category_patch_count = sum(generic_text.count(item) for item in category_literals)
    profile_lineage_match = all(
        fact.profile_fingerprint == result.profile_fingerprint
        for result in results.values() for listing in result.listings for fact in listing.facts
    )
    gates = {
        "exact_corpus_counts": count_match,
        "deterministic_replay": deterministic,
        **_operator_gate_results(operator),
        "core_role_coverage_floors": all(item["passed"] for item in core_floor_results.values()),
        "invalid_quantity_accepted_zero": operator["invalid_quantity_accepted_count"] == 0,
        "ambiguous_unit_coerced_zero": operator["ambiguous_unit_coerced_count"] == 0,
        "host_capacity_assigned_to_accessory_zero": operator["host_capacity_assigned_to_accessory_count"] == 0,
        "generic_engine_category_patch_zero": category_patch_count == 0,
        "profile_fingerprint_match_100_percent": profile_lineage_match,
        "network_calls_zero": all(dict(result.diagnostics)["network_calls"] == 0 for result in results.values()),
        "llm_authoritative_decisions_zero": all(
            dict(result.diagnostics)["llm_authoritative_decisions"] == 0 for result in results.values()
        ),
    }
    report = {
        "contract_version": "semantic-engine-v2-private-replay-summary-v1.0",
        "calibration_category_count": len(results),
        "total_listing_count": sum(result.listing_count for result in results.values()),
        "categories": summaries,
        "core_floor_results": core_floor_results,
        "operator_safety": operator,
        "generic_engine_category_patch_count": category_patch_count,
        "runtime_seconds": {
            "total_replay_wall_clock": round(time.perf_counter() - replay_started, 6),
        },
        "gates": gates,
    }
    encoded = json.dumps(report, ensure_ascii=False, sort_keys=True, separators=(",", ":"))
    privacy_safe = not re.search(r"[A-Z0-9]*B0[A-Z0-9]{8}[A-Z0-9]*|[A-Za-z]:\\|/Users/|商品标题|详细参数", encoded)
    report["privacy_leak_count"] = 0 if privacy_safe else 1
    report["gates"]["privacy_leak_zero"] = privacy_safe
    if category_patch_count:
        report["verdict"] = "BLOCKED — GENERIC_ENGINE_CATEGORY_COUPLING"
    elif all(report["gates"].values()):
        report["verdict"] = "PASS — SEMANTIC_ENGINE_V2"
    else:
        report["verdict"] = "BLOCKED — PRIVATE_MULTI_CATEGORY_SEMANTIC_REPLAY_REQUIRED"
    return report


def _parser() -> argparse.ArgumentParser:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--manifest", required=True, help="private external five-category JSON manifest")
    parser.add_argument(
        "--operator-review",
        help="private external original 60-row XLSX workbook or legacy JSON labels",
    )
    parser.add_argument("--output", required=True, help="new aggregate-only JSON output")
    return parser


def main(argv: list[str] | None = None) -> int:
    args = _parser().parse_args(argv)
    try:
        output = Path(args.output)
        if output.suffix.casefold() != ".json" or output.exists():
            raise SemanticEngineV2Error("UNSAFE_OUTPUT", "output must be a new JSON file")
        report = run_private_replay(
            Path(args.manifest),
            None if args.operator_review is None else Path(args.operator_review),
        )
        with output.open("x", encoding="utf-8", newline="\n") as handle:
            json.dump(report, handle, ensure_ascii=False, sort_keys=True, indent=2)
            handle.write("\n")
        print(json.dumps({
            "status": "SUCCEEDED", "verdict": report["verdict"],
            "total_listing_count": report["total_listing_count"],
            "privacy_leak_count": report["privacy_leak_count"],
        }, ensure_ascii=False, sort_keys=True))
        return 0 if report["verdict"].startswith("PASS") else 3
    except (OSError, ValueError, SemanticEngineV2Error) as exc:
        print(json.dumps({"status": "FAILED", "error": str(exc)}, ensure_ascii=False, sort_keys=True))
        return 2


if __name__ == "__main__":
    raise SystemExit(main())
