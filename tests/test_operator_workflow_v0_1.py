from __future__ import annotations

from copy import deepcopy
import json
import os
from pathlib import Path
from tempfile import TemporaryDirectory
import unittest
from unittest.mock import patch

from openpyxl import load_workbook

import amazon_product_intelligence.operator_workflow as operator_workflow
from amazon_product_intelligence.contracts import canonical_json
from amazon_product_intelligence.operator_workflow import (
    OPERATOR_WORKFLOW_RULESET_VERSION,
    OperatorActionType,
    OperatorWorkflowBuilderV0_1,
    OperatorWorkflowRequest,
)
from amazon_product_intelligence.production_pipeline import (
    ProviderCreditSemantics,
    ProductionPipelineOrchestrator,
    ProductionRunMode,
    ProductionRunRequest,
    ProductionRunStatus,
)


ROOT = Path(__file__).resolve().parents[1]
ASINS = ("B0DWB00001", "B0DWB00002", "B0DWB00003")


class OperatorWorkflowV01Tests(unittest.TestCase):
    @classmethod
    def setUpClass(cls) -> None:
        if not os.environ.get("MARKET_REPORT_NODE_EXECUTABLE") or not os.environ.get(
            "MARKET_REPORT_NODE_MODULES"
        ):
            raise unittest.SkipTest("artifact-tool runtime paths are not configured")
        cls.temp = TemporaryDirectory(dir=ROOT / "outputs")
        cls.output = Path(cls.temp.name) / "fixture-run"
        request = ProductionRunRequest(
            marketplace="US",
            asins=ASINS,
            output_directory=cls.output,
            run_id="sp037-operator-fixture",
            mode=ProductionRunMode.FIXTURE,
        )
        with patch(
            "amazon_product_intelligence.production_pipeline.orchestrator.HttpJsonTransport.execute",
            side_effect=AssertionError("fixture acceptance must not access HTTP"),
        ) as http_execute:
            cls.result = ProductionPipelineOrchestrator().run(request)
            cls.http_call_count = http_execute.call_count
        cls.workflow = cls.result.operator_workflow
        cls.manifest = json.loads(
            (cls.output / "run_manifest.json").read_text(encoding="utf-8")
        )
        cls.markdown = (cls.output / "operator_market_report.md").read_text(
            encoding="utf-8"
        )
        cls.workbook = load_workbook(
            cls.output / "operator_market_report.xlsx", read_only=True, data_only=False
        )

    @classmethod
    def tearDownClass(cls) -> None:
        if hasattr(cls, "workbook"):
            cls.workbook.close()
        if hasattr(cls, "temp"):
            cls.temp.cleanup()

    def test_public_contract_is_versioned_and_explicit(self) -> None:
        self.assertEqual("operator-workflow-v0.1", OPERATOR_WORKFLOW_RULESET_VERSION)
        self.assertEqual(
            {
                "OPERATOR_WORKFLOW_RULESET_VERSION",
                "OperatorActionType",
                "OperatorClaim",
                "OperatorNextAction",
                "OperatorRunHealth",
                "OperatorWorkflowBuilderV0_1",
                "OperatorWorkflowRequest",
                "OperatorWorkflowSnapshotV0_1",
                "OperatorWorkflowValidationError",
                "build_standalone_operator_workflow",
            },
            set(operator_workflow.__all__),
        )

    def test_fixture_pipeline_builds_conservative_governed_action(self) -> None:
        self.assertIs(self.result.status, ProductionRunStatus.SUCCEEDED)
        self.assertEqual(0, self.http_call_count)
        self.assertIsNotNone(self.workflow)
        self.assertEqual("COLLECT_EVIDENCE", self.workflow["operator_action"])
        self.assertEqual(
            "EVIDENCE_COLLECTION_RECOMMENDED",
            self.workflow["recommendation_type"],
        )
        integration = self.workflow["framework_integration"]
        self.assertEqual("INSUFFICIENT_EVIDENCE", integration["recommendation_applicability"])
        self.assertEqual("decision-framework-v0.1", integration["decision_framework_ruleset_version"])
        self.assertEqual(
            "recommendation-framework-v0.1",
            integration["recommendation_framework_ruleset_version"],
        )
        self.assertEqual("UNAVAILABLE_ADAPTER_GAP", integration["decision_framework_execution_status"])
        self.assertIn("not reconstructed or fabricated", integration["adapter_gap"])

    def test_missing_values_remain_typed_nulls_and_never_zero(self) -> None:
        opportunity = next(
            item
            for item in self.workflow["opportunity_summary"]
            if item["label"] == "Opportunity Score"
        )
        self.assertEqual("PENDING_DATA", opportunity["status"])
        self.assertIsNone(opportunity["value"])
        unavailable_competition = [
            item
            for item in self.workflow["competition_summary"]
            if item["status"] == "UNAVAILABLE"
        ]
        self.assertTrue(unavailable_competition)
        self.assertTrue(all(item["value"] is None for item in unavailable_competition))
        self.assertNotIn("Opportunity Score | 0", self.markdown)
        self.assertIn("PENDING_DATA", self.markdown)
        self.assertIn("UNAVAILABLE", self.markdown)

    def test_executive_summary_is_first_and_markdown_xlsx_are_semantically_equal(self) -> None:
        self.assertEqual("# Operator Brief", self.markdown.splitlines()[0])
        self.assertEqual("Operator Summary", self.workbook.sheetnames[0])
        summary = self.workbook["Operator Summary"]
        self.assertEqual("Operator Action", summary["A4"].value)
        self.assertEqual(self.workflow["operator_action"], summary["B4"].value)
        self.assertIn(self.workflow["operator_action"], self.markdown)
        self.assertIn(self.workflow["action_reason"], self.markdown)
        for label in (
            "Why This Action",
            "Evidence Readiness",
            "Top Opportunity Themes",
            "Top Risks / Blockers",
            "Missing Evidence",
            "Recommended Next Checks",
            "Run Health",
            "Provider Usage",
            "Audit / Provenance",
        ):
            self.assertTrue(
                any(summary.cell(row=row, column=1).value == label for row in range(4, 16)),
                label,
            )

    def test_every_operator_claim_and_action_retains_lineage(self) -> None:
        for name in (
            "supporting_evidence",
            "missing_evidence",
            "top_buyer_need_themes",
            "competition_summary",
            "opportunity_summary",
            "risks_and_limitations",
        ):
            for item in self.workflow[name]:
                self.assertTrue(item["provenance_reference_ids"], (name, item))
        for item in self.workflow["next_actions"]:
            self.assertTrue(item["provenance_reference_ids"], item)
            self.assertIn(item["trigger_status"], {"PARTIAL", "UNKNOWN", "UNAVAILABLE"})
            self.assertIn(item["priority"], {1, 2, 3, 4, 5})

    def test_run_health_is_explicit_and_fixture_credits_are_not_billed(self) -> None:
        health = self.workflow["run_health"]
        self.assertEqual("SUCCEEDED", health["status"])
        self.assertFalse(health["retried"])
        self.assertFalse(health["resumed"])
        self.assertEqual(4, health["logical_operation_count"])
        self.assertEqual(4, health["transport_attempt_count"])
        self.assertEqual(4.0, health["credits"])
        self.assertEqual(ProviderCreditSemantics.FIXTURE_REFERENCE.value, health["credit_semantics"])
        self.assertIn("fixture reference credits; not billed", self.markdown)

    def test_manifest_is_secret_safe_and_contains_same_snapshot(self) -> None:
        self.assertEqual(self.workflow, self.manifest["operator_workflow"])
        serialized = canonical_json(self.manifest).casefold()
        self.assertNotIn("offline-fixture-sentinel", serialized)
        for forbidden in ("x-api-key", '"authorization"', "password", "api_key"):
            self.assertNotIn(forbidden, serialized)

    def test_semantic_fingerprint_excludes_runtime_only_health(self) -> None:
        report = OperatorReportLoader.load(self.output / "market_report.json")
        base = OperatorWorkflowBuilderV0_1().build(
            OperatorWorkflowRequest(
                report=report,
                run_id="first-runtime",
                run_status="SUCCEEDED",
                provider_summary=self.result.provider_summary.to_dict(),
                recovery=self.result.recovery,
            )
        )
        resumed_summary = deepcopy(self.result.provider_summary.to_dict())
        resumed_summary["executed_operation_count"] = 1
        resumed_summary["replayed_operation_count"] = 3
        resumed_summary["transport_attempt_count"] = 1
        resumed = OperatorWorkflowBuilderV0_1().build(
            OperatorWorkflowRequest(
                report=report,
                run_id="resumed-runtime",
                run_status="SUCCEEDED",
                provider_summary=resumed_summary,
                recovery={**self.result.recovery, "resume_source_run_id": "faulted-source"},
            )
        )
        self.assertEqual(base.semantic_fingerprint, resumed.semantic_fingerprint)
        self.assertEqual(base.operator_action, resumed.operator_action)
        self.assertEqual(base.next_actions, resumed.next_actions)
        self.assertNotEqual(base.run_health, resumed.run_health)


class OperatorReportLoader:
    @staticmethod
    def load(path: Path):
        from amazon_product_intelligence.market_report import validate_market_report_payload

        return validate_market_report_payload(json.loads(path.read_text(encoding="utf-8")))


if __name__ == "__main__":
    unittest.main()
