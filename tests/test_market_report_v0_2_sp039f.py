from __future__ import annotations

import json
from pathlib import Path
from tempfile import TemporaryDirectory

import pytest
from openpyxl import load_workbook

from amazon_product_intelligence.market_report.v0_2 import market_report_v0_2_from_dict
from amazon_product_intelligence.market_report.v0_2.delivery import (
    SHEET_NAMES,
    OperatorReportDeliveryV0_2,
    compose_operator_view,
)
from amazon_product_intelligence.market_report.v0_2.models import MarketReportV0_2ValidationError
from amazon_product_intelligence.production_pipeline import (
    ProductionRunMode,
    ProductionRunRequest,
    ProductionRunStatus,
    ProductionRunValidationError,
)
from amazon_product_intelligence.production_pipeline.cli import build_parser
from amazon_product_intelligence.production_pipeline.orchestrator import ProductionPipelineOrchestrator
from amazon_product_intelligence.production_pipeline.recovery import run_request_fingerprint
from amazon_product_intelligence.xlsx_delivery.package_fingerprint import ooxml_package_content_sha256
from tests.test_market_report_v0_2_sp039e import build_snapshot
from tests.test_production_pipeline_v0_1 import RecordingDelivery
from tests.test_production_reliability_v0_1 import (
    FIXTURE_PATH,
    FaultInjectingFixtureTransport,
    response,
    runtime_factory,
)


ROOT = Path(__file__).resolve().parents[1]
ASINS = ("B0DWB00001", "B0DWB00002", "B0DWB00003")


def request(output: Path, *, version: str = "market-report-v0.2", resume_from: Path | None = None):
    return ProductionRunRequest(
        marketplace="US",
        asins=ASINS,
        output_directory=output,
        report_version=version,
        resume_from=resume_from,
        run_id="sp039f-offline",
    )


def test_operator_view_is_pure_deterministic_projection():
    report = build_snapshot()
    before = report.to_dict()
    first = compose_operator_view(report)
    second = compose_operator_view(report)
    assert dict(first.parity) == dict(second.parity)
    assert report.to_dict() == before
    assert first.parity["report_id"] == report.metadata.report_id
    assert first.parity["product_direction_semantics"] == "HYPOTHESIS"
    assert first.parity["competitor_shortlist_semantics"] == "REVIEW_ORDER_NOT_RANK"


def test_delivery_uses_one_snapshot_and_has_twelve_operator_first_sheets():
    report = build_snapshot(market_unavailable=True, direction_unavailable=True)
    with TemporaryDirectory(dir=ROOT / "outputs") as directory:
        result = OperatorReportDeliveryV0_2().deliver(report, Path(directory))
        workbook = load_workbook(result.xlsx_path, data_only=False)
        assert tuple(workbook.sheetnames) == SHEET_NAMES
        assert all(cell.data_type != "f" for sheet in workbook for row in sheet for cell in row)
        parity_rows = dict(
            workbook["Executive Summary"].iter_rows(
                min_row=2, max_row=22, min_col=1, max_col=2, values_only=True
            )
        )
        workbook.close()
        markdown = result.markdown_path.read_text(encoding="utf-8")
        assert parity_rows["report_id"] == report.metadata.report_id
        assert report.metadata.report_id in markdown
        assert "UNAVAILABLE" in markdown
        assert "hypotheses" in markdown
        assert "not a rank" in markdown
        assert result.xlsx_package_content_sha256 == ooxml_package_content_sha256(result.xlsx_path)


def test_delivery_rejects_invalid_snapshot_before_artifacts():
    payload = build_snapshot().to_dict()
    payload["metadata"]["report_version"] = "market-report-v9"
    with TemporaryDirectory(dir=ROOT / "outputs") as directory:
        output = Path(directory)
        with pytest.raises(MarketReportV0_2ValidationError):
            OperatorReportDeliveryV0_2().deliver(payload, output)
        assert not (output / "operator_market_report.xlsx").exists()
        assert not (output / "operator_market_report.md").exists()


def test_explicit_v02_fixture_pipeline_writes_validated_four_artifact_contract():
    with TemporaryDirectory(dir=ROOT / "outputs") as directory:
        output = Path(directory)
        result = ProductionPipelineOrchestrator().run(request(output))
        assert result.status is ProductionRunStatus.SUCCEEDED
        assert result.market_report_version == "market-report-v0.2"
        expected = {
            "market_report.json", "operator_market_report.xlsx",
            "operator_market_report.md", "run_manifest.json",
        }
        assert expected <= {path.name for path in output.iterdir()}
        report_payload = json.loads((output / "market_report.json").read_text(encoding="utf-8"))
        report = market_report_v0_2_from_dict(report_payload)
        manifest = json.loads((output / "run_manifest.json").read_text(encoding="utf-8"))
        assert manifest["requested_market_report_version"] == "market-report-v0.2"
        assert manifest["market_report_version"] == "market-report-v0.2"
        assert manifest["market_report_id"] == report.metadata.report_id
        assert manifest["delivery_status"] == "SUCCEEDED"
        assert manifest["provider_summary"]["credit_semantics"] == "FIXTURE_REFERENCE"
        assert "market-report-v0.2" in (output / "operator_market_report.md").read_text(encoding="utf-8")


def test_omitted_version_remains_v01_and_manifest_shape_is_unchanged():
    with TemporaryDirectory(dir=ROOT / "outputs") as directory:
        output = Path(directory)
        v01 = ProductionRunRequest(marketplace="US", asins=ASINS, output_directory=output)
        assert v01.report_version == "market-report-v0.1"
        result = ProductionPipelineOrchestrator(delivery=RecordingDelivery()).run(v01)
        assert result.status is ProductionRunStatus.SUCCEEDED
        manifest = json.loads((output / "run_manifest.json").read_text(encoding="utf-8"))
        assert manifest["market_report_version"] == "market-report-v0.1"
        assert "requested_market_report_version" not in manifest
        assert "market_report_id" not in manifest
        assert "delivery_status" not in manifest


def test_unknown_version_fails_at_request_boundary():
    with TemporaryDirectory(dir=ROOT / "outputs") as directory:
        with pytest.raises(ProductionRunValidationError):
            request(Path(directory), version="market-report-v9")


def test_xiyou_v02_live_mode_is_rejected_at_request_boundary():
    with TemporaryDirectory(dir=ROOT / "outputs") as directory:
        with pytest.raises(ProductionRunValidationError, match="explicit Sorftime"):
            ProductionRunRequest(
                marketplace="US",
                asins=ASINS,
                output_directory=Path(directory),
                report_version="market-report-v0.2",
                mode=ProductionRunMode.LIVE,
            )


def test_version_is_part_of_resume_compatibility_fingerprint():
    with TemporaryDirectory(dir=ROOT / "outputs") as directory:
        output = Path(directory)
        assert run_request_fingerprint(request(output)) != run_request_fingerprint(
            request(output, version="market-report-v0.1")
        )


def test_cross_version_resume_fails_before_provider_factory():
    with TemporaryDirectory(dir=ROOT / "outputs") as directory:
        root = Path(directory)
        source = root / "source"
        first = ProductionPipelineOrchestrator(delivery=RecordingDelivery()).run(
            request(source, version="market-report-v0.1")
        )
        assert first.status is ProductionRunStatus.SUCCEEDED
        calls = []

        def forbidden_factory(_request):
            calls.append("called")
            raise AssertionError("provider factory must not be called")

        resumed = ProductionPipelineOrchestrator(
            provider_runtime_factory=forbidden_factory
        ).run(request(root / "target", resume_from=source))
        assert resumed.status is ProductionRunStatus.FAILED
        assert resumed.error["code"] == "INCOMPATIBLE_RESUME_SOURCE"
        assert calls == []


def test_same_version_resume_preserves_semantic_outputs_and_replays_fixture_checkpoints():
    with TemporaryDirectory(dir=ROOT / "outputs") as directory:
        root = Path(directory)
        source = root / "source"
        target = root / "target"
        fixture = json.loads(FIXTURE_PATH.read_text(encoding="utf-8"))
        faulting_transport = FaultInjectingFixtureTransport(
            fixture,
            {f"asin_keywords:{ASINS[1]}": [response(503), response(503)]},
        )
        failed = ProductionPipelineOrchestrator(
            provider_runtime_factory=runtime_factory(faulting_transport, [])
        ).run(request(source))
        assert failed.status is ProductionRunStatus.FAILED
        fresh = root / "fresh"
        first = ProductionPipelineOrchestrator().run(request(fresh))
        resumed = ProductionPipelineOrchestrator().run(request(target, resume_from=source))
        assert first.status is resumed.status is ProductionRunStatus.SUCCEEDED
        first_report = json.loads((fresh / "market_report.json").read_text(encoding="utf-8"))
        resumed_report = json.loads((target / "market_report.json").read_text(encoding="utf-8"))
        assert first_report["metadata"]["semantic_fingerprint"] == resumed_report["metadata"]["semantic_fingerprint"]
        assert first_report["metadata"]["report_id"] == resumed_report["metadata"]["report_id"]
        assert (fresh / "operator_market_report.md").read_text(encoding="utf-8") == (
            target / "operator_market_report.md"
        ).read_text(encoding="utf-8")
        assert ooxml_package_content_sha256(fresh / "operator_market_report.xlsx") == (
            ooxml_package_content_sha256(target / "operator_market_report.xlsx")
        )
        assert resumed.provider_summary.replayed_operation_count > 0


def test_cli_opt_in_is_run_only_and_default_remains_v01():
    parser = build_parser()
    default = parser.parse_args([
        "run", "--market", "US", "--asin", ASINS[0], "--output-dir", "out"
    ])
    explicit = parser.parse_args([
        "run", "--market", "US", "--asin", ASINS[0], "--output-dir", "out",
        "--report-version", "market-report-v0.2",
    ])
    batch = parser.parse_args(["batch", "--batch-file", "batch.json", "--output-dir", "out"])
    assert default.report_version == "market-report-v0.1"
    assert explicit.report_version == "market-report-v0.2"
    assert not hasattr(batch, "report_version")


def test_report_id_and_semantics_change_when_governed_source_changes():
    first = build_snapshot(monthly_sales_value=120)
    second = build_snapshot(monthly_sales_value=121)
    assert first.metadata.report_id != second.metadata.report_id
    assert compose_operator_view(first).parity["monthly_sales"] == 120
    assert compose_operator_view(second).parity["monthly_sales"] == 121


def test_operational_metadata_does_not_change_semantic_identity():
    first = build_snapshot(operational_metadata={"run_id": "one", "credits": 0})
    second = build_snapshot(operational_metadata={"run_id": "two", "credits": 999})
    assert first.metadata.semantic_fingerprint == second.metadata.semantic_fingerprint
    assert first.metadata.report_id == second.metadata.report_id


def test_null_market_metrics_are_never_rendered_as_zero():
    view = compose_operator_view(build_snapshot(market_unavailable=True))
    assert view.parity["monthly_sales"] == "UNAVAILABLE"
    assert view.parity["monthly_revenue"] == "UNAVAILABLE"
    assert view.parity["monthly_sales"] != 0


def test_keyword_not_attached_is_explicit_not_zero_demand():
    view = compose_operator_view(build_snapshot())
    assert view.parity["keyword_intelligence_state"] == "NOT_ATTACHED"
    assert "demand" not in view.parity
