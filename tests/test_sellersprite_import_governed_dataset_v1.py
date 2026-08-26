from __future__ import annotations

from hashlib import sha256
from pathlib import Path

import pytest
from openpyxl import Workbook

from amazon_product_intelligence.sellersprite_import import (
    ImportContext,
    ImportValueStatus,
    SellerSpriteImportError,
    import_sellersprite_file,
)
from amazon_product_intelligence.sellersprite_import.schema_v1 import (
    FIELD_SPECS,
    MAX_LISTING_ROWS,
)


BASE_HEADERS = ["ASIN", "商品标题", "价格($)", "月销量", "评分"]


def context(**overrides: object) -> ImportContext:
    values = {
        "marketplace": "US",
        "category": "synthetic-category",
        "imported_at": "2026-08-26T12:00:00Z",
        "observed_date": "2026-08-25",
    }
    values.update(overrides)
    return ImportContext(**values)  # type: ignore[arg-type]


def write_csv(path: Path, headers: list[str], rows: list[list[object]], *, bom: bool = True) -> None:
    import csv
    import io

    buffer = io.StringIO(newline="")
    writer = csv.writer(buffer)
    writer.writerow(headers)
    writer.writerows(rows)
    payload = buffer.getvalue().encode("utf-8")
    path.write_bytes((b"\xef\xbb\xbf" if bom else b"") + payload)


def add_sheet(workbook: Workbook, title: str, headers: list[str], rows: list[list[object]]) -> None:
    worksheet = workbook.create_sheet(title)
    worksheet.append(headers)
    for row in rows:
        worksheet.append(row)


def field(dataset: object, asin: str, header: str):
    record = next(item for item in dataset.records if item.asin == asin)
    return next(item for item in record.fields if item.header == header)


def test_csv_utf8_sig_typed_import_and_governed_metadata(tmp_path: Path) -> None:
    source = tmp_path / "seller-export.csv"
    write_csv(
        source,
        BASE_HEADERS + ["销量同比增长率", "上架时间", "未知列", "LQS", "毛利率"],
        [["SYNTH00001", " Synthetic title ", "$19.90", "1,234", "4.7", "-12.5%", "2026-01-02", "ignored", "99", "40%"]],
    )

    dataset = import_sellersprite_file(source, context=context())

    assert dataset.source_type == "CSV"
    assert dataset.source_basename == "seller-export.csv"
    assert dataset.accepted_listing_count == 1
    assert dataset.unique_asin_count == 1
    assert dataset.unmapped_headers == ("未知列",)
    assert dataset.out_of_scope_headers == ("LQS",)
    assert field(dataset, "SYNTH00001", "月销量").value == 1234
    assert str(field(dataset, "SYNTH00001", "价格($)").value) == "19.90"
    assert str(field(dataset, "SYNTH00001", "销量同比增长率").value) == "-0.125"
    assert field(dataset, "SYNTH00001", "毛利率").evidence_semantics.value == "REFERENCE_ONLY_NOT_PROCUREMENT_TRUTH"
    assert all(item.header not in {"LQS", "SP广告"} for item in dataset.records[0].fields)
    assert str(tmp_path) not in dataset.to_json()


def test_full_66_header_xlsx_is_order_independent_and_prefers_raw_sheet(tmp_path: Path) -> None:
    source = tmp_path / "full.xlsx"
    workbook = Workbook()
    workbook.remove(workbook.active)
    add_sheet(workbook, "other", BASE_HEADERS, [["SYNTH00002", "wrong", "8", "1", "4"]])
    headers = [spec.header for spec in reversed(FIELD_SPECS)]
    values = {"ASIN": "SYNTH00001", "商品标题": "right", "价格($)": 12, "月销量": 3, "评分": 5}
    add_sheet(workbook, "原始数据源", headers, [[values.get(header) for header in headers]])
    workbook.save(source)

    dataset = import_sellersprite_file(source, context=context())

    assert dataset.source_sheet == "原始数据源"
    assert [record.asin for record in dataset.records] == ["SYNTH00001"]
    assert len(FIELD_SPECS) == 66


def test_explicit_sheet_resolves_multi_sheet_choice(tmp_path: Path) -> None:
    source = tmp_path / "multi.xlsx"
    workbook = Workbook()
    workbook.remove(workbook.active)
    add_sheet(workbook, "one", BASE_HEADERS, [["SYNTH00001", "one", 1, 1, 4]])
    add_sheet(workbook, "two", BASE_HEADERS, [["SYNTH00002", "two", 2, 2, 5]])
    workbook.save(source)

    with pytest.raises(SellerSpriteImportError, match="AMBIGUOUS_HEADER"):
        import_sellersprite_file(source, context=context())
    selected = import_sellersprite_file(source, context=context(sheet_name="two"))
    assert [record.asin for record in selected.records] == ["SYNTH00002"]


def test_no_fuzzy_header_and_duplicate_mapped_header_fail_closed(tmp_path: Path) -> None:
    fuzzy = tmp_path / "fuzzy.csv"
    write_csv(fuzzy, ["asin", "商品标题", "价格($)", "月销量", "评分"], [["SYNTH00001", "x", 1, 1, 5]])
    with pytest.raises(SellerSpriteImportError, match="HEADER_SCHEMA_MISMATCH"):
        import_sellersprite_file(fuzzy, context=context())

    duplicate = tmp_path / "duplicate.csv"
    write_csv(duplicate, BASE_HEADERS + ["ASIN"], [["SYNTH00001", "x", 1, 1, 5, "SYNTH00001"]])
    with pytest.raises(SellerSpriteImportError, match="DUPLICATE_MAPPED_HEADER"):
        import_sellersprite_file(duplicate, context=context())


def test_equivalent_duplicates_dedupe_and_conflicts_quarantine_all(tmp_path: Path) -> None:
    equivalent = tmp_path / "equivalent.csv"
    row = ["SYNTH00001", "same", "9.99", 12, "4.2"]
    write_csv(equivalent, BASE_HEADERS, [row, row])
    dataset = import_sellersprite_file(equivalent, context=context())
    assert dataset.accepted_listing_count == 1
    assert dataset.duplicate_row_count == 1
    assert dataset.quarantined_row_count == 0

    conflict = tmp_path / "conflict.csv"
    write_csv(conflict, BASE_HEADERS, [row, ["SYNTH00001", "different", "9.99", 12, "4.2"]])
    dataset = import_sellersprite_file(conflict, context=context())
    assert dataset.accepted_listing_count == 0
    assert dataset.duplicate_row_count == 0
    assert dataset.quarantined_row_count == 2
    assert {item.disposition.value for item in dataset.row_outcomes} == {"QUARANTINED_CONFLICT"}


def test_missing_invalid_asin_and_malformed_values_are_explicit(tmp_path: Path) -> None:
    source = tmp_path / "quality.csv"
    headers = BASE_HEADERS + ["销量同比增长率", "上架时间", "父ASIN"]
    write_csv(
        source,
        headers,
        [
            ["", "missing", 1, 1, 4, "1%", "2026-01-01", ""],
            ["bad", "invalid", 1, 1, 4, "1%", "2026-01-01", ""],
            ["SYNTH00001", "kept", "bad-money", "N/A", "9", "not-percent", "01/02/2026", "bad-parent"],
        ],
    )
    dataset = import_sellersprite_file(source, context=context())
    assert dataset.accepted_listing_count == 1
    assert dataset.rejected_row_count == 2
    assert field(dataset, "SYNTH00001", "价格($)").import_status is ImportValueStatus.PARSE_FAILED
    assert field(dataset, "SYNTH00001", "月销量").import_status is ImportValueStatus.NOT_AVAILABLE
    assert field(dataset, "SYNTH00001", "评分").import_status is ImportValueStatus.PARSE_FAILED
    assert field(dataset, "SYNTH00001", "上架时间").import_status is ImportValueStatus.PARSE_FAILED
    assert field(dataset, "SYNTH00001", "父ASIN").import_status is ImportValueStatus.PARSE_FAILED
    assert field(dataset, "SYNTH00001", "价格($)").value is None


def test_missing_core_headers_and_blank_cells_never_become_zero(tmp_path: Path) -> None:
    source = tmp_path / "subset.csv"
    write_csv(source, BASE_HEADERS, [["SYNTH00001", "title", "", "N/A", 4]])
    dataset = import_sellersprite_file(source, context=context(observed_date=None))
    assert dataset.observed_date_status == "UNKNOWN"
    assert field(dataset, "SYNTH00001", "价格($)").import_status is ImportValueStatus.BLANK
    assert field(dataset, "SYNTH00001", "月销量").import_status is ImportValueStatus.NOT_AVAILABLE
    assert field(dataset, "SYNTH00001", "FBA($)").import_status is ImportValueStatus.MISSING_HEADER
    assert field(dataset, "SYNTH00001", "价格($)").value is None
    assert dict(dataset.missing_core_field_summary)["FBA($)"] == 1


def test_deterministic_identity_input_immutability_and_zero_network(tmp_path: Path, monkeypatch: pytest.MonkeyPatch) -> None:
    source = tmp_path / "stable.csv"
    write_csv(source, BASE_HEADERS, [["SYNTH00001", "title", 10, 2, 4.5]], bom=False)
    before = sha256(source.read_bytes()).hexdigest()

    import socket

    def blocked(*args: object, **kwargs: object) -> None:
        raise AssertionError("network constructor must not be called")

    monkeypatch.setattr(socket, "socket", blocked)
    first = import_sellersprite_file(source, context=context())
    second = import_sellersprite_file(source, context=context())
    after = sha256(source.read_bytes()).hexdigest()
    assert first.dataset_id == second.dataset_id
    assert first.semantic_fingerprint == second.semantic_fingerprint
    assert first.to_json() == second.to_json()
    assert before == after == first.source_file_sha256


def test_csv_encoding_ambiguity_and_row_bound_fail_closed(tmp_path: Path) -> None:
    utf16 = tmp_path / "utf16.csv"
    utf16.write_bytes((",".join(BASE_HEADERS) + "\n").encode("utf-16"))
    with pytest.raises(SellerSpriteImportError, match="UNSUPPORTED_CSV_ENCODING"):
        import_sellersprite_file(utf16, context=context())

    too_many = tmp_path / "too-many.csv"
    rows = [[f"SYN{i:07d}", "title", 1, 1, 4] for i in range(MAX_LISTING_ROWS + 1)]
    write_csv(too_many, BASE_HEADERS, rows)
    with pytest.raises(SellerSpriteImportError, match="ROW_LIMIT_EXCEEDED"):
        import_sellersprite_file(too_many, context=context())
