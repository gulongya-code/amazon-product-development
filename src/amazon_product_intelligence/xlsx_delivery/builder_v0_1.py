"""Deterministic XLSX Operator Delivery Foundation V0.1 builder."""

from __future__ import annotations

import base64
from datetime import datetime
from hashlib import sha256
from io import BytesIO
import re
from typing import Any, Mapping
from zipfile import ZIP_DEFLATED, ZipFile, ZipInfo

import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from .errors import XlsxDeliveryValidationError
from .models import (
    XLSX_DELIVERY_RULESET_VERSION,
    CellRenderRecord,
    DeliveryDiagnostic,
    DeliveryLineageReference,
    WorkbookDeliveryRecord,
    WorkbookStyleDefinition,
    WorksheetRenderDefinition,
    XlsxDeliveryRequest,
    XlsxDeliverySnapshotV0_1,
    _DELIVERY_LAYOUT,
    _LINEAGE_COPY_FIELDS,
    _MEDIA_TYPE,
    _WORKBOOK_FILENAME,
    _chunked_delivery_values,
    _delivery_values,
    canonical_json,
    coverage_from_delivery,
    deterministic_id,
)


_FIXED_DOCUMENT_TIME = datetime(2000, 1, 1, 0, 0, 0)
_ZIP_TIME = (1980, 1, 1, 0, 0, 0)


def _style_definition() -> WorkbookStyleDefinition:
    content = {
        "font_name": "Aptos",
        "title_font_size": 16.0,
        "header_font_size": 11.0,
        "body_font_size": 10.0,
        "title_fill": "1F4E78",
        "header_fill": "5B9BD5",
        "title_font_color": "FFFFFF",
        "header_font_color": "FFFFFF",
        "body_font_color": "1F1F1F",
        "border_color": "B4C6E7",
        "wrap_text": True,
        "max_cell_chars": 30000,
    }
    return WorkbookStyleDefinition(
        style_id=deterministic_id("xlsx-style", content),
        **content,
    )


def _column_width(column: str) -> float:
    fixed = {
        "ASIN": 16.0,
        "Marketplace": 14.0,
        "Channel": 16.0,
        "Provider": 16.0,
        "Evidence Count": 15.0,
        "Recommendation Type": 24.0,
        "Relationship Type": 22.0,
    }
    if column in fixed:
        return fixed[column]
    if column in {"Title", "Keyword", "Product Endpoint", "Product"}:
        return 34.0
    if column in {"Limitations", "Explanation", "Source Reference"}:
        return 42.0
    return 48.0


def _worksheet_definition(
    *,
    ordinal: int,
    sheet_name: str,
    title: str,
    source_table_id: str,
    source_export_sheet_id: str,
    columns: tuple[str, ...],
    rendered_row_count: int,
    source_export_row_ids: tuple[str, ...] = (),
    delivery_row_ids: tuple[str, ...] = (),
    lineage_reference_ids: tuple[str, ...] = (),
) -> WorksheetRenderDefinition:
    last_row = max(2, 2 + rendered_row_count)
    identity = {
        "ordinal": ordinal,
        "sheet_name": sheet_name,
        "title": title,
        "source_table_id": source_table_id,
        "source_export_sheet_id": source_export_sheet_id,
        "columns": columns,
        "column_widths": tuple(_column_width(item) for item in columns),
        "title_row": 1,
        "header_row": 2,
        "data_start_row": 3,
        "freeze_panes": "A2",
        "auto_filter_range": f"A2:{get_column_letter(len(columns))}{last_row}",
    }
    return WorksheetRenderDefinition(
        worksheet_id=deterministic_id("xlsx-worksheet", identity),
        source_export_row_ids=source_export_row_ids,
        delivery_row_ids=delivery_row_ids,
        lineage_reference_ids=lineage_reference_ids,
        **identity,
    )


def _delivery_lineage(
    *,
    worksheet_id: str,
    source_export_snapshot_id: str,
    source_export_lineage: Mapping[str, Any],
) -> DeliveryLineageReference:
    content = {
        "worksheet_id": worksheet_id,
        "source_export_snapshot_id": source_export_snapshot_id,
        "source_export_row_id": source_export_lineage["export_row_id"],
        "source_export_lineage_id": source_export_lineage["export_lineage_id"],
        **{
            field_name: source_export_lineage[field_name]
            for field_name in _LINEAGE_COPY_FIELDS
        },
    }
    return DeliveryLineageReference(
        delivery_lineage_id=deterministic_id("xlsx-lineage", content),
        **content,
    )


def _cell(
    *,
    worksheet_id: str,
    delivery_row_id: str,
    source_export_row_id: str,
    excel_row: int,
    excel_column: int,
    chunk_index: int,
    column_name: str,
    value: Any,
    lineage_reference_ids: tuple[str, ...],
) -> CellRenderRecord:
    identity = {
        "worksheet_id": worksheet_id,
        "delivery_row_id": delivery_row_id,
        "source_export_row_id": source_export_row_id,
        "excel_row": excel_row,
        "excel_column": excel_column,
        "chunk_index": chunk_index,
        "coordinate": f"{get_column_letter(excel_column)}{excel_row}",
        "column_name": column_name,
        "value": value,
    }
    return CellRenderRecord(
        cell_id=deterministic_id("xlsx-cell", identity),
        lineage_reference_ids=lineage_reference_ids,
        **identity,
    )


def _normalize_zip(content: bytes) -> bytes:
    source_buffer = BytesIO(content)
    target_buffer = BytesIO()
    with ZipFile(source_buffer, "r") as source, ZipFile(
        target_buffer,
        "w",
        compression=ZIP_DEFLATED,
        compresslevel=9,
        allowZip64=True,
    ) as target:
        for name in sorted(source.namelist()):
            payload = source.read(name)
            if name == "docProps/core.xml":
                payload = re.sub(
                    rb"(<dcterms:modified[^>]*>)[^<]*(</dcterms:modified>)",
                    rb"\g<1>2000-01-01T00:00:00Z\g<2>",
                    payload,
                    count=1,
                )
            info = ZipInfo(filename=name, date_time=_ZIP_TIME)
            info.compress_type = ZIP_DEFLATED
            info.create_system = 3
            info.external_attr = 0o600 << 16
            target.writestr(
                info,
                payload,
                compress_type=ZIP_DEFLATED,
                compresslevel=9,
            )
    return target_buffer.getvalue()


def _render_workbook(
    *,
    worksheets: tuple[WorksheetRenderDefinition, ...],
    cells: tuple[CellRenderRecord, ...],
    style: WorkbookStyleDefinition,
) -> bytes:
    workbook = Workbook()
    workbook.remove(workbook.active)
    workbook.properties.creator = "Amazon Product Intelligence"
    workbook.properties.lastModifiedBy = "Amazon Product Intelligence"
    workbook.properties.title = "Amazon Product Analysis"
    workbook.properties.subject = "Auditable Operator Export Delivery"
    workbook.properties.description = (
        "Deterministic XLSX rendering of an Operator Export Snapshot"
    )
    workbook.properties.keywords = "operator export,xlsx,auditable,lineage"
    workbook.properties.category = "Operator Delivery"
    workbook.properties.created = _FIXED_DOCUMENT_TIME
    workbook.properties.modified = _FIXED_DOCUMENT_TIME
    workbook.calculation.calcMode = "manual"
    workbook.calculation.fullCalcOnLoad = False
    workbook.calculation.forceFullCalc = False

    thin_side = Side(style="thin", color=style.border_color)
    body_border = Border(
        left=thin_side, right=thin_side, top=thin_side, bottom=thin_side
    )
    title_fill = PatternFill(fill_type="solid", fgColor=style.title_fill)
    header_fill = PatternFill(fill_type="solid", fgColor=style.header_fill)
    title_font = Font(
        name=style.font_name,
        size=style.title_font_size,
        bold=True,
        color=style.title_font_color,
    )
    header_font = Font(
        name=style.font_name,
        size=style.header_font_size,
        bold=True,
        color=style.header_font_color,
    )
    body_font = Font(
        name=style.font_name,
        size=style.body_font_size,
        color=style.body_font_color,
    )
    title_alignment = Alignment(horizontal="left", vertical="center")
    header_alignment = Alignment(
        horizontal="center", vertical="center", wrap_text=True
    )
    body_alignment = Alignment(
        horizontal="left", vertical="top", wrap_text=style.wrap_text
    )
    cells_by_worksheet = {
        worksheet.worksheet_id: sorted(
            (item for item in cells if item.worksheet_id == worksheet.worksheet_id),
            key=lambda item: (item.excel_row, item.excel_column),
        )
        for worksheet in worksheets
    }
    for definition in worksheets:
        worksheet = workbook.create_sheet(definition.sheet_name)
        last_column = get_column_letter(len(definition.columns))
        worksheet.merge_cells(f"A1:{last_column}1")
        title_cell = worksheet["A1"]
        title_cell.value = definition.title
        title_cell.fill = title_fill
        title_cell.font = title_font
        title_cell.alignment = title_alignment
        title_cell.border = body_border
        worksheet.row_dimensions[1].height = 30.0
        for column_index, column in enumerate(definition.columns, start=1):
            header_cell = worksheet.cell(row=2, column=column_index, value=column)
            header_cell.fill = header_fill
            header_cell.font = header_font
            header_cell.alignment = header_alignment
            header_cell.border = body_border
            worksheet.column_dimensions[
                get_column_letter(column_index)
            ].width = definition.column_widths[column_index - 1]
        worksheet.row_dimensions[2].height = 30.0
        for record in cells_by_worksheet[definition.worksheet_id]:
            cell = worksheet.cell(
                row=record.excel_row,
                column=record.excel_column,
                value=record.value,
            )
            cell.font = body_font
            cell.alignment = body_alignment
            cell.border = body_border
            if type(record.value) is str:
                cell.number_format = "@"
            worksheet.row_dimensions[record.excel_row].height = 48.0
        worksheet.freeze_panes = definition.freeze_panes
        worksheet.auto_filter.ref = definition.auto_filter_range
        worksheet.sheet_view.showGridLines = True
        worksheet.sheet_properties.pageSetUpPr.fitToPage = True
        worksheet.page_setup.fitToWidth = 1
        worksheet.page_setup.fitToHeight = 0

    raw = BytesIO()
    workbook.save(raw)
    return _normalize_zip(raw.getvalue())


class XlsxDeliveryBuilderV0_1:
    """Render a serialized Operator Export Snapshot as a real XLSX workbook."""

    def build(self, request: XlsxDeliveryRequest) -> XlsxDeliverySnapshotV0_1:
        if not isinstance(request, XlsxDeliveryRequest):
            raise XlsxDeliveryValidationError(
                "request must be XlsxDeliveryRequest"
            )
        source = request.operator_export_snapshot
        source_snapshot_id = source["snapshot_id"]
        style = _style_definition()
        tables_by_id = {
            item["table_id"]: item for item in source["table_definitions"]
        }
        lineages_by_id = {
            item["export_lineage_id"]: item for item in source["lineage_index"]
        }

        worksheets: list[WorksheetRenderDefinition] = []
        cells: list[CellRenderRecord] = []
        delivery_lineages: list[DeliveryLineageReference] = []
        for ordinal, export_sheet in enumerate(source["sheet_definitions"], start=1):
            table = tables_by_id[export_sheet["table_id"]]
            delivery_layout = _DELIVERY_LAYOUT[ordinal - 1]
            if table["table_key"] != delivery_layout[0]:
                raise XlsxDeliveryValidationError("source export sheet order mismatch")
            columns = delivery_layout[3]
            source_rows = tuple(sorted(
                (
                    row for row in source["rows"]
                    if row["table_id"] == table["table_id"]
                    and row["sheet_id"] == export_sheet["sheet_id"]
                ),
                key=lambda item: item["source_output_row_id"],
            ))
            row_plans: list[tuple[Mapping[str, Any], tuple[tuple[Any, ...], ...]]] = []
            rendered_row_count = 0
            for source_row in source_rows:
                actual_columns, values = _delivery_values(
                    table["table_key"], source_row["values"]
                )
                if actual_columns != columns:
                    raise XlsxDeliveryValidationError("delivery column mapping mismatch")
                chunks = _chunked_delivery_values(
                    columns, values, style.max_cell_chars
                )
                row_plans.append((source_row, chunks))
                rendered_row_count += len(chunks)
            provisional = _worksheet_definition(
                ordinal=ordinal,
                sheet_name=delivery_layout[1],
                title=delivery_layout[2],
                source_table_id=table["table_id"],
                source_export_sheet_id=export_sheet["sheet_id"],
                columns=columns,
                rendered_row_count=rendered_row_count,
            )
            worksheet_lineages: list[DeliveryLineageReference] = []
            worksheet_cells: list[CellRenderRecord] = []
            delivery_row_ids: list[str] = []
            excel_row = 3
            for source_row, chunk_rows in row_plans:
                row_lineages = tuple(sorted(
                    (
                        _delivery_lineage(
                            worksheet_id=provisional.worksheet_id,
                            source_export_snapshot_id=source_snapshot_id,
                            source_export_lineage=lineages_by_id[lineage_id],
                        )
                        for lineage_id in source_row["lineage_reference_ids"]
                    ),
                    key=lambda item: item.delivery_lineage_id,
                ))
                row_lineage_ids = tuple(
                    item.delivery_lineage_id for item in row_lineages
                )
                worksheet_lineages.extend(row_lineages)
                for chunk_index, values in enumerate(chunk_rows):
                    row_content = {
                        "worksheet_id": provisional.worksheet_id,
                        "source_export_row_id": source_row["export_row_id"],
                        "chunk_index": chunk_index,
                        "excel_row": excel_row,
                    }
                    delivery_row_id = deterministic_id("xlsx-row", row_content)
                    delivery_row_ids.append(delivery_row_id)
                    for column_index, (column, value) in enumerate(
                        zip(columns, values, strict=True), start=1
                    ):
                        worksheet_cells.append(_cell(
                            worksheet_id=provisional.worksheet_id,
                            delivery_row_id=delivery_row_id,
                            source_export_row_id=source_row["export_row_id"],
                            excel_row=excel_row,
                            excel_column=column_index,
                            chunk_index=chunk_index,
                            column_name=column,
                            value=value,
                            lineage_reference_ids=row_lineage_ids,
                        ))
                    excel_row += 1
            final_worksheet = _worksheet_definition(
                ordinal=ordinal,
                sheet_name=delivery_layout[1],
                title=delivery_layout[2],
                source_table_id=table["table_id"],
                source_export_sheet_id=export_sheet["sheet_id"],
                columns=columns,
                rendered_row_count=rendered_row_count,
                source_export_row_ids=tuple(
                    item["export_row_id"] for item in source_rows
                ),
                delivery_row_ids=tuple(delivery_row_ids),
                lineage_reference_ids=tuple(sorted(
                    item.delivery_lineage_id for item in worksheet_lineages
                )),
            )
            worksheets.append(final_worksheet)
            cells.extend(worksheet_cells)
            delivery_lineages.extend(worksheet_lineages)

        worksheets_tuple = tuple(worksheets)
        cells_tuple = tuple(sorted(cells, key=lambda item: item.cell_id))
        lineage_tuple = tuple(sorted(
            delivery_lineages, key=lambda item: item.delivery_lineage_id
        ))
        workbook_bytes = _render_workbook(
            worksheets=worksheets_tuple,
            cells=cells_tuple,
            style=style,
        )
        metadata = {
            "format": "XLSX",
            "media_type": _MEDIA_TYPE,
            "ruleset_version": XLSX_DELIVERY_RULESET_VERSION,
            "source_export_snapshot_id": source_snapshot_id,
            "renderer": "openpyxl",
            "renderer_version": openpyxl.__version__,
            "formula_escape": "LEADING_APOSTROPHE",
            "max_cell_chars": style.max_cell_chars,
            "worksheet_count": len(worksheets_tuple),
            "source_export_row_count": sum(
                len(item.source_export_row_ids) for item in worksheets_tuple
            ),
            "rendered_row_count": sum(
                len(item.delivery_row_ids) for item in worksheets_tuple
            ),
            "cell_count": len(cells_tuple),
        }
        workbook_content = {
            "filename": _WORKBOOK_FILENAME,
            "media_type": _MEDIA_TYPE,
            "content_base64": base64.b64encode(workbook_bytes).decode("ascii"),
            "content_sha256": sha256(workbook_bytes).hexdigest(),
            "size_bytes": len(workbook_bytes),
            "worksheet_ids": tuple(item.worksheet_id for item in worksheets_tuple),
            "metadata": metadata,
        }
        workbook = WorkbookDeliveryRecord(
            workbook_delivery_id=deterministic_id(
                "xlsx-workbook", workbook_content
            ),
            **workbook_content,
        )
        diagnostic_content = {
            "code": "XLSX_RENDERING_PRESENTATION_ONLY",
            "severity": "INFO",
            "message": (
                "Workbook cells reproduce Operator Export values; oversized text is "
                "split into deterministic continuation rows and formulas are escaped."
            ),
            "source_export_snapshot_id": source_snapshot_id,
        }
        diagnostics = (
            DeliveryDiagnostic(
                diagnostic_id=deterministic_id(
                    "xlsx-diagnostic", diagnostic_content
                ),
                **diagnostic_content,
            ),
        )
        coverage = coverage_from_delivery(
            worksheets=worksheets_tuple,
            cells=cells_tuple,
            lineage=lineage_tuple,
            diagnostics=diagnostics,
        )
        content = {
            "ruleset_version": XLSX_DELIVERY_RULESET_VERSION,
            "source_export_snapshot_id": source_snapshot_id,
            "source_bundle_fingerprints": tuple(
                source["source_bundle_fingerprints"]
            ),
            "style": style,
            "worksheet_definitions": worksheets_tuple,
            "cells": cells_tuple,
            "workbook": workbook,
            "coverage": coverage,
            "diagnostics": diagnostics,
            "lineage_index": lineage_tuple,
        }
        snapshot = XlsxDeliverySnapshotV0_1(
            snapshot_id=deterministic_id("xlsx-delivery-snapshot", content),
            **content,
        )
        return snapshot.validate_against_export_snapshot(
            request.operator_export_snapshot
        )


__all__ = ("XlsxDeliveryBuilderV0_1",)
